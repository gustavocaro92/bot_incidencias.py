import logging, os, threading, base64, tempfile, re
from http.server import HTTPServer, BaseHTTPRequestHandler
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, ConversationHandler, CallbackQueryHandler, filters, ContextTypes, PicklePersistence
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime

try:
    from PIL import Image as PILImage
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import pyotp
    PYOTP_OK = True
except ImportError:
    PYOTP_OK = False

try:
    from spellchecker import SpellChecker as _SpellChecker
    _spell_es = _SpellChecker(language="es")
    # Vocabulario técnico + nombres propios de Ecuador (no marcar como error)
    _spell_es.word_frequency.load_words([
        # Telecomunicaciones / FO
        "telconet","otdr","fibra","optica","óptica","sdh","nodo","tramo",
        "hilo","empalme","atenuacion","atenuación","splitter","backup","main",
        "cuadrilla","enlace","resolución","resolucion","incidencia","reposicion",
        "reposición","empalmador","fusion","fusión","datacenter","olt","onu","pon",
        "adss","opgw","jdf","mts","dbm","hospedaje","cordón","cordon","manga",
        "acometida","armario","derivación","derivacion","multipunto","enlazado",
        "anillo","backbone","uplink","downlink","patch","cord",
        # Provincias de Ecuador
        "esmeraldas","manabí","manabi","guayas","pichincha","imbabura",
        "carchi","sucumbíos","sucumbios","orellana","napo","pastaza",
        "tungurahua","bolívar","bolivar","chimborazo","cañar","canar",
        "azuay","loja","zamora","morona","galápagos","galapagos","cotopaxi",
        # Cantones y ciudades
        "pedernales","atacames","muisne","quinindé","quininde","rioverde",
        "esmeraldas","tachina","rocafuerte","portoviejo","manta","jipijapa",
        "montecristi","chone","calceta","bahia","caráquez","guayaquil",
        "durán","duran","milagro","daule","samborondon","quito","sangolqui",
        "cayambe","ibarra","tulcán","tulcan","ambato","riobamba","latacunga",
        "cuenca","machala","mera","puyo","tena","macas","salinas","playas",
        "naranjal","naranjito","babahoyo","ventanas","quevedo","pasaje",
        "huaquillas","arenillas","balsas","zaruma","naranjito","cotacachi",
        "otavalo","guaranda","riobamba","alausí","alausi","macará","macara",
        # Operadoras
        "claro","movistar","cnt","otecel","conecel",
    ])
    SPELL_OK = True
except ImportError:
    SPELL_OK = False

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)
BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
PORT      = int(os.environ.get("PORT", 8080))
SALIDA    = "/tmp/reportes"
FOTOS_DIR = "/tmp/reportes/fotos"
os.makedirs(SALIDA,    exist_ok=True)
os.makedirs(FOTOS_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# IDs DE TELEGRAM DE LOS COORDINADORES
# A todos estos les llega el Excel cada vez que un técnico genera un reporte
# ══════════════════════════════════════════════════════════════════════════════
COORDINADORES_TG_IDS = [
    5380409177,
    5917376513,
]

# ══════════════════════════════════════════════════════════════════════════════
# USUARIOS AUTORIZADOS
# ─────────────────────────────────────────────────────────────────────────────
# Cada usuario pone en Telegram:
#   email: sucorreo@telconet.ec
#   totp:  SuContraseñaPersonal
#
# Para AGREGAR un usuario nuevo:
#   1. Agrega una linea con su correo, su contraseña y su nombre completo
#   2. Comunica al usuario su contraseña
#   3. Reinicia el bot
# ══════════════════════════════════════════════════════════════════════════════
# Secret TOTP para FortiToken Mobile — se lee desde variable de entorno Railway
TOTP_SECRET_GLOBAL = os.environ.get("TOTP_SECRET", "")

# Sesiones activas: {telegram_user_id: {"email": ..., "nombre": ...}}
_SESIONES: dict = {}


_LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAA7AQIDASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAUIBAYHAwkCAf/EADkQAAEEAgEDAQYEBAMJAAAAAAEAAgMEBREGBxIhMQgTIkFRYRQVcYEWMlLRM5GUFyNCVmJjdJKh/8QAGQEBAAMBAQAAAAAAAAAAAAAAAAEDBQIE/8QAKBEAAgIBAgcAAAcAAAAAAAAAAAECEQMEIQUSEyIxQVFhcaHR4fDx/9oADAMBAAIRAxEAPwC5aIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIvG9brUacty5PHBXiaXSSSO01o+pK4xzr2leDYFk0WHZaz1pnge5b7uHf3e75foCrceHJlfYrKcufHhVzdHbV+BLEZTEJGGQDZaHDYH6KjPOfaU59yDvix80WCqH0ZTGpCPvIdn/LS1DgvV/l3F7GTuYu9u5kA1s1mxGJpB2knwX7Hz+i9seGzreSs8EuKwT2i2vp9GEVKOmHXfJ1eUOznOMxyPKQsYWw1Kr2NhJPzc3bR48LrlP2m8Dkr0NDD8Rz123O4MjiHuw5zj6AacVXk4fmi9lZZj4ngmu50/n+He0WHhrVq7i4LVzHyY+eRvc+tI9rnR/YlvhZi8TVGgnasIiKCQiIgCIiAIiIAiIgIvkXIsDxytHZz2Yo4yGR3ayS1M2Nrj9ASoL/an031v+OePa/8APj/utjzOGw+ahZDmMVRyMTD3NZarsla0/UBwKq1x7j2Ak9sWXEPweMfjRXnP4R1Vhh2I3EHs12//ABejDjhNO72Vnmz5Z42qrd0WKk6i8Cjow3pOYYRtWcubFMbrOx5adEA786WOeqnTYevOuPf6+P8Auq9e3HisXh6vFKmJx1PHwbsH3VaBsTN/Bs6aAF+uD3cZFNhXZ3kPRSTFNbEbcIoM/Fe77RtpJZ/ifU79VdHTReNTV7lEtXNZHj22os1Fy3jEud/IY8/jn5UAH8GLDTLogEfDvfkEFTSqFczGHwntkZDI5PIVaNBkkX++leGxhvuY9efTWl0D2heZ2eVcBlp9K+Qw5WaJ/vMrHi5u+dtbWtjt+LXdrevkuJaZ80UvaLI6tOMm/Kb2OuZTnPC8Xa/CZLluCqWAdGKa/Exw/UF2x+6mMdfo5Kqy3jrta5XeNtlglbIxw+xBIKrH0mZ7OJ4PSGdGJblmxgXhlC9s/vdfF4Pj/wBfCkOTO4lN015E3oA1rMgCwZD8C2dsjoQfi7DJ4Pje+1Hp1dK/za2IjqXXM6f4J7ndcvzfhuItGplOV4OlYaQHQz342PG/q0u2FJ4jLYrMVhaxGTpZCA+BLVnbKz/NpIVW+iQ9n5/D42csbj259rz+N/OS4Sd+/OidN16+nn6qKw35JJ7RGJd0RivNx7ZAMk6EPFYs2O7+bz2a3vf7Lp6VW1uq9tbHK1bpPZ36T3LQ1+fcIsZT8rh5bhH3i/sFcXY+8u/p1ve/spfOZnE4Og7IZjI1aFRv801iQMYP3Ko7034DT6i9XOTYS3clpzMjsT1p4vIjlbIACR8x5PjYUv1W5Rz7jXCbfTTqDSktlpBx2VBLvfMB9C4/zDwfPr8l09HHmUYs4WulyOUo7ev5LZXuoXBaMVaW5y7C12WohNA6S4xolYfRzdnyPHqvKn1L6fXLUdWpzTAzzyuDY4470Zc4n0AG1SnqCdN6bH3lSMjBV/itt3C343eXj+n6ruXTG9hW8xqjP57o5bgce2KPEUY2WDKf5O0lg+eknpIxje5OPWSnLlpLwdjs9TOnta1LVsc1wMU8LyySN96MOY4HRBG/BC9cZ1E4Hk78OPx3MMJbtzuDIoYrrHPe4+gAB8lce9snjfHsf0xfkaGBxdS7JejL7EFRjJXbJ3twGzv5ravZo41xt/S3A5d2AxLsi0OcLZpxmYODjo9+t7/dVPFjWLqb/C5ZsjzdPb6ddljjlidFKxr2OGnNcNghU29pfo4/il2TknH67n4Oy7csTQT+Ecfl8/g+5PqdK5ax8lSq5KhNQvQRz1p2Fkkb2ghwP2KjS6mWnna8exrNJHUw5X59M+WdhpjL2/RfjHxF7GtHz8ldg9pnpLa4BmvxmPjkmwd1+q0mtmM7/wAN339PPz2to6Cezzk+QxV83ytsuNxJ05kBBbPYHqPBHwtP19Stvr4kuo3sYC0+Zvppd393NA6X9OOR89yYpYSr214yBPbkGooR9SfmfXwFc/pJ0p4308og0oRbyj26nvSjb3fUN/pb9h+63Dj+FxfH8VDisNRhpU4RpkUTdAff7n7lZ6ydVrp5+1bRNnR8Ox6fue8vv7BEReE0QiIgCIiAIiIAiIgCIiALAZhcOzK/mzMXSbkNFv4oQNEuj6ju1tZ6JZDVmBl8Jh8x7v8ANsVSv+732fiIGydu/XWx4Ud/BPDv+VcJ/oY/7LYEUqTXshxi/KIa7xTjF6y6zc49irE7wA6SWoxzjoaGyR9F7Ynj+CxEz5sVhsfRkkb2vfXrtjLh9CQPIUmicz8WOVXdEFk+G8SydhtjIcaxFqZru8SSVGF2/rvXlS1GnToVxXo1YKsIOxHDGGNH7DwvdEcm9iVFLdIg8tw/imWl97k+N4m3J3d/fLUY5xd9SdbKksbjcdjIfc42hVpxH1ZBE2MH9gFlInM2qCik7o1bhfFxg5pZZ6WJdZcX7uwRds8gc7enfCPH7/Jf2/xq3k80bGYlx2Txpf8ADRtVO9sTderSSQXb+eh4W0IlvyKVUapb4HgLebguWcXRlq164hhrOhBYzRPo3WtefRYR4BTi5PBmKVHB1oID8FRmPYAf+svA33D1Hhbwinml9I5Y/DV+c8auckkqRMuVoqcXcZop64nbITrR7XeNj5H5KQ4dh5MFgYcZJMyYxOcQ5jA0aJJHgeimEUW6oml5CIigk5b7RlXljOKMznFbvbJjn+9sU3wNmZOz5fA4EEg+f2VSR1r6lUM/LlIOSWGyTO73xaBhJP8A2z8I/YL6DTRsmhfFK0OY9pa5p9CD4IVEvaj6bP4Ryx16jCRhr7jJXcANMcTt0Y/TYWrw/JCSeOSV+jF4liyRkssG697/AKm+cI9rHIMMdblnHIrfjRsUH+7d+pY7YP7EKyPT7mGJ5xx6POYZtltZ7i3U8RYQ4eCPofP0Kol0G6c3+oXLY6kTHMx8JD7ljR7WM35Gx/xHRAX0CwWKo4TEVsVjoGQVa0YjYxo14A9T9yq9djw46UV3FvDsmfJbm+1GaiIs41QiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiALVOqvCcfz7h1vAXu2N0je6Cft2YpB5BHz1sDevULa0XUZOLtHMoqacX4NM6P8AT/GdO+IwYal2y2SA63Z7dGaT5n9N70PktzRElJydvyIQUIqMfAREXJ0EREAREQBERAf/2Q=="

def _get_logo_path():
    import base64 as _b64, tempfile as _tmp
    try:
        data = _b64.b64decode(_LOGO_B64)
        t = _tmp.NamedTemporaryFile(suffix=".png", delete=False, dir="/tmp")
        t.write(data); t.flush(); t.close()
        return t.name
    except Exception as e:
        logger.warning(f"Logo no disponible: {e}")
        return None

_OTDR_IMG_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAD5ASwDASIAAhEBAxEB/8QAHQAAAAcBAQEAAAAAAAAAAAAAAAMEBQYHCAIBCf/EAEUQAAIBAwIDBQUGAwUGBwEBAAECAwAEEQUhBhIxBxNBUXEiMmGBkQgUFUKhsSNSwTM1c4LRJENicpKyFhclNFOi8MLh/8QAGgEAAQUBAAAAAAAAAAAAAAAAAgABAwQFBv/EACkRAAICAgICAQQCAwEBAAAAAAABAgMEEQUSITFBEyIyUSNhFEKBFTP/2gAMAwEAAhEDEQA/ANh0KFCkIFChQpCBQoV4Tg4pC2e0KR3OqabbPyXOoWsLeTyqD9M11balp9ySLa9t5iP5JAaLq/0D2X7FVCvFZWGVIPpXuabTQ+0ChXorknBGdtsn4Uw57QrgTRE8okQsOoB3HrXYOdgD9MUhAoUAfT5GhSEChQoUhAoUKFIQKFChSEChQoUhAoUKFIQKFChSEChQoUhAoUKFIQKFChSECk8g9s0pFJn9804LZ5QoUKcQpoUKFCEChQoUhAZlVeZiAB1JOBWc+2HtouWv7jSOGrgRWsTd3JcKPakPiFPgKtHt51mbROzbUZ7Z+S4nIgjbOMcxwT8hmsWXxc55j0JxWtx2IrPvkY/I5jg+kR4v+JLqZ2mnuZZHY5JdyTn503HiG8EgeG5mjceIYj9sUxzu2SCTtRKtvW06Y/ox/r2fsm+n8f8AF1jvZ8RahF8O+Yj6E1OeDu3nizSVaTV2j1S3UgOsnsMB5hh47VSqOR0NGyySNpk6oNwyt+9Rzxq5/kh45M4v2bC0z7QHZ/dnSo3vpLeXULhbfupU9qF2A5S//CTtnes8dq3bB2injHVLmDXLvT9DW+uLOzit5Qqy9y3IVJGTnOCTtnNUtxIkYsxdifu5Y29kA7k+tMtnNLLABNK7lWLZdicsep9TtXP5NH056R02Pc7K02SO+4o4hu9Ul1OfWb43kr87yrMQS3nT3pXan2h6YF+5cYaxGB4G4LD6GoJzNznc/Oj4GBYBunjVXRYTNR9gfb3xLNr7WPGN9JqVlJFnvCgDwkeIx7w8T41rO0uYbu2iubdxJDKoaN1OQwIyCDXzX4QuTYa3ZcgG8LF8+RJxn6VfU32ibfgvswsNF0yKO84nXnjWKcExQIG2Zzkc23QAim0Oa0odBk9Kxzwp9rjiKTU4Ite4d0qW1LhZpLN5I3AzuQrFgT861ZwXxPpHFvD9vrmi3Sz2s65BAwVPirDwNNoQ9UKrHtT7X9J4F12DSbjTrm9mkh72Qxyhe7GSBt5nFM+l/aM4IuFIu4NStX/lMav+xqxHEtkuyXgrTy6oy6tlzZr3FV5pfbN2dX5ATiCOFj4TRsmPnjFSew4x4VvgPunEOmyk9ALhQf1NBKmyPtBxvrl/sPdCiIb2zm/sbqGTPTlcHNHBlO2d/Ko3FokU4v0z2hXnNXtMEChQoUhAoV7XlIQKFChSEChQoUhApO/vn1pSKTSe+adAyPKFChTiFNChQoQgUKFDrtSEUf8Aaw1Hl0rR9IVj/FledxnwUYH6k1mfUkCgk433q4/tUatntFgs5CRHb2Sco/5ixz/+8qpniG8it9MmuUPMwQcp8jnb966fBSrxkzl87dmR1Ga9lgiPtuATSMzwk+zIv1pgkuZZ5CzsSWO9Fs6BsMcH41BLkfPosrjdr2SZJkPRgfSj4bhkJ5GxnqPOorHLj3HI9DRi3Ew3Ez/WnjyMflEc+Ll8MkjpZSTpLPp1tOVYNiReZTg+I8qSXGlaTNI7raGAueYrEcKCfIdBTWl7cr0lyPiM0dHqU4O6q4+O1P8A5NE/MkD/AIuVWtRZ3Lw5ZyNmO4nTb82DvRa8MyI+UvEdcdCCN6VJqYx7cRHoaNXU4B1EgHoP9aZV4k/QXfNh7OorC/h1BruIIwWFVTlfByAfP4moleaPrhmeaaynYyEsWXcHepol9asAe+A9QaOiniYZjlU+Gxpng48vTCXI5MPyRBdOtLqO5PewSoEQtjkIJNbW+xbdXdrwXq9nd5WK3nW4idiQCrrv1GwBXz8azUk0w6SHHkd/6U8W/FnEVrpz6fbazewWrxmJoUkIUqfA+Y3NC+Li/wAZEi5eWvuQ5dqXETcT8earq/ec0UkxSDB2Ea+yB+mfnUXJAOKJ58DBztsB4UA3xrYpgq4KJiXWuc3L9ihSMV2kjLujFT8DikoY0Yh3FG0n7REm16F9vqmp25HcX1xHg59mQinjTuOuL7FgbbiDUEx4d8SPoajmaA60Lpg/cUH9aa+Sx7Dtp7Q7RQF1xpsdO9iRv3FSPTftD8ZQEferXTboePNGVJ+hqlwcV7neoZYdMvcSaObfH1JmitN+0lMCPv8Aw9G3+FLj9xUhsvtFcMS8ou9O1C2JIzy8rDFZVztXqnbHh5VDLjMeXwWI8pejZFl24cAXIHPqE9ufKSE/0zT5a9p3AlyoMXElmMjOGYr+9Yc3xjO1dx7VXlw9Xw2TQ5a1e0b80/iDRNQx9x1S0uMjI7uZSf3pzBBGxr582t7dWkiy2s8kMikEMjcpyKvLsV7ZLuC7g0Xii4M0D+zHcH3lPhnzFU8ji3XHtDyXsflI2S1PwaUoUXbTw3EKzwyB43GVYHINGVkNaNhPYKTv7xpRSd/fNOhjyhQoU4wpoUKFCEChtnB6UD0NV/xTxPcfiklnY3DRhCASpxvUdligS01Ox6RnT7VcxfthvASeVLaFQM9PZz/Wqa4oueXSxCDjncZ9BWmO1/hBOJbafVXU/i8cfsygnMnKBgN5+QrLPFJbvo4HXBT3h4hvKtvHzozxesfZjZXHTqye8vQRw1ZC9vsPgoqkkHxpWeEoJrX77PxJYacstzJDbRXMUzd4EA5m5kBwMkjceFG8Gx8guZmwF2GfLz/pVuaB2N8Q8S8EaBr9pqNtas9nLI8M8fMG7yV3ydx+Vh59KrVKEpam/BbW/gpscF6lL/d+t8O35/KsGorGx+UgSiZuDeNYASOH9QmQbc9uBMp9ChNWtqXYfx8sYeC00q4zh+9hDRNgeRIA3qMjs6400dD974RvZHVjmW0myx9OU5NXHi0yX2yGcpform4TUNPcx39rcWz/AMk8TRk/9QFcpektgYYgZODnFaT7ENE4rk0nWJdUg1NuW5jS2tb+RsqpznAb5VKNQ4Qsr1yNZ4Ts+7ALM81vGwGOu4GRXPZWdLHv+l13/Zo1YkLa+/bTMlJeKR7QbPwGa7W6jO3tD1FS+VNI1BnuI+DdEliZiVW0vHhlUZwMjmwD06CkkumcKlik+jcR6fIo5ysV1HOCu+4DLnGx8a2I4tsoppGbOcIy0MAnjO3PkdK79jwwP0pxn0HhdmbueKru1PPgLd6a/X1Qnb5VJ4Ow7ji60uDU9JuNL1KzmiWWJluDE7KdweVwCPnQyrsqW5IdRUyFRMVIKMwxv72aUC8ulG0pPrvTrqHZd2kad7T8LahKB425Euf+kmmDU9O4j0s8uo6LqVrjr31q64+oplkTj6YEsSL+BYmo3QO/Iwz4ij01Nx70QPoajYvpFBLDGDvkUeuoJgZqWGbb+yF4Fb9okSalGd3R1/ajV1G18ZeX1FR5L+B1yWIx5ijo7q2Yf2yfWp48lYvbIZcVVIkcdzDIAUnQjy5qOWTm2GaYImikGVZGHTqDRqhV918ejGrEOUa9ld8V+mPgPxoxW9n0pmtknmnjiilk53YKoB6knarG1zsp4w0k37QT2+ox2kkUcfdqS9y0gB/hrjcDxPwNTx5OD9lefG2/6kSD+p+VGA0ZZ6DxRc6sNJj0aQ3jMVCNGUJYflBbA5uu3wpJqn33Sr6Ww1GxeC5hYLIjMOZSRkbZ/XpU8c2p/JXlg3R+BQGB8DXatTfFewsM4kXffKmjku7dukyDfzqRXwfpkTotXtClm9oV1FO0UqSps6EFT8aTF8+viPKi1YtKqg7kgUb/ALIttGhuy7tIu9GuYLe8ZnsXZeZS26qV+laRtZo7i3jnicPHIoZGHiD0rEsCMIBtg8oA+laV7AuIvxbg+GwnkJurMcm/ingf6fSud5ChKXeJ0fGZMpL6ciy8Umf3z60oBpO/vmstGueUKFCnGFNe7V5QG5oQhFrt79w0q4udgUQkZ8wKoW0uLifUDPIXJkkLfLNWJ2x6jejTE07TpOR5T7bDqB5VAeHbW/T2Jpu8ztvVG+1dtGzhVdYdiUyp3sJPLnKYrF3atbG17QNXgKFAlwSAT575FbciQrYDPVv2rIv2jbdYe0y4kUAd/DHI3w2I/wD5osSbVmkQ50FKHYi2mwsvDcqIA0twGVB5s2FX55IrbYtbbQ+D9K0tNQeye0s40jETqpmCIARgg5Hn8KyT2d6a2p8WcLaUEBEmo27SgD8qnvG/7K1rxZqEltd21tHoz3ckjBoWUqehy646ggDrWhOWjJrS+ST6YyNptv3MjNCYl5SwzzDA+lHuiMuGijYf8Sg/0om0ubSVAkMkYKgDlzg+mPPwpQCNx40MX8jyXkSNp9g5JezjXPinsn6gUnutFsriB4cS8jgh0aQlWBGCCPGnLavaTSb2xJtLRWWqdhvZ3fZ/9EhgdurRAp/2kVGrz7OPDnKW0/U9Rs3HuslwSRt4cwOKvHG9e+FWYZdkVpMi+mt7M2at9nvWoYGWw4ynaNwVK3MSP1+JIO+4q2NF0250zSbXTo7OUw2sKQhkXOeVQPD9qmd9Z219AYLqESRnqM4pvXhvT4gFt5b23HNzYjuGA+lDbkztWphJaGYXCjmVxIvKNwVNdGaKUAGRWA6BjTnPomonmW21uZEdy7LLGsm+AABnoOvSnBdOtmRBLbRSMqgOwTHMcb1BvYWyE6lw9w5qYI1DRdMus+Mtqj7/ADFRvUOyHs2v1Ik4Vs4c7lrctEc/5SKtWTRbBt+5dDn8rEUSdBiyTHdzLnoDuBT9tCKL1L7O3AM4b7o+q2LHoUuQ4X5Op/eovqX2ZoiGOm8X3CnwW5tQw+qsK0s2i3a+5dI//MMUW+l6lEuQkUoP8p3/AFFEpBGTb/7OfGUHtWWr6TeEDAUu8f7g/vUavuyLtIsZnjGircSR78ttewyN/wBPOD+lbRkt7+I+3Zybb7YqB2nD3EouYbOY28FpHeNcC75naWTJJCFcYAyQOvhUU7ZxfhFvGxqrU3OWtGVJtE7QNBuo55eHdetp4GEqN+HyNysDkEHlI6gU/aZ2z8e6XdpDPfx3LWy8ndXVuMrgHY9CpyzHz3rSNvpPFlsbWztNchns1EfevKW7/P52zjcbnAJ8B8c557SbnhvXOPdbvLu11Xl+/Sp30ad4mEYoAM7AYAP9as4cXfLUvBXzKq8ZrrLew6fty1q5utOmvtIsn/D7830awsyK0hBByMnxKkeXLjpUht+27hW6sY7bU+DIyYdPNvEVSN8SlVHN7Q2GV5s9ck+eao+W3gDMqqQFYhSBgkeGa8Fqj+6WHo5q++Mn7iUXkRftF/y8c9jutjvr/SJIpmcc2ITAOVV9lRyHA6FfnVZ3PdzTvPBCkUcjMyRqxYIM5ABO5FRebQynDMmt/i1vzC4EItHGZmGMl/ToKcuH5pDpsXOxz8CaqSrlVLrL2FqMlvQ7zMisO7LYIyR5Hy+NHaSvf6nbp5tuP1pC+WORTlwkjNrsIAzgH9jW/jSbq2zmMiCVzSLEEZEYH/DtVm9hc0tvxbZRRjCzxyBx4Yxn/Sq9KgQMSN8ACrU7C7bvuLC+By2lqR8/dqjmy+xl7BT+qi8xRD+8fWjwPGiH94+tYSOiPKFChTjCmmvinV4tE0eW+kEjMowiopYknp0p0qtO2CYXLwWkftKq+1yuQQc+VQWycY+Czj1/UnrRDdS1651fUC8+zH3QMgj670/aOgCrsM43pq0XTjGFJywz+YU9KBBuoA9KyPLe2b6SSSQpvZ0ijwWwF/Ssdds+rx692l3fcsrLCVttumVzn9Sa0L2scWw8PcKX96Z1WfkKQL/M5Gw/aslaPzzamk0zkuSXZvEnqT6mr+DHcnJ/Bl8hYoR6ouv7Nlkb/tdtJBGDHp9hcTtkdHfCJ+hetLWEPf8AES3C3RnjRHbu+UEQnpjPX/8AGqU+yPp5abivXGiGM29iu2ASis7Y+ci/Srv4aS2kubu7t7JrQMQhRkCkkEknA9avzW/ZmxSaHa5tredMSwRv47ik6Wcsf/trqSMDorjvF9N+lLSaA2pLwgG/I13N3ewXUMUtp3meYs9vucbblcfoKV293bStyrcIH8UY4YeoNGNCv3r7yoxJgLnPQb/617LDBMuJ4UlHk6g02xzvqwAIr0A4pEtgsfM1tPPCc5C8xZfTB6fKibu41G1txzRxTK0iIGiJVhk7+yc+H/EKYZDmOleHOaJivbeR+7MvJJ/8ci8jfQ9fWjVdWmaMZyoBJxtvSHaOh60EJGdzuc0MDmwRj4mgp+FIY6JPmaAxXJO9eikEkdV6MeVeLucV1gjwNIc9Ub9K4vyFtZHYc3KpIyuf0oyMikfEdwbfR5pF2IGc4zQS21pDoqi/LWEdze3cNxHJAjTSFyQBjcADwOB0261WvCGrwWfDVlbyxxFzErzcyAkuwDNn5k1Ku0K0uND4K4i1B5rtri+tBA+JMiV5ZcZC/lwM4+e9U+2pm1hV7mGe1TbDSwsB8NwMVl5GHkxq3Tt7ZT5e52OMYFgyzcIahNi90zSpW6ZaBM/Wkc/CPAN65B0VIs+MUzKD8gcVCfxHS7pgshs5id8lhmnOxv8Au1CRAhc7ANn+vSs+ORyVPjcjEUpohPbPomj8O3cFvpEtw8U8IkKTPzsh5iMZ6+FNemKYrOFD1Ar3tKuPxPidIS5bCquc538f3oxcBlA6DFdbTOy2EZWezXhJqvY92sHNp4cqfbJwfhTtwFb82tliCQq0b907jRrdcHPcjm9cU48A27IZ5ioPNsCOorqILrSjm5S7XNktCBpYIv5jk+gNXR9n+09jVL8L7zrHzefj/pVM25BuZJm2EUW3zz/pWkuyfTPwvgqzTkCST5mfbf2umflisjNnqOjY4+G7NktB2pM3vH1pQOlJ294+tZKNoFChQpxjzVryOw0+e6kbHdrkfE+FUuZ59V1OS6nPNljjx2qU9o2uC7m/D7c+xGdyPzGmPRISANgMneqF9m3pGti1uC2OltAiwjbwpBqUioCeYYUU6TlUiIBxVc9rvFMHDfCN7ejDXDgR2656uf8A9+lQRh2ekWlLotv0UB2/cSnXuJ/w6JwbaxYq2DkNJnf6Hb61EOGYgZpJGBYKMfWm27d3uCZX55ScyOfzMdyfqaeNLk+7aHdzkH3GIIHX2dq2K6vpx0YN131JuS9Gs/sy6dLb9jsV1HyRXGp3dxeCZxsAXCKceXKgqz9AWY2XfXDK00rFnKrgHfGw9AKjnDFhbcO9lGjaXdxPyW2mwxOoQkluQE5xv72d6kelXdlFp1tD3qxsIhlG2I+tNL2MvQ44r2uFljfBjlVgemN8mu8t0IFM/Q2jzI86GaGAdyBXhIBxQjaPRRckMUjxtICTG/OB4ZxijBXDbttSHR7NDDNGUeKN0/ldcim7TbIiS5nhNxZHveVBzAqQABkA5AGc05jOKCnbFIfY3ajLqNtayuEiuCBhSnsNk7AkdD9aUx6hbErE8hicjZZU5GPw3/ptXd1bx3Jj7wH+G6yLg43ByPltRk0ayoUdVZT1DDIPypDAZgEL4JAGdt8/SuNOu4L2yjurcsY5BkcylSPUHxpBqlq9tZXFzp6T97GjOsUT4D4GcYOQM/CjOFlmOgWklwWad1MkhbGcsSTnG3iPpSCXodUIzXfhXC+ldA7dMU2xHqrTLxcLh7JYbe47hychuUNg/OnvmwM7iotxw8Xf26GVubIHKsgBO/y2pL+x/RTX2lpb2x4K0+zaRHudRvB3bR++e7VzvnwBIOPjVG8U8RXF/wB1bfhMelRNAsbIHZzIy4BY5G3QnbzNWh9onUrROK9DsNX+8SQW1pJcd3nlZWkOF907Z5T5dTVW8R/dJtWiilW+t/u0pSVZ5OchV3BXm6Z28a6HBxlGuOzOyLE5+BbecTWU/DNtYDSwbpeYXMrj2CntcoAxgY8+pr3h+XhKPhW6OqMq6pA4FvCoPeTZC+1ke6ASfpTfra6fcWCiyvO9l71FKSQIGIOw9pRv1py1XS0g0x7gw6dJ/CY/wJmA6Z2BGCR69RV26qMo6IE0yDzhJuImCZ5EJYZOSN+lO9nF31zFH4M4X60yaZ7eozydeo39alPC8Qm1mBWBIVuf4ZFYaTnbpfsmtfWpk21QL3LIMlQAv6U6cHQiOwYjGGamrUCChUeVP+iL3emR8oAJGa37ftgkjmaX2m2PfDti+para6egBa6lCD6j+ma1bZwrBaxQxjCIgVR5ADFUD2IaZ9745t5SoZLKFpPgGwAD9SfpWhcHrXO5k256OnwK0o9jwCkzD2j60ppM3vH1qki+wUKFCnGKQhZrm8Zs5J8TT9pq8i/EU2aZb8qhgPaNOyL3cXXfxrFTb8s6PS1oKvZjyEHGPOswfaK1uTUeLYNIWRe5tIwzBeneN5/L960nqLERNk7Y3rGnGN4b7i3Vr4tz81xIVJ32BIH7CtPjYdrNmZytrrpUV7ZHJzzSMal2k2cczaXpjZP3q7hV8dOXmy3y5VNRKMF5VGMksBVsdjekxaz2t6HYTp3sEKSTyoOhULyY/wDuavyknL+jJj4iWovHUMxhis7q5iDj+PHO7EFTvkDLD5eVP0vHsyOBaatDeYAPcyRKM/t8akVt2dcM3F9cM9haRd0wRXspyGRgW9lvly+e5pLfdjulSsXtNUuoiSSBIquB+1X64cfNeZNMpyszIv7YpoQxcbQR3bXdzZ27c0YUCJWjx1JJ5Cd/Df4U5W/Heny2zvCt3C4Yqj/eCyE4yNnPSo7q3ZFqtpBJPbavamKNSx5w0ewG5OPhTLBwnrCxALKZIwoZe5ukYHOfytg48qeeBjtbrt2BHPvi/wCStlqadxZBLBEJdTCTcgZ82xKZwfHypdHxJFJe29tFLa3TTqx5lJRU5RnDZ6E/CqXlPFejuxFjdm3C45pbXII9cYIpvbiW/N2sl3ETg5aNDyqf8pGKF8RbJbhp/wDQnzGOve0aLi1BjkPasR5xOGH6kV5batp0ylopmxzMuSh6jruBVAaHxILe4aS8kuijAACLAKf9JGf/APaebbjb7mWGnahNAF5WEco5FJJPMdx8KrT47Ih7gyxXnUWepIvFbm2cDu543z5GjUXmGVwQPI1Sacc3N9yJdx87LMrryIDzEEEA8p6U/RceWMdoJrmxEQBxzRs6b1WlTZD8olmNsJPSZZwYHBrrwzg/Sq/0zi0S30veSXccDoj268gfHvA5Prj6U7Him1W0knhvorhlDYjeMqWI8AenXaonFoNLySK/k7qwuZhtyRMc/KjbVFit4oV/3aBOnkMUxNqk+oaT/s6WbvKFJj+8YwOYEg/HAru84kltLiGG60m4M86s0MELq8knKfAZGfZycf1oRyRRjxrs9aY7HiKyuNQkszFc23dRq7yXMRiTJ/KCepHj4Ut/F9LaQRrqNqWzjAlX/Whb8jxTflC87KTg7CoJxLPZXGvvGY4mmRtjICOX5jpvUxhvreQFopkkC5JKnOMAHfy6iofJqK97m5s51jf35O7zzHrzDyFFFdnpCltezPPaoeFb3ju+/F76NLq35IArTsgChQSQeUqRzE9QTtUeXgzQNUnC22vtyyYRW76OcFeg5hkHPTP7VEOPbh9R4v1K+jkjZZrqQgBx/NkADPpSCzFmtrMt7p9xJMwJhlXHsnG2RjcZ+NdhTX1rijIte5FizdkeqRgPpmpWFwOYEBudc46Z3Ix8qZeI+FuJuHuHG++abA1nAzBrpZedhzZGDnGd222qH22o6jZN/s17cwYOQEdlx9KdeIOIdXl0a702XiG41OydEYcxyAxIbG++xGKe9dYNgxW2MWhqeRnO5LHJPWpjwWjHUXlx7MaZJ9ai2koEs0xnfepnwapFpNJjYsBmsbCj3uQXIzUKW0Pr5cnxLVK7ZQkUUYAAUAVGbNee7iTrlt/SpPDke1jOB0rYyHpHOYu3IvD7Pun93Y6lqrJjvpVhQ/BRk/qRVrA+zUQ7I7H7jwFpyEYaZTM2evtHI/TFS4dK5W6TlNnY4setaQDSZvePrSk0mb3j61EiZsFChQohFYWUPKoOKOuEAj6Uot48Rg1xdDbFZCibkZ7GLWYmktGjUkFhjI8Kxzxrpb6JrOp2Eoy8UzYOPBjkfowNbUvYTJbvyjflNZU+0bGsPF8IQcrT2yNJ8SpI/wBPpWjx9nRyT/RR5OtzhFr4ZXGjjn1KFcZ9oH0+NaB+ylYCbjbiDVwqf7DZLFFJIcKrHmY5PgM8maofhxV+9M7bALufIVqH7Klg8PZvqepRW0NxLqd8VCTyd1zJnkK83+Tb1FWGzOSLq0ewT8OSO4treELcGdO4uDIshzkPzHc5JJ+VOwO1IdPs0SCxMkbpJb2/drGXLBc4zueuMYHjtS7FAH6E2o3tjbDuby4jiDg4Dnr4VFLiKeSGbGk6JfL3fI3cSAEkfl+h8qf+JuV4IoY4bCa5kf2Fujtyjc4qOx2qnUImbhmHlllBM1tc4C+bEA7+dM2/2M9MmOnxrBaRRLEIVVRhAc8vwz40Xf6fYXSt96sracY37yJW/cGlC4UBR0A2HlXeRjJx89qkjZOK8MD6cH7RXvEHDGhzzyixsNAP8QIiySdy2QPaHskbk4xRQ7KeH7u0SR2vLOcqOYQz86g48ObNPF9ZG+kQnTdMvVeVskTe1y82zjz2H71LVUIqxqAqqAFUDZR5VZr5DIh6myGzj8ez3EqK/wCxkZzY61zY90Twgn6g01v2V8S2isIJLe5HVRFctGfoRirxr2rP/s5D8S00QPicff27Rn+80Dj/AE2Z3istRZOXlBAE23kcZPXNMcl5renRzW0+nSWyPgM5geN8joeb4elaeBrx0V/eVT6jNSQ5Spv+SpEL42+Mt12v/pmPTdba35GkE8ijGcyqc/UVJbbi/T1ulljOpWYjiITc8vMfQ9NhVn8Q8KrqLyyQW+lEsV5FuLRWC4znfGd80mHZlwxd2qteacLa6I/iNaysq5+A8KN38bY9yi0OquQh6mmVxqvHOp3y22mjUppIZnJkUsPaUDOD7I2+FLor5xCqBF5ceIzRfan2e2PC2kw6/pd3dMsNwolSbDBUbILZAyN8U1WN8kkUXMcFsL8zXMcj0ja/ov7TsOHTnR/KvuH62nwcrGI/8NiuenX6Cj9d4ju9P0TVL5oRlbF05hIQBgbHHSi7FE7rcA/GudesLbV9DudLed7dZ0KmRd8dMZHiPOoMTK+nP7g8zEU19qMlWZgjvZZr2xS7jdieQOAfXO+9KyuhmDDwanZPz++jhwBnw6U+8ZcMy6JqX3K90pY3Jys0Z9iZR+ZfDemFYdJCE/er+ylB6MOYE/LBrvaMiNkVo5O2mUJNMJvvugUfh+oyzY3ImTlb96b9UZmsI1LJzSS+0QME4G37mnOe3t+6L22r29yR+R4yCT8xTVqTBrm2GF2wSANqHOlqrwBUvuF9uOWNFA5dh0qccMqyaLFlQveZcgfT+lQhcggj5CrEsYzHY28ZGOWMD1qlxUW5tlHmJtQURfpC5vwcdFqT6cjy3McSAszsFUeZJqP6JGe8kkI2xgGp52aWQvuM9LgwCDchiPgvtf0q7mS0mzPwofckaZ0m3W00y3tFACwRLGAPJRj+lK1rlPzHzO1dCuWk9vZ2MV1SQDSZvePrSmkz+8aSEwUKFCnGIEvspgeVJ7nZcmj2YAYpFduTtvgdazPRr1+zgy4BrJP2lZebtBt4lcHltBkE7jLH/StVXEgCE5wKyL9olg/afMVbmAtkAwfi1S0b7+Acv/5kc0eKVNJu7xYiyqCMhc7gA4/Wtl9lHDCWvZNw9p0F48ZktYrmRVIYCUjn5vMEE/vWQezrRr/ijiK24atL4Wi3cjOXcZVQqli2MjOwqRX/AAPr2j/e7jh7iNdVFh3Czrp7yxTKZXKr/DIBzkfLIO9aMkY6ZtGH8Xt55JsQzySgBjzFVOPIb8ufGjjq1xGALjTJ85/3UinP1xWRrWbt24fhurqK416W0sZGSdlnF3HGUALg7t7vjj4+Ro7S/tBdoVikcl49lfW77q9xZlA3wDLgH5UyiFs01xDqNvPcwc+kGdQp5xNzBlz/ACkZG9I+FzpX4yoi0640yVfaAE+YmGMYxn4D6VTWkfabm5wNV4RhYeL2t6d/8rLt8qlNn2/dneor3eoabqdkT/Pbo6//AFYn9KHqLZeo3AIIOfKir4strJgrnkbl5mwC2NgT4VWumdoXZlqCJ914otLRydhLI1vg/wCYAVKIvumrWjR2WvreQTAbLOkqtjpg7n6b0/UI8ttPH47DnQTHHEoMdzHcE8hG+GAPw26/GpUetROw0m70zUHv4ZluJWyGWUsFOTk7ZwOvX4U7LqV8AWn04EZ27l8k/I4ptDjrQpuXWIs4uLW6g3wC0eQf+kmj01LT2x/tkKknGHblP0OKZoQsWuqLSRGGVdWHmGBrqbvBCxjA5+UlOboT4fKloQ0T/wDiaORDbx2E6FyGDMQeUn98VIVyRlhhj1Gc0x2dzrwvI4bqwh+7Mo5pY33DeO3lT2mwFCwkhv4ksLTU9CvdOv1Btp4HWT2cnBHX5day7Hb3uicQzaNePztbv/DkzgSIRlTv5j9TjrWqNYlSKz9q4W3Z2CLIRkAnpt0qou13TJ7jRvx9TpszWL8rTWu0jIz7B8Yzu2enhVa+ClEvYt3SaiM+l3jSx4BIwcEZp1gQMMnOc+WaiegTZZSSRzjpUy0oZU53rIbaejfjprYRqeh6XrVi9lqVmssUmwyoJT4g+BrOXGPDttw7r82lXF/cwSRjvIpHjLCVCfZIP6fKtWWsYY4PSoD9oHhJda4QbVLKPGo6YpkRkHtPH1ZPTxHzrZ4rLlVPTfgx+SxozjtLyZsv7drcIyz2t2jZIKAhk9cgU2ThZtacEBVj2AAwKUpfh4wHSOTHUsMmkmmky3tzI+CTuDXS59r6R18nN0w0x4so+9u4Y/5nA+WasqOPog3xsKgHDMIm1yAEEgZOBVjwJsD8an4yPWLZh8xPtYoi3Sk5bY/FjVp9g1kJeMkuSM9xBI31wtVjaLy2yfGrq+z1b/7VqdxjeNFiz4bkk/tQ5830YXHQTsRcg6V0K4WuxXOs6cFJm94+tKTSVvfPrT7GZ7QoUKcYriSQdc9KQXEpDHfOa8llIGS21JZJOY5zWW2bMEJtSuGVcKcfGsqdqFiNZ411G7t5eSaKTuyrHCvgD6HOa0jxFqCW9vcTO4WOFCzNnoAKzFJO93dTXTkl5ZGdifMnNa/DUQucnL4Mnm8qeNXHqOPY9rNvwTxXNq2taTf3cDWc1qps2AeNpABzqc9QAfrU84K7R+DtF4quNQutR4gurzVJUspr7UxyPY2YXZsjZ5A2CMHYDz61spYHIYj0NdczFeUnK+TDIralxn9nOw5Rr84lhz8a8L6RonBfCTXNjrtpIJDqt2tw+bYy3BLt4Z9gfm8xT/x3a8RW2lcdT8Y3FkvC33Yx6DCskZjL5/gmADcbYJ6eNUvNa2swIltYH5hg/wAMD9qJu7CC5hWFjOEQYQd8xC7eAJIqvPjpr8SxHlK37WiJ8xLE7gZ2oxHYDZiPnTvLoIXeC58P94tJ30O9UZR4pBjwfeoXh2x9onhnUy9sRCeQHBYkeRo+1v57Z+e3dom/mQ4P6UXPY30S8z2snL/MNx+lJwsi7sjgfFTUcqprxonV8H6ZM9F7R+NNKI+5cT6vGF3C/eWZR8jmpjpPb72hWn9vqFtegeFxbjJ/zLg1T0ZzuDRvNUbjr2Sdv7NEaR9pfU1IXVOG7OVcYLQTlSfkQalmmfaH4Ku8fiOi6jZ+HMESQZ8em9ZLz9K9DlfdYgeWBQ62O5tGxf8AzO4JvL9H0vVdKS3Y5ZLtmt5FPnkjFSvQ+J7bUkP3S8dAELhor6OVWH8oGc5+VYS79+XqfrQS8kjcMmQc55lYqwPmCPGm6hKbPoJa8QT5lH3pf4Sc7d/bYHKDv7Q60ssOJTdrKY1tpRDu5jl5SBnHQjzrA9jxtxXYjFlxFqkIHTNwXA+TZFP9l218fWg5Jb+01BWGHF3aKeceTFOUnzqNxJIyNs6pq00sEYj0/vEDEsso5wdtsYPnTDcSQSaDe21xo9rbSXSCAxQ7CQEMCTnpjOazPpf2idSggaDUOGLVgygM1jeSQscHY5bmx6dKlWl/aP0aSNUu/wDxBZnlVcSQRXatjzxyk758KDon4DjPUkwaLcTQ3zWzli8UhQg/lILD+nWrD0Sc+zzeNVXecT6LrvFc+paBeCeCco8h7oxYcjG6noSAPrVg6Bcc4Qk9RtWNfDrM6eiXatMnlngrzCjJo0ljeJ15lcFWB8QQRikmms3c4pRJJynpn4eBoYS09oGytTWmYe490puHOMtU0ZlZfu9ywX4qxLL+hpLogH3dnx7RODVv/aw4YWK7seKrSNuWb+Bdt+VSAeQ/PpVQ6cvJaR4I9oZ2PnXSrId0I6+Dl8it1TeyW8EITqbuB7idfXarAtVyv6VB+CcR20su3M78vy2qb6ceZAM9Dmuiwo6p2chnPvfsdLdBmNBjAIxnpV+fZ/tOTh2+uzjM11jP/Ko/1qhYN54x0HU1o3sUg7jgmIjpLM7/AK4/pWdyEv4zQ4yO7CbCuxXNdCsU3wGkze8fWlNJmB5j604wKFChTjFMtNz5zmkF3OUBVXwPWjSTyH0pq1G6t7a3luLqVI4olLMzHAAFYzN6K0Qjtj1RbLhl7aN+Wa8PIuDuR+b9KpqIArsMUXx9x+mvcTSzNHINPi/h2wU+Gd2O/n+lE6fqVjcoBDOCfEHY113Dyrpq0/Gzj+ajbff68IXbV6K5LIMe2uT0Ga9B2z4VsxtjL0znnXJPTR6TivAa8OepBFBTUi8kckeivcnffrtXlDBp9DeToHbc14QCcEAg+dACvQN6bqv0JSa9M4e0s3P8S1ifwOV8KTy6Np7qRBFJGx6BMnf0xS2p12LtpX/iC5S+mt47gwj7uJsAMc77n4VXupr1tos022N/kVXcaFy/2V3ggkFXA8PSk7aReqoMXdyqfHOK0Txl3DalNbzpaR2a2yyxzNY973z5bK82CFGw6b4NKLfgfhzVNIhurnSBZTzxCR1gkIKk+A/0PSs9Qqn8aNRW3RX2szLPZX0K5ktXC+YGR+lInLBsFWXHmMZqxdYhjs9RuLS3lZ4oXKqx6/Omq4SOQYkRX/5hmlPAWtxFDkJb+4hbv4dKLY7VKprGycktbr0xttTdcaZbc3sc67+earTwZotwzoP4GBwSc70U5K5xtkU6z6Y+WMcu2OhGKRy2Nwp6K3xU5qq8ea9lqN0ZfI79m+qNpvEkKFv4Vx/DIOwB8D+wrTfDE5eOPHSskLHc28izLE/NGQ64HiN6092eXr3OlWk2COdAd+o2rF5Ctwe9HRcZd2j1TLStLp0VQrEeFLwzMgJJNR+Cf2RvTzpkyTqYxnmFZ5qtCHizQrPinhu70O/5xDcAe0uOZCDswz4j+tVdcdgsSLzW2uydyoAUPEOYYHmCP2q5likWbIO1LVX2cVPVbOK8MqX0Qs/JFEp2V6tpsIWwuYLoA5Ocg0t0fhXXo5GE1i2VyOoPzq6XiRh7SijLO1Xv1IGxrZq5q+EOnswL+Ax7J9t6Kcaxuba7CXELRnBGWGwrSnZZCIeA9LA6tEWPzJqO3el2l1C8NxGroykbjOKX6Nr40mzi0/uFMMChRg4NHZyKvhprRBTxf+NPcfJOq9UbVH4uKtPYLzJKufh0pfY65pt0xSK4HN5MMZqFWJk0qZr4HLBpO3vH1o8SKwypBFJ3941JFp+iF7R5QoUKIRRc8ixRMxPhVKdvur3k+jpp1sXS3lkAdf8A5AN9/hVu6irTIsYzhjg1CuLeGk1NgJ1LCNSV+B3rCcvJ0Shsyx9xmU8vJuOuNq7eBoI+9ckEePjVpanws9pNJ3keVB9kgdRUI4ptiblbWBMlBzPjpjetCm1y9FTIrUV5RHZLyaUqs38VF6ZOCKc9LXkBeC5uLdz/ACttSdLB+pj3+ApfZW0oX3D1q9XdOPpmXOiuXtD5p818oHPPBOv/ABpgn507wRtInMYIiT4IaYII5UUeB+Bp0tJnUDNXK8yxfJWnhUv4HF7KcKG+6TKPMAH+tFtAAcMe7z051Iz+lOOm6nJEyspYAbVKNO1dJUHOI38CGAII+lWoZsipPi65EEWFmzylWI+Irxo3Xdlqzo4eHr8gT6dbBsYzyD9xiupOEOHpRlFmhY+Mcxx9DVqOd/RSs4vXplW4+FeMoKkHB9RtVjXPAFufatdUffoJYw2KbbvgDWFGYZLW4HgVJQ/SpllRktNFeXHzj6Ila3V5aAm0vJ4DjAMblf2+X6062vGXFEERjXWLh1IIIkIY/UjIrq74U4gtvf0yVl/mjww/Smm6tLi2blnt5Yj/AMaEU6sqYDpuh4E8jySSM7uWZjksepoiY7+go/Y9CPpU67NdCsrhhqN/AsjBsokicyqPPBqPIy4Uw7Ms4eDbk2dV4K1lPkd/KkzjNak/CdJvIzZ3thbTQsuwaIYHptVXdpvZe+mwyarw+GltlOXts5dB4keYrMq5eq16a0a+RwduPHsnsquCAT3McJPKJHCknwycU969o2nWTfdobN+aGISysJiCF36fzE4NMDllbIJBB+YoyTUr0ujzTCZo2BjMiB8fWrEoOXkoQ+32Ga/pEFjFFPDKWD/lc5+NW32ZzE6DZljyt3Q2+VUtrGpXOoNH94ZcR+CjAPxqzOyq9kuNBQu2XjkZM+GAdhWVyy3Ujd4WWr2t/Bbunyl23HSnnTHMVxzqSKjukOzMvketSOyjZhkCub0dYmSqMJcRrNFjl6HzzR3IMDamLSLs2txySEhG61IxyuodDsaOK0QTb2EhMnlwKWWqAeAGKJXAOds0Yr4FGkVZvyKpXIUlRkimS4RjKSwwTvToGPnRF2nNhttqNoKGkIlRvj0o+IHzP1oEYoA4NNvQpafgcrbWNRsgBFNzqPyv0qZ2sxuLSGcjBkjViPUVXUh3xVgaQf8A0m0/wE/YVbxZNszsutRW0Ka9oUKumcUmbZAQPAUlubRHDkDqD4U5hc1w6AHZS3w86wdHT7UVtlbcaaeF09uSIyTs4EaKPeJ/pUG/8vbgjvLiJu+fdiRvWjtH4S+9y/iF8g5v92pGeWnOfhqLoEzt5VoULSMzKv7vRlO54EkjP9kaQTcJ3EOQsTY9K1Xd8KQuN4l+lM17wZGQeWJcdelWYlByMzPw/OnWJz8qLGjyod4ZB8q0TccGpg/wh9Ka73gwH3I8YHlUsXoBso+G0K7KCfPNKoUdegx6VZ1zwQ3Of4B3HlTfLwgysQIG2qeMgOzIbazSRH3ifnTraavNEAp9tfjTlNw73XWEj1FJ30Yg+yhH+WpO+gGOWn6wrj3iPMGnaHUgV25fnUWGnSxH2FPXypTDHcx/lJHxqWNoHUlcOpKuNwM9cUqMtpMB3qpIMfmUH96iCXGNipo9L3l8WAo+4LgmPV3ofDt4uJ9Ls2z4rGFP1FNFtBHpmoTWtqBFarumc4UfrXSX4JwH3ojVrlpbGQRnMrEAtncA1Sz/ALqzQ4txru8kk0TUvvkid2xKx5BPmalBAkiGQNvAjO1QThJ+VQsYXYYPrVgaerPCC69a5tSezqXFNFF9sfAj2lx+MaPbloZW/jRovQnyAqp7uGSNuR1KsDgqwwc/Oth6xauIu8jzlPaGPhSG5j0y/hC3llbTqy4IkjU+vhXQ8flzlHo2ctymAq5/Uj8+zHNwr5OUYeoqw+xmVvutzb5yonDAeWRVxar2e8C6gDz6WsDEbtbuUP74/So63B+l8KX0f4VcXMsVyd1nIblI8jij5GXeoi42PW/aJfoaeypqW6epEVRjQRstS6zGYq5s6ls9aBX6gZpbp87wKEfoNtq5RQTQdeUHI+dEkC2OUbiTda65t6SWTFVHkaVoPbXPTO9GivKPkMSQdDXksnh4Vxkc3L1xXLqQhNE2NFa9njHNcb81chiW3owUDC15OZei+tWBo/8AdNn/AICf9oqv5ui+tWBpH91Wf+An/aKuYftlLO/FCyhQoVeMop6IF3VEBZm2UAbt6VMeGuFjzLeXwDH8sfgPifjTlwzwpa6UFmmY3N1jeV/6DwFSNQAMAAD4VSqp0vKNHIyu3iIlW1i5QAgAAwABXJtI89P0pYBXpAqdLRQbb9jfJYofAfSiH09SOi07YrwqvlTiI9NpiE45FIx4Ck0uixMdk/SpSVXyH0rkxjyp9jMhU+gqz5CDYU3XHDas5PJj0qw2gBzsKJe1BHRTRdgdFZXvC8ZT3KarvhZAg9jerceyVhvGp+VJJ9NjbrEn0p1NjNFO3PC2MgITt1FNtxw3KE2U59KumbTI9wI1puudIXGORd6kU2NopG50FwxHdnPpTZcaY6g8oPxzV2XeghicKM0z3vD3snKLv8KNWi0UzPaMp6FTSZ1dCTvjG9WdfcOAk5i8aj+q8PMq+wmNvlQXTUoNEtD62JjVwkQgJJxl+arN0qRZLcYOSKqS3aWzuDCQVKmp3wxqGMK5O4rm/Ozr9pw8Epl3UjGdqiWuWz20jTRrgEe74VLQysAwI3pDq8Uclvh8VPVY657K9tSuh0kV6dZZGIyy4oi5vUv7mDLFuQnY+FIuKLWS1nd0PsE528KT8Ntz3OW33rUvyldS2jIqxXj39WWHoKoqAYBqR2bjkI+NRbSXKjGTvjFSGxflG/j0rINhjsh2r1ix2BPpRMRyQc+NHqQGz8KdAMOjJCgGlUTAgE0jDijY5By4HWiG0HfnyNq7yDsc0WpzQYnOwNNsDQXcgIyldtq9jkHiDmupCpUFxuKTk+0eWnFoVPhlG3Sp/pf912f+Cn7VXcTkgg9asTS/7qs/8Ff2FXMPy2Uc78UKqFChV4yg4nNChQqJsJgoUKFIQKFChSEChQoUhANc10a5pCAaLdaMrl+tIYIdAc7Ck0sKk5IpY1ES0+xMQSWwJO9IrizBbwxTq/Wks3vGnTBGC8tI8H2B9KZNR01GGQtSe68abrn3TTsSKm430MW8y3cUfssfaOOlNWm3HdMu+4qweOP7inqs094VlZEFF+Do8G1yr0yc6XfK8YVm6UsuZEmQDIzUa0jrTxD7wqtsuEd4ssg8T8q+8N8VFOEoWjluFcElX2z5VPNf/s2qH6F/725/5qSk1tIU/ue2SuzPKqY86fLWX2VzTHb+6lOsPRfSkmRMeoZAVNHRN7NILT+lLYfdowRShLDO1dKcGuIvcr0+9S2IWW5DHA/WjCcHFFWXvCjW9806BfsKm32PlQRQU6ChL73yr2P3TRDo55cNtVi6V/dVn/gr+wqvD7wqxNK/um0/wV/areJ4bM7kPxQpoUKFXzKP/9k="

def _get_otdr_path():
    import base64 as _b64, tempfile as _tmp
    try:
        data = _b64.b64decode(_OTDR_IMG_B64)
        t = _tmp.NamedTemporaryFile(suffix=".png", delete=False, dir="/tmp")
        t.write(data); t.flush(); t.close()
        return t.name
    except Exception as e:
        logger.warning(f"Imagen OTDR no disponible: {e}")
        return None

class KeepAlive(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers()
        self.wfile.write(b"Bot Incidencias FO - Activo OK")
    def log_message(self, *args): pass

def iniciar_servidor():
    HTTPServer(("0.0.0.0", PORT), KeepAlive).serve_forever()


# ── Estados ────────────────────────────────────────────────────────────────────
(AUTH,
 CODIGO, FECHA_INC, HORA_INC, FECHA_ARR, HORA_ARR, FECHA_REP, HORA_REP,
 FECHA_RES, HORA_RES, MOTIVO, TESTIGOS, REMEDIO, ENLACES, PERSONAL, LUGAR,
 COORDS,
 FOTOS_ANTES, FOTOS_DURANTE, FOTOS_FIN,
 TRAZAS_ANTES, TRAZAS_DESPUES,
 OBS, RESUMEN,
 MANT_CONFIRM,
 MANT_EST, MANT_HILO, MANT_DA, MANT_PA, MANT_DD, MANT_PD, MANT_OBS,
 MANT_MORE_HILO, MANT_MORE_EST,
 CONCLUSIONES, CONFIRMAR,
 MOTIVO_ZONA, MOTIVO_TIPO, MOTIVO_TRAMO,
 OBS_MANUAL,
 REMEDIO_METROS, REMEDIO_MANUAL,
 ENLACES_ZONA, ENLACES_TRAMO, ENLACES_MAS, ENLACES_MANUAL,
 CUADRILLA_ZONA, CUADRILLA_SEL) = range(48)
MENU = 48   # ← menú principal por pestañas

# ══════════════════════════════════════════════════════════════════════════════
# CORRECTOR ORTOGRÁFICO — campos de texto libre
# ══════════════════════════════════════════════════════════════════════════════
def _verificar_ortografia(texto: str) -> list:
    """
    Detecta errores conocidos del diccionario propio de Telconet.
    No usa corrector general para evitar falsos positivos.
    Devuelve lista de tuplas (palabra_incorrecta, sugerencia).
    """
    _PROPIAS = {
        # Marca
        "telsonet":"Telconet","telconett":"Telconet","telconed":"Telconet",
        "telkonet":"Telconet","telcomet":"Telconet",
        # Verbos mal escritos en reportes
        "tizne":"tiene","tizna":"tiene","tisne":"tiene",
        "notifisa":"notifica","notifísa":"notifica",
        "informe":"informa","informo":"informó",
        # Tildes faltantes
        "kilometro":"kilómetro","kilometros":"kilómetros",
        "tecnico":"técnico","tecnicos":"técnicos",
        "optica":"óptica","fibra optica":"fibra óptica",
        "atenuacion":"atenuación","resolucion":"resolución",
        "reposicion":"reposición","derivacion":"derivación","fusion":"fusión",
        # Errores tipográficos comunes
        "serca":"cerca",
        "llego":"llegó","salio":"salió","realizo":"realizó",
        "encontro":"encontró","detecto":"detectó","reporto":"reportó",
        "reparo":"reparó","identifico":"identificó","escalo":"escaló",
        "traslado":"trasladó","procedio":"procedió",
    }
    errores = []
    texto_lower = texto.lower()
    for mal, bien in _PROPIAS.items():
        patron = r'\b' + re.escape(mal) + r'\b'
        if re.search(patron, texto_lower):
            original = re.search(patron, texto, re.IGNORECASE)
            if original and original.group(0) != bien:
                errores.append((original.group(0), bien))
        if len(errores) >= 5:
            break
    return errores[:5]

def _aplicar_correcciones(texto: str, errores: list) -> str:
    """Reemplaza automáticamente las palabras incorrectas por sus sugerencias."""
    resultado = texto
    for mal, bien in errores:
        resultado = re.sub(r'\b' + re.escape(mal) + r'\b', bien, resultado)
    return resultado

def _msg_spell(texto_original: str, errores: list) -> str:
    """
    Muestra el texto ya corregido con las palabras cambiadas marcadas en negrita.
    Estilo Word: ves directamente cómo queda.
    """
    texto_corregido = _aplicar_correcciones(texto_original, errores)
    # Marcar cada corrección con tachado → negrita
    cambios = "\n".join(f"  🔴 ~~{mal}~~ → *{bien}*" for mal, bien in errores)
    return (
        f"✏️ *Corrección ortográfica:*\n"
        f"{cambios}\n\n"
        f"📝 *Texto corregido:*\n"
        f"`{texto_corregido}`\n\n"
        f"¿Confirmar?"
    )

def _ikb_spell():
    """Un solo botón de confirmar + opción de reescribir."""
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Confirmar corrección", callback_data="sc:fix"),
        InlineKeyboardButton("✏️ Reescribir",          callback_data="sc:retry"),
    ]])

# ══════════════════════════════════════════════════════════════════════════════
# INLINE KEYBOARDS — No reemplazan el menú de abajo
# ══════════════════════════════════════════════════════════════════════════════

def _ikb_zonas(prefix="z"):
    """Teclado inline de zonas (2 por fila)."""
    rows = []
    for i in range(0, len(ZONAS_ORDEN), 2):
        row = [InlineKeyboardButton(ZONAS_ORDEN[i], callback_data=f"{prefix}:{i}")]
        if i + 1 < len(ZONAS_ORDEN):
            row.append(InlineKeyboardButton(ZONAS_ORDEN[i+1], callback_data=f"{prefix}:{i+1}"))
        rows.append(row)
    return InlineKeyboardMarkup(rows)

def _ikb_lista(items, prefix, extra=None):
    """Teclado inline de lista (1 por fila)."""
    rows = [[InlineKeyboardButton(str(item)[:55], callback_data=f"{prefix}:{i}")]
            for i, item in enumerate(items)]
    if extra:
        for label, cdata in extra:
            rows.append([InlineKeyboardButton(label, callback_data=cdata)])
    return InlineKeyboardMarkup(rows)

def _ikb_listo(sec):
    """Botones Listo/Omitir inline por sección."""
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Listo",  callback_data=f"lo:{sec}:l"),
        InlineKeyboardButton("⏭ Omitir", callback_data=f"lo:{sec}:o"),
    ]])

IKB_MOTIVO_TIPOS = InlineKeyboardMarkup([
    [InlineKeyboardButton("⚡ Corte del Tramo",         callback_data="mt:0")],
    [InlineKeyboardButton("📉 Atenuacion en el Tramo",  callback_data="mt:1")],
    [InlineKeyboardButton("🔧 Ventana de Mantenimiento",callback_data="mt:2")],
    [InlineKeyboardButton("✏️ Escribir manualmente",    callback_data="mt:m")],
])
_MOTIVO_TIPOS = ["Corte del Tramo","Atenuacion en el Tramo","Ventana de Mantenimiento del Tramo"]

# ── Mantenimiento: helpers de panel ──────────────────────────────────────────
def _mant_nodos_a_filas(nodos: dict) -> list:
    """Convierte mant_nodos → mant_filas (formato que usa generar_excel)."""
    filas = []
    for est, hilos in nodos.items():
        for hilo_num, datos in hilos.items():
            filas.append({
                "est": est, "hilo": hilo_num,
                "da": datos.get("da",""), "pa": datos.get("pa",""),
                "dd": datos.get("dd",""), "pd": datos.get("pd",""),
            })
    return filas

def _mant_hilo_completo(datos: dict) -> bool:
    return all(datos.get(k,"") for k in ["da","pa","dd","pd"])

def _mant_hilo_icono(datos: dict) -> str:
    filled = sum(1 for k in ["da","pa","dd","pd"] if datos.get(k,""))
    if filled == 4: return "✅"
    if filled > 0:  return "🔶"
    return "⬜"

# ── Panel principal ───────────────────────────────────────────────────────────
def _texto_mant_panel(nodos: dict) -> str:
    if not nodos:
        return ("🔧 TABLA DE MANTENIMIENTO\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━\n"
                "No hay nodos aún.\n"
                "Toca ➕ Agregar nodo nuevo para comenzar.")
    lineas = ["🔧 TABLA DE MANTENIMIENTO", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for nombre, hilos in nodos.items():
        completos = sum(1 for d in hilos.values() if _mant_hilo_completo(d))
        lineas.append(f"📡 {nombre}  —  {completos}/{len(hilos)} hilo(s) completo(s)")
    lineas += ["━━━━━━━━━━━━━━━━━━━━━━━━",
               "Toca un nodo para llenarlo o agrega uno nuevo:"]
    return "\n".join(lineas)

def _ikb_mant_panel(nodos: dict) -> InlineKeyboardMarkup:
    rows = []
    for nombre, hilos in nodos.items():
        completos = sum(1 for d in hilos.values() if _mant_hilo_completo(d))
        icono = "✅" if completos == len(hilos) and hilos else "🔶" if completos else "📡"
        rows.append([InlineKeyboardButton(
            f"{icono} {nombre}  ({len(hilos)} hilo{'s' if len(hilos)!=1 else ''})",
            callback_data=f"mp:nodo:{nombre}")])
    rows.append([InlineKeyboardButton("➕ Agregar nodo nuevo", callback_data="mp:nuevo")])
    if nodos:
        rows.append([InlineKeyboardButton("📝 Observaciones", callback_data="mp:obs")])
    rows.append([InlineKeyboardButton("✅ Listo — volver al menú", callback_data="mp:menu")])
    return InlineKeyboardMarkup(rows)

# ── Panel de nodo (grid de hilos) ─────────────────────────────────────────────
def _texto_mant_nodo(nombre: str, hilos: dict) -> str:
    completos = sum(1 for d in hilos.values() if _mant_hilo_completo(d))
    lineas = [f"📡 NODO: {nombre}",
              f"━━━━━━━━━━━━━━━━━━━━━━━━",
              f"✅ Campos OK: {completos*4}/{len(hilos)*4}",
              "",
              "Toca el hilo que quieres llenar o editar:"]
    return "\n".join(lineas)

def _ikb_mant_nodo(nombre: str, hilos: dict) -> InlineKeyboardMarkup:
    rows = []
    # Grid de hilos de 2 por fila (como la imagen)
    items = list(hilos.items())
    for i in range(0, len(items), 2):
        fila = []
        for hilo_num, datos in items[i:i+2]:
            icono = _mant_hilo_icono(datos)
            fila.append(InlineKeyboardButton(
                f"{icono} Hilo {hilo_num}",
                callback_data=f"mn:hilo:{hilo_num}"))
        rows.append(fila)
    rows.append([InlineKeyboardButton("➕ Agregar hilo", callback_data="mn:add")])
    rows.append([InlineKeyboardButton("↩ Volver al panel", callback_data="mn:back")])
    return InlineKeyboardMarkup(rows)

# ── Panel de hilo (grid de campos) ────────────────────────────────────────────
def _texto_mant_hilo(nombre: str, hilo_num: str, datos: dict) -> str:
    def val(k): return datos.get(k,"") or "—"
    filled = sum(1 for k in ["da","pa","dd","pd"] if datos.get(k,""))
    return (
        f"📡 {nombre}  —  Hilo {hilo_num}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Campos OK: {filled}/4\n"
        f"\n"
        f"  📏 Dist. Antes:   {val('da')} km\n"
        f"  📉 Pérd. Antes:   {val('pa')} dB\n"
        f"  📏 Dist. Después: {val('dd')} km\n"
        f"  📉 Pérd. Después: {val('pd')} dB\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Toca el campo que quieres llenar o editar:"
    )

def _ikb_mant_hilo(datos: dict) -> InlineKeyboardMarkup:
    def ico(k): return "✅" if datos.get(k,"") else "⬜"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"{ico('da')} 📏 Dist. Antes",   callback_data="mh:da"),
         InlineKeyboardButton(f"{ico('pa')} 📉 Pérd. Antes",   callback_data="mh:pa")],
        [InlineKeyboardButton(f"{ico('dd')} 📏 Dist. Después", callback_data="mh:dd"),
         InlineKeyboardButton(f"{ico('pd')} 📉 Pérd. Después", callback_data="mh:pd")],
        [InlineKeyboardButton("↩ Volver al nodo", callback_data="mh:back")],
    ])

# Teclado legado — conservado por compatibilidad
IKB_MANT_CONF = InlineKeyboardMarkup([[
    InlineKeyboardButton("✅ SI", callback_data="mc:si"),
    InlineKeyboardButton("❌ NO", callback_data="mc:no"),
]])
IKB_MANT_MAS_HILO = InlineKeyboardMarkup([[
    InlineKeyboardButton("➕ Otro hilo",    callback_data="mh:si"),
    InlineKeyboardButton("✅ Siguiente",   callback_data="mh:no"),
]])
IKB_MANT_MAS_EST = InlineKeyboardMarkup([[
    InlineKeyboardButton("➕ Otra estación", callback_data="me:si"),
    InlineKeyboardButton("✅ Finalizar",     callback_data="me:no"),
]])
IKB_ENLACES_MAS = InlineKeyboardMarkup([
    [InlineKeyboardButton("➕ Agregar otro",        callback_data="em:mas")],
    [InlineKeyboardButton("✏️ Escribir manualmente",callback_data="em:manual")],
    [InlineKeyboardButton("✅ Finalizar enlaces",    callback_data="em:fin")],
])
IKB_CONFIRMAR = InlineKeyboardMarkup([[
    InlineKeyboardButton("✅ SI, GENERAR EXCEL", callback_data="cf:si"),
    InlineKeyboardButton("❌ Cancelar",          callback_data="cf:no"),
]])

# ══════════════════════════════════════════════════════════════════════════════
# TRAMOS POR ZONA
# ══════════════════════════════════════════════════════════════════════════════
TRAMOS_POR_ZONA = {
    "UIO": [
        "GOSSEAL - TABACUNDO","GOSSEAL - CAYAMBE","DATA CENTER - CHIRIBOGA",
        "DATA CENTER - TANDAPI","GOSSEAL - LATACUNGA","ARMENIA - MACHACHI",
        "DATA CENTER - NANEGALITO","MACHACHI - PUJILI","GOSSEAL MACHACHI",
    ],
    "GYE": [
        "AURORA - TARIFA F2","DATACENTER - DAULE F2","DATACENTER - PROGRESO F1",
        "DATACENTER - PROGRESO F2","DATACENTER-MILAGRO F1","DATACENTER - NARANJAL MAIN",
        "DATACENTER - NARANJAL BACKUP",
    ],
    "Ibarra": [
        "Ibarra SDH - Cotama SDH","Cotama SDH - Tabacundo","Ibarra SDH - Urcuqui",
        "Urcuqui - Lita SDH","Cotama SDH -San José de Minas","Apuela - Pacto",
        "Ibarra SDH - Cayambe SDH","Cayambe - Pifo","Tabacundo - Malchingui",
        "Malchingui - Guayllabamba","Tabacundo - Cayambe","Guayllabamba-San Jose de Minas",
        "Quinche - Guayllabamba","Cuzubamba - Pifo","Bolivar SDH - Ibarra SDH F02",
        "Tulcan SDH - El Angel SDH F02","El Angel SDH - Ibarra SDH F02",
        "Tulcan SDH - Bolivar SDH F02","San Gabriel SDH - Ibarra SDH",
        "Tulcan SDH - Bolivar SDH","Julio Andrade - Santa Barbara",
        "Sta. Barbara - La Bonita","Tulcan SDH - El Angel SDH",
        "Bolivar SDH - Ibarra SDH F01","El Angel SDH - Ibarra SDH",
        "Tulcan- Rumichaca Principal","Tulcan-Rumichaca Backup",
        "Bolivar - Pimapiro","Pimampiro - Ibarra",
    ],
    "TS.MILAGRO": [
        "MILAGRO - BABAHOYO MAIN","DATA CENTER - MILAGRO BACKUP F02",
        "DATA CENTER - MILAGRO MAIN","MILAGRO - BABAHOYO BACKUP",
        "MILAGRO - YAGUACHI BACKUP","MILAGRO - CUMANDA","MARCELINO - EL ROSARIO",
        "NARANJITO-PUENTE PAYO","CUMANDA - CHILLANES","VIRGEN DE FATIMA - TRIUNFO F01",
        "VIRGEN DE FATIMA - TRIUNFO F02","MILAGRO - VIRGEN DE FATMA",
        "TELEPUERTO - VIRGEN DE FATMA MAIN","TELEPUERTO - VIRGEN DE FATMA BAKUP",
        "VIRGEN DE FATIMA - PUERTO INCA","VIRGEN DE FATIMA - TAURA",
        "TRIUNFO - TRONCAL - PUERTO INCA F02","TRIUNFO - TRONCAL - PUERTO INCA F01",
        "LA TRONCAL - ZHUD","EL TRIUNFO - CUMANDA F01","EL TRIUNFO - CUMANDA F02",
        "CUMANDA - PALLATANGA","NARANJAL - PUERTO INCA MAIN","NARANJAL - PUERTO INCA BACKUP",
        "NARANJAL - EL GUABO MAIN","NARANJAL - EL GUABO BACKUP","NARANJAL - MOLLETURO",
        "DATA CENTER - DAULE","DAULE - PEDRO CARBO","PEDRO CARBO - JIPIJAPA MAIN",
        "PEDRO CARBO - JIPIJAPA BACKUP","DAULE - PALESTINA","DAULE - SALITRE",
        "SALITRE - LAS LOJAS","SALITRE - SAMBORONDON","SAMBORONDON - CERRO SANTA ANA",
        "SAMBORONDON - TARIFA",
    ],
    "TS.RIOBAMBA": [
        "Guamote - Alausi","Riobamba - Alausi","Alausi - Manga COMPUD","Alausi - Chunchi",
        "Puyo - Arajuno","TEN-Eloy Alfaro SDH (O) / PUNIBOCANA",
        "Archidona (E)-F01-BB / Y DE NARUPA","Cajabamba - Pallatanga",
        "Columbe - Cajabamba","Riobamba - Cajabamba","Cajabamba - Guamote",
        "Chillanes (E)-F01-BB /","Chimbo - Chillanes","Guamote - Columbe",
        "Guano - Comil","Tena - Chontapunta","Guamote - Macas","Guano - Riobamba",
        "Guanujo - Salinas","Guanujo - SDH Guaranda","Balsapamba - Guaranda",
        "Riobamba - Guaranda","PUY-Puyo SDH (O) / RIO NEGRO","PUY-Shell (E) / PTO SANTA ANA",
        "Chambo - Parque Industrial","Penipe (E)-F01-BB / MANGA CAHUAJI",
        "Quimiag - Penipe","Puyo - Tena","Puyo - Chuwitayo","Riobamba - Quimiag",
        "Salinas - Simiatug","San Juan - La Silveria","San Miguel - San Pablo",
        "San Miguel - Chillanes","Simiatug-F01-BB / FACUNDO VELA",
        "Simiatug (E)-F01-BB / EL SALADO","Velez - Cerro Calvario","Riobamba - Manga Urbina",
    ],
    "TS.EL COCA": [
        "24 DE MAYO - ARCHIDONA","LORETO - 24 DE MAYO","COCA - LORETO",
        "COCA - JOYA DE LOS SACHAS","COCA - DAYUMA","SHUSHUFINDI - JOYA DE LOS SACHAS",
        "JOYA DE LOS SACHAS - JIVINO VERDE","LA BELLEZA - LORETO",
        "LA BELLEZA - CHONTAPUNTA","DAYUMA - LA BELLEZA",
    ],
    "TS.SALINAS": [
        "Salinas-Playas","Progrreso-El Azucar","Salinas-Progreso","Azucar-Julio Moreno",
        "Salinas-La Entrada","Colonche-Libertador Bolivar","Progreso-Cerecita",
        "Sabana Grande-Cerecita","Cerecita-Julio Moreno","Playas - Progreso F02",
        "Playas - Progreso F01","Playas-Posorja F01","Playas-Posorja F03",
        "Colonche-Julio Moreno",
    ],
    "TS.CUENCA": [
        "ZHUD - CAÑAR PRINCIPAL","GUALLETURO - TRONCAL","ZHUD - CAÑAR BACKUP",
        "CUENCA - CAÑAR PRINCIPAL","CUENCA - MOLLETURO","CUENCA - CAÑAR BACKUP",
        "CUENCA - GIRON","CUENCA - GUALACEO","CUENCA - JIMA","JIMA - OÑA",
        "CUENCA - SIGSIG","DESCANSO - PAUTE","EL CABO - SAN JUAN","AZOGUES - PAUTE",
        "CUENCA - PAHUANCAY","GIRON - LENTAG","SANTA ISABEL - GIRON","MENDEZ - TIWINZA",
        "GUALACEO - ÑUÑURCO","GUALACEO - SIGSIG","GUALAQUIZA - GRANADILLAS",
        "GUALAQUIZA - SAN JUAN BOSCO","SANTA ISABEL - CASACAY","LIMON -GUALAQUIZA",
        "CHUWITAYO - MACAS","MACAS - MENDEZ","SAN JUAN BOSCO - MENDEZ",
        "MENDEZ - PAHUANCAY","SANTA ANA - LOGROÑO","SANTA ISABEL - PUCARA",
        "ZHUD - CHUNCHI","JARATA - RAMADA","ZHUD - JAVIN","GUALLETURO - JAVIN",
        "PUCARA - TENDALES",
    ],
    "TS.MACHALA": [
        "Machala - Pasaje","El Guabo - Machala F02","El Guabo - Machala F01",
        "Pasaje - Santa Rosa","Machala - Santa Rosa","Guabo - Naranjal F01",
        "Guabo - Naranjal F02","Guabo - Sta Isabel","Ponce Enrriquez - Pucará",
        "Balao - Tenguel","Huaquillas - Santa Rosa","Arenillas - Balsas",
        "Guabo - Huaquillas","Piñas - Atahualpa","Zaruma - Atahualpa","Piñas - Zaruma",
        "Zaruma - Chaguarpamba","Balsas - Piñas","Balsas - La Avanzada",
        "Balsas . Olmedo","Balsas - Alamor",
    ],
    "TS.QUEVEDO": [
        "CARACOL - MANGA DE DERIVACIÓN CALUMA F01","CARACOL - RICAURTE F01",
        "BABAHOYO – JUJAN F02","BABAHOYO – SIMÓN BOLÍVAR F1","BABAHOYO – SIMÓN BOLÍVAR F02",
        "BABAHOYO – BALSAPAMBA F01","PUEBLOVIEJO – RICAURTE F01","BALZAR - OLMEDO F01",
        "LA CATORCE - LA BRAMADORA F01","SANTA MARIA - CHONE CHOFERES F01",
        "MONTALVO – CALUMA F01","RICAURTE – CALUMA F01","ECHEANDIA - CALUMA F01",
        "EL EMPALME - PUEBLO NUEVO F01","LAS NAVES - ECHEANDIA F01",
        "BABAHOYO – JUJAN F01","LA MANA - MORASPUNGO F01","LA MANA - ZUMBAHUA F01",
        "MOCACHE - MANGA DERIVACIÓN SAN CARLOS F01","MONTALVO - FEBRES CORDERO F01",
        "PALESTINA - COLIMES PRINCIPAL F01","PALESTINA - COLIMES RESPALDO F02",
        "MORASPUNGO - FACUNDO VELA F01","PICHINCHA - SAN SEBASTIÁN F01",
        "VINCES – PALENQUE F01","VINCES - PALESTINA F01",
        "PUEBLO NUEVO - MANGA DERIVACIÓN 4 MANGAS F01","QUINSALOMA - LAS NAVES F01",
        "QUINSALOMA - MORASPUNGO F01","BABAHOYO - BALSAPAMBA F02",
        "LA UNIÓN - SANTA MARIA DEL TOACHI F01","MOCACHE - MANGA DERIVACIÓN QUEVEDO F01",
        "BABAHOYO - PALESTINA F02","QUEVEDO - PATRICIA PILAR PRINCIPAL F01",
        "QUEVEDO – BABAHOYO F02","QUEVEDO - LA MANA F02","QUEVEDO - PALESTINA F01",
        "QUEVEDO - PATRICIA PILAR RESPALDO F03","QUEVEDO - PATRICIA PILAR RESPALDO F02",
        "QUEVEDO - PATRICIA PILAR PRINCIPAL F04","QUEVEDO - PICHINCHA F01",
        "QUEVEDO - PICHINCHA F02","QUEVEDO – VALENCIA F01","QUEVEDO - VENTANAS F01",
        "QUINSALOMA - MANGA DERIVACIÓN LA ERCILIA F01","SAN CARLOS – VALENCIA F01",
        "QUEVEDO - PALESTINA F02","VALENCIA – LA MANA F01","VENTANAS – BABAHOYO F01",
        "EL VERGEL – MANGA DERIVACIÓN LA UNIÓN F01","EL VERGEL – MANGA DERIVACIÓN QUEVEDO F01",
        "LA CATORCE - SANTA MARÍA F01","VINCES – ANTONIO SOTOMAYOR PRINCIPAL F01",
        "VINCES – ANTONIO SOTOMAYOR RESPALDO F02","BABAHOYO - VINCES F01",
        "VINCES - SAN JUAN F01",
    ],
    "TS.LAGO AGRIO": [
        "BAEZA - ARCHIDONA","BAEZA - EL CHACO","EL CHACO - REVENTADOR",
        "LAGO AGRIO - SHUSHUFINDI","GENERAL FARFAN - LAGO AGRIO","LAGO AGRIO - GENERAL FARFAN",
        "LAGO AGRIO - JIVINO VERDE","TARAPOA - LAGO AGRIO","LUMBAQUI - LAGO AGRIO",
        "TARAPOA - PUERTO EL CARMEN","LUMBAQUI - LA BONITA","REVENTADOR - LUMBAQUI",
        "SHUSHUFINDI - JOYA DE LOS SACHAS","JIVINO VERDE - SHUSHUFINDI",
        "NODO SHUSHUFINDI - NODO SHUAR","BAEZA - PIFO",
    ],
    "TS.SANTO DOMINGO": [
        "Mendoza - Bramadora","Santo Domingo - Patricia Pilar Main",
        "Santo Domingo - Patricia Pilar BCK","La Concordia-Mendoza",
        "La Concordia-San Jacinto de Bua","Santo Domingo - Mendoza",
        "La Concordia-La Union","Santo Domingo - Los Bancos","Tandapi - Santo Domingo",
        "Los Bancos-Pedro Vicente Maldonado","Valle Hermoso - Los Bancos","Los Bancos-Mindo",
        "BNC-Mindo (E)/Los Bancos (O)-Nanegalito (O)-F02-BB",
        "Luz de America - Puerto Limon","Luz de America - Santa Maria del Toachi",
        "Mendoza-Pedernales-F02","Mendoza-Pedernales-F01","Nanegalito-Pacto",
        "Puerto Limon - Nuevo Israel","Nuevo Israel - San Jacinto",
        "Puerto Quito - Pedro Vicente Maldonado",
        "Pedro Vicente Maldonado - Las Golondrinas","Santo Domingo - La Concordia",
        "Santo Domingo - Chiriboga","Tandapi - Aloag","Santo Domingo - Valle Hermoso",
        "Mendoza-Flavio Alfaro","Puerto Quito - La Concordia",
        "Puerto Quito - Las Golondrinas","Patricia Pilar - Santa Maria del Toachi",
    ],
    "TS.MANTA": [
        "Santa Ana - Ayacucho","Ayacucho - San Sebastian","Bahia - San Jacinto",
        "Bahia - Tosagua","Bahia - San Vicente","Bellavista - Olmedo",
        "San Sebastian - Sucre","Canoa - Jama","Canoa - San Vicente",
        "Charapoto - Crucita","Charapoto - Rocafuerte","Charapoto - San Jacinto",
        "Rocafuerte - Portoviejo main","Crucita - Charapoto BK","Crucita - Rocafuerte Main",
        "Manta - Rocafuerte","Flavio Alfaro - San Isidro","Flavio Alfaro - Chone",
        "Flavio Alfaro - Manga del Mono","Jama - La Cabuya","Portoviejo - Junin",
        "Los Bajos - Montecristi","Chone - Tosagua","Chone - Santa Maria",
        "Chone - San Isidro","Manta - Portoviejo Main","Manta - Rocafuerte",
        "Manta - Montecristi","Manta - Puerto Lopez","Jipijapa - Montecristi",
        "Jipijapa - Pajan main","Puerto Lopez - Jipijapa","Jipijapa - Pedro Carbo",
        "Jipijapa - Montecristi","Jipijapa - Sucre","Portoviejo - Santa Ana",
        "La Entrada - San Lorenzo","Rocafuerte - Tosagua","Portoviejo - San Sebastian",
        "Sucre - Santa Ana","Jipijapa - Sucre",
    ],
    "TS.ESMERALDAS": [
        "Tonsupa - El Salto 24H","Tonsupa - El Salto 48H","Pedernales - El Carmen 24H",
        "Pedernales - El Carmen 48H","Borbon - La Tola F01","Borbon - Lagarto 48H",
        "San Lorenzo - Lagarto 24H","Chamanga - El Salto 24H","Chamanga - El Salto 48H",
        "Borbon - La Tola F02","Rio Verde - Tachina 24H","San Lorenzo - Borbon 24H",
        "San Lorenzo - Lita 48H","Esmeraldas - Tachina 24H F01","Esmeraldas - Tachina 48H F02",
        "Esmeraldas - Tonsupa 24H","Esmeraldas - Tonsupa 48H","Esmeraldas - Viche 24H",
        "Esmeraldas - Viche 48H","Muisne - El Salto 24H","Muisne - Tonhigue 48H",
        "Pedernales - Chamanga 48H","Pedernales - Cojimies 48H F01",
        "Pedernales - Cojimies 48H F02","Pedernales - Chamanga 24H","Pedernales - Jama 24H",
        "Valle Alto - Golondrinas 48H","Valle Alto - La Union 24H",
        "Valle Alto - Puerto Quito 48H","Valle Alto - Viche 48H","Valle Alto - Viche 24H",
        "Lagarto - Rio Verde 24H","San Lorenzo - Lagarto 24H","San Lorenzo - Lita 24H",
        "Pedro Vicente - Golondrinas",
    ],
    "TG.AMBATO": [
        "LATACUNGA-GOSSEAL(Chasqui 49Km)","LATACUNGA - AMBATO","LATACUNGA- SALCEDO",
        "LATACUNGA- PUJILI/ BCK","LATACUNGA - PUJILI /PRINCIPAL","PUJILI - ZUMBAHUA",
        "ZUMBAHUA-LA MANA (Pilalo)","PUJILI - SAQUISILI","CUICUNO - LASSO",
        "CUICUNO - SIGCHOS","CUSUBAMBA - CUNCHIBAMBA","PUJILI-CUSUBAMBA",
        "SALCEDO-CUSUBAMBA","ZUMBAHUA-SIGCHOS","AMBATO-PELILEO PRINCIPAL",
        "AMBATO -PELILEO R2 BACKUP","PELILEO-BAÑOS","BAÑOS-PUYO(RIO NEGRO)",
        "SALCEDO-PILLARO","PATATE - BAÑOS","PELILEO-PATATE","PATATE-PILLARO",
        "AMBATO -PILLARO","AMBATO-SAN FRANCISCO","SAN FRANCISCO-CEVALLOS",
        "CEVALLOS-QUERO","QUERO-MOCHA","MOCHA - TISALEO","TISALEO- SANTA ROSA",
        "SANTA ROSA-AMBATO","PELILEO-HUAMBALO F01","PELILEO-HUAMBALO F02",
        "HUAMBALO-PENIPE","SANTA ROSA-PILAHIUN","PILAHUIN-SIMIATUG(EL SALADO)",
        "AMBATO-CUNCHIBAMBA","AMBATO-RIOBAMBA(Urbina 35KM)","SAQUISILI - CUICUNO",
    ],
    "TS.LOJA": [
        "Alamor-Celica","Amaluza-Jimbura","Quilanga-Amaluza","Cariamanga-Sozoranga",
        "Cariamanga-Lucero","Gonzanama_Cariamanga","Yamana-Casanga","San Pedro-Catacocha",
        "Catacocha-Celica","Catamayo-El Tambo","TelepuertoLoja-Catamayo",
        "Catamayo-SanPedro","El Cisne-Chataco","Yantzaza-Zamora","San Pedro-El Cisne",
        "Gonzanama-El Tambo","El Tambo-Malacatos","Gonzanama-Quilanga",
        "Jimbura-San Andres","Los Encuentros-Lundingold","Macara-Sozoranga",
        "Zapotillo-Macara","Telepuerto Loja-Malacatos","Malacatos-Vilcabamba",
        "Teleouerto Loja-Olmedo","Palanda-Zumba","Palanda-Valladolid",
        "Zumbi-Paquisha","Paquisha-Los Encuentros","Pindal-Zapotillo","Alamor-Pindal",
        "San Andres-Zumba","Oña-Telepuerto Loja","Telepuerto Loja-Zamora",
        "Yangana-Valladolid","Vilcabamba-Yangana","Zumbi-Yacuambi",
        "Yantzaza-Gualaquiza","El Pangui-Tundayme",
    ],
}

ZONAS_ORDEN = [
    "UIO","GYE","Ibarra","TS.MILAGRO","TS.RIOBAMBA","TS.EL COCA","TS.SALINAS",
    "TS.CUENCA","TS.MACHALA","TS.QUEVEDO","TS.LAGO AGRIO","TS.SANTO DOMINGO",
    "TS.MANTA","TS.ESMERALDAS","TG.AMBATO","TS.LOJA",
]

# ══════════════════════════════════════════════════════════════════════════════
# CUADRILLAS POR ZONA (valores únicos, sin repetidos)
# ══════════════════════════════════════════════════════════════════════════════
CUADRILLAS_POR_ZONA = {
    "UIO": [
        "FO UIO INT 01","FO UIO INT 02","FO UIO INT 04",
        "FO UIO INT 05","FO UIO INT 06","FO UIO INT 10",
        "FO UIO INT 12","FO UIO URB 05",
    ],
    "GYE": [
        "GYE FO INT 1","GYE FO INT 2","GYE FO INT 3",
        "GYE FO INT 4","GYE FO INT 5",
    ],
    "Ibarra": [
        "IBA PROV IBA INT 01","IBA PROV CAY INT 01",
        "IBA PROV ANG INT 01","IBA PROV SGA INT 01",
    ],
    "TS.MILAGRO": [
        "MIL PROV MIL INT 01","MIL PROV MIL INT 02","MIL PROV MIL INT 03",
        "MIL PROV TRI INT 01","MIL PROV NAR INT 01",
        "MIL PROV DAU INT 01","MIL PROV DAU INT 02",
    ],
    "TS.RIOBAMBA": [
        "RIO PROV ALA INT 01","RIO PROV PTZ INT 01","RIO PROV TEN INT 01",
        "RIO PROV RIO INT 01","RIO PROV RIO INT 02","RIO PROV GDA INT 01",
    ],
    "TS.EL COCA": [
        "COC PROV LRT INT 01","COC PROV JDS INT 01","COC PROV COC INT 01",
    ],
    "TS.SALINAS": [
        "STL PROV SLN INT 01","STL PROV SLN INT 02","STL PROV ENT INT 01",
        "STL PROV PRO INT 01","STL PROV PLY INT 01",
    ],
    "TS.CUENCA": [
        "CUE PROV ZHU INT 01","CUE PROV GLT INT 01",
        "CUE PROV CUE INT 01","CUE PROV CUE INT 02",
        "CUE PROV CUE INT 03","CUE PROV CUE INT 04",
        "CUE PROV ONA INT 01","CUE PROV STA INT 01","CUE PROV STA INT 02",
        "CUE PROV SUC INT 01","CUE PROV SUC INT 02","CUE PROV GQZ INT 01",
    ],
    "TS.MACHALA": [
        "MCH PROV MCH INT 01","MCH PROV GBO INT 01","MCH PROV HUA INT 01",
        "MCH PROV PIN INT 01","MCH PROV BLS INT 01",
    ],
    "TS.QUEVEDO": [
        "QVD PROV VNC INT 01","QVD PROV BBY INT 01",
        "QVD PROV PAL INT 01","QVD PROV PAL INT 02",
        "QVD PROV LCT INT 01","QVD PROV QVD INT 01","QVD PROV QVD INT 02",
        "QVD PROV VTN INT 01","QVD PROV LMN INT 01","QVD PROV LMN INT 02",
        "QVD PROV PCH INT 01",
    ],
    "TS.LAGO AGRIO": [
        "LAG PROV CHA INT 01","LAG PROV SHU INT 01",
        "LAG PROV LUM INT 01","LAG PROV TRP INT 01",
    ],
    "TS.SANTO DOMINGO": [
        "STO PROV CAR INT 01","STO PROV STO INT 01","STO PROV STO INT 02",
        "STO PROV STO INT 03","STO PROV STO INT 04","STO PROV CON INT 01",
        "STO PROV TAN INT 01","STO PROV BAN INT 01",
        "STO PROV LDA INT 01","STO PROV PTQ INT 01",
    ],
    "TS.MANTA": [
        "MNT PROV PVO INT 01","MNT PROV PVO INT 02","MNT PROV BHI INT 01",
        "MNT PROV PAJ INT 01","MNT PROV JAM INT 01","MNT PROV MNT INT 01",
        "MNT PROV FLV INT 01","MNT PROV CHO INT 01",
        "MNT PROV JIP INT 01","MNT PROV PLZ INT 01",
    ],
    "TS.ESMERALDAS": [
        "ESM PROV ESM INT 01","ESM PROV ESM INT 02",
        "ESM PROV PED INT 01","ESM PROV PED INT 02",
        "ESM PROV SL INT 01","ESM PROV SL INT 02","ESM PROV TCH INT 01",
        "ESM PROV QUI INT 01","ESM PROV QUI INT 02",
    ],
    "TG.AMBATO": [
        "AMB PROV LAT INT 01","AMB PROV PUJ INT 01","AMB PROV PUJ INT 02",
        "AMB PROV AMB INT 01","AMB PROV AMB INT 02","AMB PROV BAN INT 01",
    ],
    "TS.LOJA": [
        "LOJ PROV CEL INT 01","LOJ PROV AMA INT 01","LOJ PROV CAR INT 01",
        "LOJ PROV CHA INL 01","LOJ PROV LOJ INT 01","LOJ PROV ZAM INL 01",
        "LOJ PROV VIL INT 01","LOJ PROV YAN INL 01","LOJ PROV MAC INT 01",
        "LOJ PROV ZUM INT 01","LOJ PROV ZAP INT 01","LOJ PROV SAR INT 01",
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
# CAUSAS Y OBSERVACIONES PREDEFINIDAS
# ══════════════════════════════════════════════════════════════════════════════
CAUSAS = [
    "Accidente de tránsito",
    "Camión con carga elevada",
    "Maquinaria pesada (retroexcavadora, volqueta, etc.)",
    "Trabajos por terceras personas",
    "Robo / vandalismo",
    "Deslave",
    "Caída de árboles / ramas",
    "Inundación",
    "Incendio forestal",
    "Vientos fuertes",
    "Incendio por corto circuito",
    "Cambio de poste",
    "Mordida por roedores",
    "Manipulación indebida de la red",
    "Recogimiento de hilos en mangas",
    "Atenuaciones multiplex en el tramo",
    "Daños en empalmes",
    "✏️ Escribir manualmente",
]

OBSERVACIONES_PREDEFINIDAS = {
    "Accidente de tránsito":
        "Se evidencia impacto vehicular sobre infraestructura de soporte (poste/canalización), provocando corte total o parcial del cable de fibra óptica y pérdida de continuidad en el tramo.",
    "Camión con carga elevada":
        "Se observa tensión mecánica por arrastre de cable debido a paso de vehículo con altura fuera de norma, generando tensión y microcurvaturas o ruptura de la fibra.",
    "Maquinaria pesada (retroexcavadora, volqueta, etc.)":
        "Daño directo por excavación o movimiento de tierra sin señalización previa, ocasionando corte total del cable o afectación a ductos y canalización subterránea.",
    "Trabajos por terceras personas":
        "Intervención no autorizada sobre la red, generando cortes, desconexiones o manipulación indebida de la infraestructura de fibra óptica.",
    "Robo / vandalismo":
        "Sustracción o daño intencional del cableado, dejando fibras expuestas, cortes múltiples o pérdida total del servicio en el tramo afectado.",
    "Deslave":
        "Movimiento de tierra que provoca desplazamiento o ruptura de la infraestructura, afectando la continuidad del cable y generando alta atenuación o corte.",
    "Caída de árboles / ramas":
        "Impacto directo sobre el tendido aéreo, generando esfuerzo mecánico, deformación del cable o ruptura de fibras ópticas.",
    "Inundación":
        "Ingreso de agua en cámaras, ductos o mangas, provocando deterioro de empalmes, incremento de atenuación y posible degradación del servicio.",
    "Incendio forestal":
        "Exposición del cable a altas temperaturas, causando daño en la chaqueta, debilitamiento estructural y posible afectación del núcleo de fibra.",
    "Vientos fuertes":
        "Oscilación excesiva del cable aéreo que genera fatiga mecánica, microcurvaturas y aumento progresivo de la atenuación.",
    "Incendio por corto circuito":
        "Daño térmico ocasionado por falla eléctrica cercana, afectando la integridad del cable y provocando pérdida de señal en el tramo.",
    "Cambio de poste":
        "Intervención en infraestructura que genera manipulación del cable, pudiendo ocasionar microcurvaturas, tensión indebida o desconexión accidental.",
    "Mordida por roedores":
        "Deterioro de la cubierta del cable por acción biológica, dejando fibras expuestas y vulnerables a ruptura o incremento de atenuación.",
    "Manipulación indebida de la red":
        "Intervenciones no técnicas que provocan desorden en la fibra, desconexiones, curvaturas excesivas o afectación en empalmes.",
    "Recogimiento de hilos en mangas":
        "Mala organización interna en la manga de empalme, generando microcurvaturas y aumento de atenuación en las fibras afectadas.",
    "Atenuaciones multiplex en el tramo":
        "Presencia de pérdidas distribuidas a lo largo del enlace, posiblemente asociadas a microcurvaturas, empalmes defectuosos o degradación del cable.",
    "Daños en empalmes":
        "Empalmes con pérdidas elevadas por mala fusión, contaminación o deterioro, afectando la calidad de señal y continuidad del servicio.",
}

def _kb(opciones, columnas=2):
    """Crea ReplyKeyboardMarkup con las opciones dadas en N columnas."""
    filas = [opciones[i:i+columnas] for i in range(0, len(opciones), columnas)]
    return ReplyKeyboardMarkup(filas, one_time_keyboard=True, resize_keyboard=True)

KB_CAUSAS = _kb(CAUSAS + ["🔙 Menú"], columnas=1)

# ══════════════════════════════════════════════════════════════════════════════
# REMEDIOS PREDEFINIDOS
# ══════════════════════════════════════════════════════════════════════════════
REMEDIO_CON_METROS = "Se ejecuta reposición de tramo de fibra óptica"   # señal para pedir metros

REMEDIOS = [
    "Se ejecuta reposición de tramo de fibra óptica",
    "Se efectúa empalme por fusión en puntos de corte",
    "Se procede a la Re fusión de empalmes defectuosos",
    "Se ejecuta reubicación de red sobre nueva infraestructura",
    "Se recorre reserva para solventar la novedad y generación de manga",
    "Se efectúa limpieza, inspección y reconectorización en ODF",
    "Se implementa protección mecánica adicional mediante canalización/tubería",
    "✏️ Escribir otra solución",
]

TEXTOS_REMEDIO = {
    "Se ejecuta reposición de tramo de fibra óptica":
        "Se ejecuta reposición de tramo de fibra óptica mediante tendido nuevo y empalmes por fusión, garantizando niveles de atenuación dentro de parámetros operativos.",
    "Se efectúa empalme por fusión en puntos de corte":
        "Se efectúa empalme por fusión en puntos de corte, con medición y certificación de pérdidas ópticas para restablecimiento del servicio.",
    "Se procede a la Re fusión de empalmes defectuosos":
        "Se procede a la Re fusión de empalmes defectuosos, corrigiendo niveles de atenuación fuera de umbral.",
    "Se ejecuta reubicación de red sobre nueva infraestructura":
        "Se ejecuta reubicación de red sobre nueva infraestructura, mitigando riesgos y asegurando continuidad del servicio.",
    "Se recorre reserva para solventar la novedad y generación de manga":
        "Se recorre reserva para solventar la novedad y generación de manga.",
    "Se efectúa limpieza, inspección y reconectorización en ODF":
        "Se efectúa limpieza, inspección y reconectorización en ODF, garantizando niveles ópticos adecuados.",
    "Se implementa protección mecánica adicional mediante canalización/tubería":
        "Se implementa protección mecánica adicional mediante canalización/tubería para prevenir recurrencia del daño.",
}

KB_REMEDIOS = _kb(REMEDIOS + ["🔙 Menú"], columnas=1)

# SI_NO, OMITIR y LISTO se definen junto al menú principal (más abajo)

# ══════════════════════════════════════════════════════════════════════════════
# AUTENTICACION
# ══════════════════════════════════════════════════════════════════════════════
async def start(u, c):
    uid = u.effective_user.id
    # Si ya tiene sesion activa, saltar directamente al reporte
    if uid in _SESIONES:
        c.user_data.clear()
        c.user_data["auth_email"]  = _SESIONES[uid]["email"]
        c.user_data["auth_nombre"] = _SESIONES[uid]["nombre"]
        await u.message.reply_text(
            f"👤 Hola {_SESIONES[uid]['nombre']}, sesión activa.\n\n"
            "Bot Incidencias Fibra Óptica\n"
            "Escribe /cancelar para salir.")
        return await mostrar_menu(u, c)

    c.user_data.clear()
    await u.message.reply_text(
        "🔒 Debes autenticarte antes de poder interactuar.\n"
        "Ingresa tu correo y el código de 6 dígitos\n"
        "de *FortiToken Mobile* con este formato:\n\n"
        "email: tucorreo@telconet.ec\n"
        "totp: 123456",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove())
    return AUTH

async def p_auth(u, c):
    """Valida correo @telconet.ec + TOTP. Bloquea tras 3 intentos fallidos."""
    texto = u.message.text.strip()
    uid   = u.effective_user.id

    # Inicializar contador de intentos
    if "auth_intentos" not in c.user_data:
        c.user_data["auth_intentos"] = 0

    MAX_INTENTOS = 3

    # Parsear el mensaje
    email = None
    totp  = None
    for linea in texto.splitlines():
        l = linea.strip().lower()
        if l.startswith("email:"):
            email = l.split(":", 1)[1].strip()
        elif l.startswith("totp:"):
            totp = linea.split(":", 1)[1].strip()

    def _intento_fallido(motivo):
        """Suma intento y devuelve el mensaje con contador."""
        c.user_data["auth_intentos"] += 1
        intentos = c.user_data["auth_intentos"]
        restantes = MAX_INTENTOS - intentos
        if intentos >= MAX_INTENTOS:
            return None   # señal de bloqueo
        return (f"🚫 **ACCESO DENEGADO.** No tienes autorización. "
                f"(Intento {intentos}/{MAX_INTENTOS})\n\n"
                f"{motivo}\n\n"
                f"Intentos restantes: {restantes}\n\n"
                "email: tucorreo@ejemplo.com\n"
                "totp: ")

    # Formato incorrecto
    if not email or not totp:
        msg = _intento_fallido("⚠️ Formato incorrecto. Usa exactamente el formato indicado.")
        if msg is None:
            logger.warning(f"BLOQUEADO (formato): TG={uid}")
            await u.message.reply_text(
                f"🚫 **ACCESO DENEGADO.** No tienes autorización. "
                f"(Intento {MAX_INTENTOS}/{MAX_INTENTOS})\n\n"
                "Has superado el numero de intentos permitidos.\n"
                "Escribe /start para reiniciar.",
                reply_markup=ReplyKeyboardRemove())
            return ConversationHandler.END
        await u.message.reply_text(msg); return AUTH

    # Dominio obligatorio @telconet.ec
    if not email.endswith("@telconet.ec"):
        msg = _intento_fallido("\u274c Solo se permiten correos @telconet.ec.\n"
                               "Ejemplo: lpilacuan@telconet.ec")
        if msg is None:
            logger.warning(f"BLOQUEADO (dominio): {email} TG={uid}")
            await u.message.reply_text(
                f"\U0001f6ab **ACCESO DENEGADO.** No tienes autorizaci\u00f3n. "
                f"(Intento {MAX_INTENTOS}/{MAX_INTENTOS})\n\n"
                "Has superado el numero de intentos permitidos.\n"
                "Escribe /start para reiniciar.",
                reply_markup=ReplyKeyboardRemove())
            return ConversationHandler.END
        await u.message.reply_text(msg); return AUTH

    # Usar la parte antes del @ como nombre
    nombre = email.split("@")[0]

    # Validar código TOTP de FortiToken Mobile
    if PYOTP_OK and TOTP_SECRET_GLOBAL:
        valido = pyotp.TOTP(TOTP_SECRET_GLOBAL).verify(totp.strip(), valid_window=1)
    else:
        valido = False
        logger.error("TOTP no disponible — verifica variable TOTP_SECRET en Railway y que pyotp esté instalado")

    if not valido:
        msg = _intento_fallido("❌ Código FortiToken incorrecto o expirado.\n"
                               "Abre FortiToken Mobile y usa el código actual de 6 dígitos.")
        if msg is None:
            logger.warning(f"BLOQUEADO (password): {email} TG={uid}")
            await u.message.reply_text(
                f"🚫 **ACCESO DENEGADO.** No tienes autorización. "
                f"(Intento {MAX_INTENTOS}/{MAX_INTENTOS})\n\n"
                "Has superado el numero de intentos permitidos.\n"
                "Escribe /start para reiniciar.",
                reply_markup=ReplyKeyboardRemove())
            return ConversationHandler.END
        await u.message.reply_text(msg); return AUTH

    # ✅ Sesion iniciada
    _SESIONES[uid] = {"email": email, "nombre": nombre}
    c.user_data["auth_email"]    = email
    c.user_data["auth_nombre"]   = nombre
    c.user_data["auth_intentos"] = 0
    logger.info(f"LOGIN OK: {email} — {nombre} (TG={uid})")

    await u.message.reply_text(
        f"✅ Hola {nombre}, acceso concedido.\n\n"
        "Bot Incidencias Fibra Óptica\n"
        "Escribe /cancelar para salir.")
    return await mostrar_menu(u, c)


# ── Pasos de texto ─────────────────────────────────────────────────────────────
async def _p(u, c, key, msg, estado):
    c.user_data[key] = u.message.text.strip()
    await u.message.reply_text(msg, reply_markup=KB_MENU)
    return estado

# ── Validadores ───────────────────────────────────────────────────────────────
def _validar_codigo(texto):
    """Acepta formatos: 20220621-0012  o  20220621/0012  o  20220621_0012"""
    return bool(re.fullmatch(r"\d{8}[-/_ ]\d{4}", texto))

def _validar_fecha(texto):
    """DD/MM/AAAA — también acepta MM-DD-AAAA y MM.DD.AAAA"""
    t = texto.replace("-","/").replace(".","/")
    try:
        datetime.strptime(t, "%d/%m/%Y")
        return t   # devuelve normalizado con /
    except ValueError:
        return None

def _validar_hora(texto):
    """HH:MM — también acepta H:MM y HH.MM"""
    t = texto.replace(".",":").strip()
    if re.fullmatch(r"\d{1,2}:\d{2}", t):
        try:
            h, m = t.split(":")
            if 0 <= int(h) <= 23 and 0 <= int(m) <= 59:
                return f"{int(h):02d}:{int(m):02d}"
        except ValueError:
            pass
    return None

# ── Pasos de texto ─────────────────────────────────────────────────────────────
async def _p(u, c, key, msg, estado):
    c.user_data[key] = u.message.text.strip()
    await u.message.reply_text(msg, reply_markup=KB_MENU)
    return estado

async def p_codigo(u,c):
    c.user_data["codigo"] = u.message.text.strip()
    d = c.user_data
    if c.user_data.get("_editing_campo") == "codigo":
        await u.message.reply_text(f"✅ Código actualizado: {d['codigo']}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, d, c)
    await u.message.reply_text(
        "✅ Código registrado.\n\n"
        "Paso 3/16 — Fecha de la incidencia\n"
        "Escribe en formato DD/MM/AAAA\n"
        "📅 Ejemplo: 21/06/2022", reply_markup=KB_MENU)
    return FECHA_INC

async def p_cuadrilla_zona(u, c):
    """Ahora solo recibe texto de fallback; la selección normal viene por callback inline."""
    zona = u.message.text.strip()
    if zona not in CUADRILLAS_POR_ZONA:
        await u.message.reply_text(
            "Selecciona la ZONA de la cuadrilla:",
            reply_markup=_ikb_zonas("cz"))
        return CUADRILLA_ZONA
    # Si llegó texto válido, procesar igual
    c.user_data["cuadrilla_zona"] = zona
    cuads = CUADRILLAS_POR_ZONA.get(zona, [])
    kb = _ikb_lista(cuads, "cs", extra=[("✏️ Escribir manualmente", "cs:m")])
    await u.message.reply_text(f"Zona: {zona}\n\nSelecciona la cuadrilla:", reply_markup=kb)
    return CUADRILLA_SEL

async def cb_cuadrilla_zona(upd, c):
    """Callback: usuario tocó una zona en el inline keyboard."""
    q = upd.callback_query; await q.answer()
    idx = int(q.data.split(":")[1])
    zona = ZONAS_ORDEN[idx]
    c.user_data["cuadrilla_zona"] = zona
    cuads = CUADRILLAS_POR_ZONA.get(zona, [])
    kb = _ikb_lista(cuads, "cs", extra=[("✏️ Escribir manualmente", "cs:m")])
    await q.edit_message_text(f"Zona: {zona}\n\nSelecciona la cuadrilla:", reply_markup=kb)
    return CUADRILLA_SEL

async def p_cuadrilla_sel(u, c):
    """Recibe cuadrilla escrita manualmente."""
    cuadrilla = u.message.text.strip()
    c.user_data["cuadrilla"] = cuadrilla
    if c.user_data.get("_editing_campo") == "cuadrilla":
        await u.message.reply_text(f"✅ Cuadrilla actualizada: {cuadrilla}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        f"✅ Cuadrilla: {cuadrilla}\n\nNúmero de Caso\nEjemplo: 20220621-0012B", reply_markup=KB_MENU)
    return CODIGO

async def cb_cuadrilla_sel(upd, c):
    """Callback: usuario tocó una cuadrilla o 'escribir manualmente'."""
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    if data == "m":
        await q.edit_message_text("✏️ Escribe el nombre exacto de la cuadrilla:")
        return CUADRILLA_SEL
    zona = c.user_data.get("cuadrilla_zona","")
    cuads = CUADRILLAS_POR_ZONA.get(zona, [])
    cuadrilla = cuads[int(data)]
    c.user_data["cuadrilla"] = cuadrilla
    if c.user_data.get("_editing_campo") == "cuadrilla":
        await q.edit_message_text(f"✅ Cuadrilla actualizada: {cuadrilla}")

        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)

        return await _volver_historial(_FM(q.message.chat_id, q.bot).reply_text,
                                       c.user_data, c)
    await q.edit_message_text(f"✅ Cuadrilla: {cuadrilla}\n\nNúmero de Caso\nEjemplo: 20220621-0012B")
    return CODIGO

async def p_fecha_inc(u,c):
    f = _validar_fecha(u.message.text.strip())
    if not f:
        await u.message.reply_text(
            "❌ Fecha inválida.\n\n"
            "Usa el formato: DD/MM/AAAA\n"
            "📅 Ejemplo: 21/06/2022\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return FECHA_INC
    c.user_data["fecha_inc"] = f
    await u.message.reply_text(
        f"✅ Fecha de incidencia: {f}\n\n"
        "Paso 2 de 2 — Ahora escribe la HORA:\n"
        "Formato: HH:MM  (24 horas)\n"
        "🕐 Ejemplo: 08:30", reply_markup=KB_MENU)
    return HORA_INC

async def p_hora_inc(u,c):
    h = _validar_hora(u.message.text.strip())
    if not h:
        await u.message.reply_text(
            "❌ Hora inválida.\n\n"
            "Usa el formato: HH:MM (24 horas)\n"
            "🕐 Ejemplo: 08:30\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return HORA_INC
    c.user_data["hora_inc"] = h
    if c.user_data.get("_editing_campo") == "fecha_inc":
        await u.message.reply_text(f"✅ Hora de incidencia registrada: {h}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "✅ Hora registrada.\n\n"
        "Paso 4/16 — Fecha de arribo al sitio\n"
        "Usa el formato: DD/MM/AAAA\n"
        "📅 Ejemplo: 21/06/2022", reply_markup=KB_MENU)
    return FECHA_ARR

async def p_fecha_arr(u,c):
    f = _validar_fecha(u.message.text.strip())
    if not f:
        await u.message.reply_text(
            "❌ Fecha inválida.\n\n"
            "Usa el formato: DD/MM/AAAA\n"
            "📅 Ejemplo: 21/06/2022\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return FECHA_ARR
    c.user_data["fecha_arr"] = f
    await u.message.reply_text(
        f"✅ Fecha de arribo: {f}\n\n"
        "Paso 2 de 2 — Ahora escribe la HORA:\n"
        "Formato: HH:MM  (24 horas)\n"
        "🕐 Ejemplo: 10:45", reply_markup=KB_MENU)
    return HORA_ARR

async def p_hora_arr(u,c):
    h = _validar_hora(u.message.text.strip())
    if not h:
        await u.message.reply_text(
            "❌ Hora inválida.\n\n"
            "Usa el formato: HH:MM (24 horas)\n"
            "🕐 Ejemplo: 10:45\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return HORA_ARR
    c.user_data["hora_arr"] = h
    if c.user_data.get("_editing_campo") == "fecha_arr":
        await u.message.reply_text(f"✅ Hora de arribo registrada: {h}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "✅ Hora registrada.\n\n"
        "Paso 6/16 — Fecha inicio de reparación\n"
        "Usa el formato: DD/MM/AAAA\n"
        "📅 Ejemplo: 21/06/2022", reply_markup=KB_MENU)
    return FECHA_REP

async def p_fecha_rep(u,c):
    f = _validar_fecha(u.message.text.strip())
    if not f:
        await u.message.reply_text(
            "❌ Fecha inválida.\n\n"
            "Usa el formato: DD/MM/AAAA\n"
            "📅 Ejemplo: 21/06/2022\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return FECHA_REP
    c.user_data["fecha_rep"] = f
    await u.message.reply_text(
        f"✅ Fecha de reparación: {f}\n\n"
        "Paso 2 de 2 — Ahora escribe la HORA:\n"
        "Formato: HH:MM  (24 horas)\n"
        "🕐 Ejemplo: 11:00", reply_markup=KB_MENU)
    return HORA_REP

async def p_hora_rep(u,c):
    h = _validar_hora(u.message.text.strip())
    if not h:
        await u.message.reply_text(
            "❌ Hora inválida.\n\n"
            "Usa el formato: HH:MM (24 horas)\n"
            "🕐 Ejemplo: 11:00\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return HORA_REP
    c.user_data["hora_rep"] = h
    if c.user_data.get("_editing_campo") == "fecha_rep":
        await u.message.reply_text(f"✅ Hora de reparación registrada: {h}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "✅ Hora registrada.\n\n"
        "Paso 8/16 — Fecha de resolución\n"
        "Usa el formato: DD/MM/AAAA\n"
        "📅 Ejemplo: 21/06/2022", reply_markup=KB_MENU)
    return FECHA_RES

async def p_fecha_res(u,c):
    f = _validar_fecha(u.message.text.strip())
    if not f:
        await u.message.reply_text(
            "❌ Fecha inválida.\n\n"
            "Usa el formato: DD/MM/AAAA\n"
            "📅 Ejemplo: 21/06/2022\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return FECHA_RES
    c.user_data["fecha_res"] = f
    await u.message.reply_text(
        f"✅ Fecha de resolución: {f}\n\n"
        "Paso 2 de 2 — Ahora escribe la HORA:\n"
        "Formato: HH:MM  (24 horas)\n"
        "🕐 Ejemplo: 15:30", reply_markup=KB_MENU)
    return HORA_RES

async def p_hora_res(u,c):
    h = _validar_hora(u.message.text.strip())
    if not h:
        await u.message.reply_text(
            "❌ Hora inválida.\n\n"
            "Usa el formato: HH:MM (24 horas)\n"
            "🕐 Ejemplo: 15:30\n\n"
            "Intenta de nuevo:", reply_markup=KB_MENU)
        return HORA_RES
    c.user_data["hora_res"] = h
    if c.user_data.get("_editing_campo") == "fecha_res":
        await u.message.reply_text(f"✅ Hora de resolución registrada: {h}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "Paso 10/16 — Motivo aparente de la incidencia\n\n"
        "Selecciona la zona:",
        reply_markup=_ikb_zonas("mz"))
    return MOTIVO_ZONA

async def p_motivo_zona(u, c):
    zona = u.message.text.strip()
    if zona not in TRAMOS_POR_ZONA:
        await u.message.reply_text("Selecciona la zona:", reply_markup=_ikb_zonas("mz"))
        return MOTIVO_ZONA
    c.user_data["motivo_zona"] = zona
    await u.message.reply_text(f"Zona: {zona}\n\nTipo de incidencia:", reply_markup=IKB_MOTIVO_TIPOS)
    return MOTIVO_TIPO

async def cb_motivo_zona(upd, c):
    q = upd.callback_query; await q.answer()
    zona = ZONAS_ORDEN[int(q.data.split(":")[1])]
    c.user_data["motivo_zona"] = zona
    await q.edit_message_text(f"Zona: {zona}\n\nTipo de incidencia:", reply_markup=IKB_MOTIVO_TIPOS)
    return MOTIVO_TIPO

async def p_motivo_tipo(u, c):
    tipo = u.message.text.strip()
    tipos_validos = _MOTIVO_TIPOS + ["✏️ Escribir motivo manualmente"]
    if tipo == "✏️ Escribir motivo manualmente":
        c.user_data["motivo_tipo"] = ""
        await u.message.reply_text("✏️ Escribe el motivo completo de la incidencia:", reply_markup=KB_MENU)
        return MOTIVO
    if tipo not in tipos_validos:
        await u.message.reply_text("Selecciona el tipo:", reply_markup=IKB_MOTIVO_TIPOS)
        return MOTIVO_TIPO
    c.user_data["motivo_tipo"] = tipo
    zona = c.user_data.get("motivo_zona","")
    tramos = TRAMOS_POR_ZONA.get(zona, [])
    kb = _ikb_lista(tramos, "tr", extra=[("✏️ Escribir tramo manualmente", "tr:m")])
    await u.message.reply_text(f"Tipo: {tipo}\n\nSelecciona el tramo en {zona}:", reply_markup=kb)
    return MOTIVO_TRAMO

async def cb_motivo_tipo(upd, c):
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    if data == "m":
        c.user_data["motivo_tipo"] = ""
        await q.edit_message_text("✏️ Escribe el motivo completo:")
        return MOTIVO
    tipo = _MOTIVO_TIPOS[int(data)]
    c.user_data["motivo_tipo"] = tipo
    zona = c.user_data.get("motivo_zona","")
    tramos = TRAMOS_POR_ZONA.get(zona, [])
    kb = _ikb_lista(tramos, "tr", extra=[("✏️ Escribir tramo manualmente", "tr:m")])
    await q.edit_message_text(f"Tipo: {tipo}\n\nSelecciona el tramo en {zona}:", reply_markup=kb)
    return MOTIVO_TRAMO

async def p_motivo_tramo(u, c):
    tramo = u.message.text.strip()
    tipo = c.user_data.get("motivo_tipo","")
    if tipo:
        c.user_data["motivo"] = f"{tipo}: {tramo}"
    else:
        c.user_data["motivo"] = tramo
    if c.user_data.get("_editing_campo") == "motivo":
        await u.message.reply_text(f"✅ Motivo actualizado: {c.user_data['motivo']}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        f"✅ Motivo: {c.user_data['motivo']}\n\nTestigos del incidente\n\nSelecciona la causa:",
        reply_markup=_ikb_lista(CAUSAS, "ca"))
    return TESTIGOS

async def cb_motivo_tramo(upd, c):
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    if data == "m":
        await q.edit_message_text("✏️ Escribe el nombre del tramo:")
        return MOTIVO_TRAMO
    zona = c.user_data.get("motivo_zona","")
    tramos = TRAMOS_POR_ZONA.get(zona, [])
    tramo = tramos[int(data)]
    tipo = c.user_data.get("motivo_tipo","")
    c.user_data["motivo"] = f"{tipo}: {tramo}" if tipo else tramo
    if c.user_data.get("_editing_campo") == "motivo":
        await q.edit_message_text(f"✅ Motivo actualizado: {c.user_data['motivo']}")

        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)

        return await _volver_historial(_FM(q.message.chat_id, q.bot).reply_text,
                                       c.user_data, c)
    await q.edit_message_text(
        f"✅ Motivo: {c.user_data['motivo']}\n\nTestigos del incidente\n\nSelecciona la causa:",
        reply_markup=_ikb_lista(CAUSAS, "ca"))
    return TESTIGOS

async def p_motivo(u,c):
    texto = u.message.text.strip()
    tipo = c.user_data.get("motivo_tipo", "")
    valor = f"{tipo}: {texto}" if tipo else texto
    errores = _verificar_ortografia(texto)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "motivo", "valor": valor, "errores": errores,
            "next_state": TESTIGOS, "current_state": MOTIVO,
            "next_msg": f"✅ Motivo: {valor[:60]}\n\nTestigos del incidente\n\nSelecciona la causa:",
            "next_markup": _ikb_lista(CAUSAS, "ca"),
        }
        await u.message.reply_text(_msg_spell(valor, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return MOTIVO
    c.user_data["motivo"] = valor
    if c.user_data.get("_editing_campo") == "motivo":
        await u.message.reply_text(f"✅ Motivo actualizado: {valor}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        f"✅ Motivo: {valor}\n\nTestigos del incidente\n\nSelecciona la causa:",
        reply_markup=_ikb_lista(CAUSAS, "ca"))
    return TESTIGOS

async def p_testigos(u,c):
    """Recibe testigos escritos manualmente."""
    texto = u.message.text.strip()
    errores = _verificar_ortografia(texto)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "testigos", "valor": texto, "errores": errores,
            "next_state": REMEDIO, "current_state": TESTIGOS,
            "next_msg": "✅ Testigos registrados.\n\nRemedio definitivo\n\nSelecciona:",
            "next_markup": _ikb_lista(REMEDIOS, "rem"),
        }
        await u.message.reply_text(_msg_spell(texto, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return TESTIGOS
    c.user_data["testigos"] = texto
    if c.user_data.get("_editing_campo") == "testigos":
        await u.message.reply_text(f"✅ Testigos registrados.", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "✅ Testigos registrados.\n\nRemedio definitivo\n\nSelecciona:",
        reply_markup=_ikb_lista(REMEDIOS, "rem"))
    return REMEDIO

async def cb_causas(upd, c):
    """Callback: selección de causa (para testigos y obs)."""
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    campo = c.user_data.get("_causas_campo", "testigos")
    if data == str(len(CAUSAS)-1):  # "✏️ Escribir manualmente"
        await q.edit_message_text("✏️ Escribe el texto:")
        return TESTIGOS if campo == "testigos" else OBS_MANUAL
    causa = CAUSAS[int(data)]
    if campo == "testigos":
        c.user_data["testigos"] = causa
        if c.user_data.get("_editing_campo") == "testigos":
            await q.edit_message_text(f"✅ Testigos: {causa[:60]}")
            class _FM2:
                def __init__(self, chat_id, bot):
                    self.chat_id = chat_id; self.bot = bot
                async def reply_text(self, txt, **kw):
                    await self.bot.send_message(self.chat_id, txt, **kw)
            return await _volver_historial(_FM2(q.message.chat_id, q.bot).reply_text,
                                           c.user_data, c)
        await q.edit_message_text(f"✅ Causa: {causa}\n\nRemedio definitivo\n\nSelecciona:",
            reply_markup=_ikb_lista(REMEDIOS, "rem"))
        return REMEDIO
    else:
        obs_text = OBSERVACIONES_PREDEFINIDAS.get(causa, causa)
        c.user_data["obs"] = obs_text

        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)

        await q.edit_message_text("✅ Observación registrada.")
        return await _volver_historial(_FM(q.message.chat_id, q.bot).reply_text,
                                       c.user_data, c)

async def p_remedio(u, c):
    """Recibe remedio escrito manualmente (fallback)."""
    seleccion = u.message.text.strip()
    texto = TEXTOS_REMEDIO.get(seleccion)
    if texto:
        c.user_data["remedio"] = texto
    else:
        c.user_data["remedio"] = seleccion
    if c.user_data.get("_editing_campo") == "remedio":
        await u.message.reply_text("✅ Remedio registrado.", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "✅ Remedio registrado.\n\nEnlaces caídos\n\nSelecciona la zona:",
        reply_markup=_ikb_zonas("ez"))
    return ENLACES_ZONA

async def cb_remedio(upd, c):
    """Callback: selección de remedio inline."""
    q = upd.callback_query; await q.answer()
    idx = int(q.data.split(":")[1])
    seleccion = REMEDIOS[idx]
    if seleccion == "✏️ Escribir otra solución":
        await q.edit_message_text("✏️ Escribe el remedio definitivo:")
        return REMEDIO_MANUAL
    if seleccion == "Se ejecuta reposición de tramo de fibra óptica":
        c.user_data["_remedio_base"] = seleccion
        await q.edit_message_text("¿Cuántos metros de fibra se utilizaron?\nEjemplo: 500 metros")
        return REMEDIO_METROS
    texto = TEXTOS_REMEDIO.get(seleccion, seleccion)
    c.user_data["remedio"] = texto
    if c.user_data.get("_editing_campo") == "remedio":
        await q.edit_message_text(f"✅ Remedio registrado: {seleccion[:60]}")
        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)
        return await _volver_historial(_FM(q.message.chat_id, q.bot).reply_text,
                                       c.user_data, c)
    await q.edit_message_text(
        f"✅ Remedio registrado.\n\nEnlaces caídos\n\nSelecciona la zona:",
        reply_markup=_ikb_zonas("ez"))
    return ENLACES_ZONA

async def p_remedio_metros(u, c):
    metros = u.message.text.strip()
    texto_base = TEXTOS_REMEDIO["Se ejecuta reposición de tramo de fibra óptica"]
    c.user_data["remedio"] = f"{texto_base} Fibra utilizada: {metros}."
    if c.user_data.get("_editing_campo") == "remedio":
        await u.message.reply_text(f"✅ Remedio con {metros} metros registrado.", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        f"✅ Fibra: {metros}\n\nEnlaces caídos\n\nSelecciona la zona:",
        reply_markup=_ikb_zonas("ez"))
    return ENLACES_ZONA

async def p_remedio_manual(u, c):
    texto = u.message.text.strip()
    errores = _verificar_ortografia(texto)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "remedio", "valor": texto, "errores": errores,
            "next_state": ENLACES_ZONA, "current_state": REMEDIO_MANUAL,
            "next_msg": "✅ Remedio registrado.\n\nEnlaces caídos\n\nSelecciona la zona:",
            "next_markup": _ikb_zonas("ez"),
        }
        await u.message.reply_text(_msg_spell(texto, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return REMEDIO_MANUAL
    c.user_data["remedio"] = texto
    if c.user_data.get("_editing_campo") == "remedio":
        await u.message.reply_text("✅ Remedio registrado.", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, c.user_data, c)
    await u.message.reply_text(
        "✅ Remedio registrado.\n\nEnlaces caídos\n\nSelecciona la zona:",
        reply_markup=_ikb_zonas("ez"))
    return ENLACES_ZONA

# ── ENLACES: flujo zona → tramo (inline) ─────────────────────────────────────
def _ikb_zona_enlaces():
    rows = []
    for i in range(0, len(ZONAS_ORDEN), 2):
        row = [InlineKeyboardButton(ZONAS_ORDEN[i], callback_data=f"ez:{i}")]
        if i+1 < len(ZONAS_ORDEN):
            row.append(InlineKeyboardButton(ZONAS_ORDEN[i+1], callback_data=f"ez:{i+1}"))
        rows.append(row)
    rows.append([InlineKeyboardButton("✏️ Escribir manualmente", callback_data="ez:m")])
    rows.append([InlineKeyboardButton("✅ Finalizar enlaces",    callback_data="ez:fin")])
    return InlineKeyboardMarkup(rows)

async def p_enlaces_zona(u, c):
    """Fallback texto para enlaces zona."""
    await u.message.reply_text("Selecciona la zona del enlace caído:", reply_markup=_ikb_zona_enlaces())
    return ENLACES_ZONA

# ── DATOS DEL REPORTE — selector de campo individual ─────────────────────────
def _historial_datos_texto(d):
    """Genera el texto del historial de datos del reporte."""
    campos_ok = sum(bool(d.get(k)) for k in [
        "codigo","cuadrilla","fecha_inc","fecha_res",
        "personal","lugar","motivo","obs",
        "testigos","remedio","coords"])
    total = 11
    campos_vacios = total - campos_ok
    avance = "✅ ¡Todo completo!" if campos_vacios == 0 else f"⏳ Faltan {campos_vacios} campo(s)"
    return (
        f"📋 DATOS DEL REPORTE — HISTORIAL\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📁 Caso:        {d.get('codigo','—')}\n"
        f"👥 Cuadrilla:   {d.get('cuadrilla','—')}\n"
        f"📍 Lugar:       {str(d.get('lugar','—'))[:40]}\n"
        f"📅 Incidencia:  {d.get('fecha_inc','—')} {d.get('hora_inc','—')}\n"
        f"📅 Arribo:      {d.get('fecha_arr','—')} {d.get('hora_arr','—')}\n"
        f"📅 Reparación:  {d.get('fecha_rep','—')} {d.get('hora_rep','—')}\n"
        f"📅 Resolución:  {d.get('fecha_res','—')} {d.get('hora_res','—')}\n"
        f"⚡ Motivo:      {str(d.get('motivo','—'))[:40]}\n"
        f"🔗 Enlace:      {str(d.get('enlaces','—'))[:40]}\n"
        f"👥 Testigos:    {str(d.get('testigos','—'))[:40]}\n"
        f"🔧 Remedio:     {str(d.get('remedio','—'))[:40]}\n"
        f"📍 Coords:      {str(d.get('coords','—'))[:40]}\n"
        f"👤 Personal:    {str(d.get('personal','—'))[:40]}\n"
        f"📝 Observación: {str(d.get('obs','—'))[:50]}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Campos OK: {campos_ok}/{total}  —  {avance}\n\n"
        f"Toca el campo que quieres llenar o editar:"
    )

async def _volver_historial(reply_fn, d, c):
    """Limpia la bandera de edición y muestra el historial con el selector."""
    c.user_data.pop("_editing_campo", None)
    await reply_fn(_historial_datos_texto(d), reply_markup=_ikb_datos_campos(d))
    return MENU
def _ikb_datos_campos(d):
    """Teclado inline con TODOS los campos de datos del reporte — 2 columnas."""
    def ic(key):
        return "✅" if d.get(key) else "⬜"
    def ic2(k1, k2):
        return "✅" if d.get(k1) and d.get(k2) else "⬜"
    return InlineKeyboardMarkup([
        # Fila 1: cuadrilla + caso
        [InlineKeyboardButton(f"{ic('cuadrilla')} 👥 Cuadrilla",      callback_data="dc:cuadrilla"),
         InlineKeyboardButton(f"{ic('codigo')} 📁 Nº de Caso",        callback_data="dc:codigo")],
        # Fila 2: fechas incidencia + arribo
        [InlineKeyboardButton(f"{ic2('fecha_inc','hora_inc')} 📅 Incidencia", callback_data="dc:fecha_inc"),
         InlineKeyboardButton(f"{ic2('fecha_arr','hora_arr')} 📅 Arribo",     callback_data="dc:fecha_arr")],
        # Fila 3: fechas reparación + resolución
        [InlineKeyboardButton(f"{ic2('fecha_rep','hora_rep')} 📅 Reparación", callback_data="dc:fecha_rep"),
         InlineKeyboardButton(f"{ic2('fecha_res','hora_res')} 📅 Resolución", callback_data="dc:fecha_res")],
        # Fila 4: motivo + enlace(s)
        [InlineKeyboardButton(f"{ic('motivo')} ⚡ Motivo",             callback_data="dc:motivo"),
         InlineKeyboardButton(f"{ic('enlaces')} 🔗 Enlace(s)",         callback_data="dc:enlaces")],
        # Fila 5: testigos + remedio  ← NUEVOS
        [InlineKeyboardButton(f"{ic('testigos')} 👥 Testigos",         callback_data="dc:testigos"),
         InlineKeyboardButton(f"{ic('remedio')} 🔧 Remedio",           callback_data="dc:remedio")],
        # Fila 6: personal + lugar
        [InlineKeyboardButton(f"{ic('personal')} 👤 Personal",         callback_data="dc:personal"),
         InlineKeyboardButton(f"{ic('lugar')} 📍 Lugar",               callback_data="dc:lugar")],
        # Fila 7: coordenadas + observación  ← COORDENADAS NUEVO
        [InlineKeyboardButton(f"{ic('coords')} 🗺️ Coordenadas",        callback_data="dc:coords"),
         InlineKeyboardButton(f"{ic('obs')} 📝 Observación",           callback_data="dc:obs")],
        # Fila 8: listo
        [InlineKeyboardButton("✅ Listo — volver al menú",             callback_data="dc:menu")],
    ])

async def cb_datos_campo(upd, c):
    """Callback: usuario elige qué campo de datos editar."""
    q = upd.callback_query
    campo = q.data.split(":")[1]

    # Toast visible — el usuario VE que el botón respondió
    TOASTS = {
        "cuadrilla": "👥 Abriendo Cuadrilla…",
        "codigo":    "📁 Abriendo Número de Caso…",
        "fecha_inc": "📅 Abriendo Fecha Incidencia…",
        "fecha_arr": "📅 Abriendo Fecha Arribo…",
        "fecha_rep": "📅 Abriendo Fecha Reparación…",
        "fecha_res": "📅 Abriendo Fecha Resolución…",
        "motivo":    "⚡ Abriendo Motivo…",
        "enlaces":   "🔗 Abriendo Enlace(s)…",
        "testigos":  "👥 Abriendo Testigos…",
        "remedio":   "🔧 Abriendo Remedio…",
        "coords":    "🗺️ Abriendo Coordenadas…",
        "personal":  "👤 Abriendo Personal…",
        "lugar":     "📍 Abriendo Lugar…",
        "obs":       "📝 Abriendo Observación…",
        "menu":      "↩️ Volviendo al menú…",
    }
    await q.answer(TOASTS.get(campo, "✏️ Procesando…"))

    # Usar c.bot y upd.effective_chat.id (más confiables que q.bot y q.message.chat_id)
    _bot     = c.bot
    _chat_id = upd.effective_chat.id

    # Quita los botones del historial
    try:
        await q.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass

    # Patrón igual al de cb_fotos_sec (que SÍ funciona)
    class _FM:
        def __init__(self):
            self.chat_id = _chat_id
            self.bot     = _bot
            self.text    = ""
            self.location = None
            self.photo    = None
        async def reply_text(self, txt, **kw):
            await _bot.send_message(chat_id=_chat_id, text=txt, **kw)

    class _FU:
        def __init__(self):
            self.message       = _FM()
            self.effective_user = q.from_user

    fu = _FU()

    # Envía el prompt usando el fake-update (mismo patrón que en el resto del bot)
    async def _prompt(texto, kb=None):
        try:
            if kb:
                await _bot.send_message(chat_id=_chat_id, text=texto, reply_markup=kb)
            else:
                await _bot.send_message(chat_id=_chat_id, text=texto)
        except Exception as e:
            logger.error(f"[cb_datos_campo] _prompt error campo={campo}: {e}")

    if campo == "menu":
        c.user_data.pop("_editing_campo", None)
        return await mostrar_menu(fu, c)

    # Marca el campo en edición para que los handlers regresen aquí al terminar
    c.user_data["_editing_campo"] = campo

    if campo == "cuadrilla":
        await _prompt(
            "👥 CUADRILLA\n\nSelecciona la zona de la cuadrilla:",
            _ikb_zonas("cz"))
        return CUADRILLA_ZONA

    if campo == "codigo":
        await _prompt(
            "📁 NÚMERO DE CASO\n\n"
            "✏️ Escribe el número de caso:\n"
            "Ejemplo: 20220621-0012B")
        return CODIGO

    if campo == "fecha_inc":
        await _prompt(
            "📅 FECHA Y HORA — INCIDENCIA\n\n"
            "Paso 1 de 2 → Escribe la FECHA (DD/MM/AAAA):\n"
            "Ejemplo: 21/06/2022")
        return FECHA_INC

    if campo == "fecha_arr":
        await _prompt(
            "📅 FECHA Y HORA — ARRIBO\n\n"
            "Paso 1 de 2 → Escribe la FECHA (DD/MM/AAAA):\n"
            "Ejemplo: 21/06/2022")
        return FECHA_ARR

    if campo == "fecha_rep":
        await _prompt(
            "📅 FECHA Y HORA — REPARACIÓN\n\n"
            "Paso 1 de 2 → Escribe la FECHA (DD/MM/AAAA):\n"
            "Ejemplo: 21/06/2022")
        return FECHA_REP

    if campo == "fecha_res":
        await _prompt(
            "📅 FECHA Y HORA — RESOLUCIÓN\n\n"
            "Paso 1 de 2 → Escribe la FECHA (DD/MM/AAAA):\n"
            "Ejemplo: 21/06/2022")
        return FECHA_RES

    if campo == "motivo":
        await _prompt(
            "⚡ MOTIVO DE LA INCIDENCIA\n\n"
            "Selecciona la zona del tramo afectado:",
            _ikb_zonas("mz"))
        return MOTIVO_ZONA

    if campo == "enlaces":
        await _prompt(
            "🔗 ENLACES CAÍDOS\n\n"
            "Selecciona la zona del enlace:",
            _ikb_zona_enlaces())
        return ENLACES_ZONA

    if campo == "testigos":
        c.user_data["_causas_campo"] = "testigos"
        await _prompt(
            "👥 TESTIGOS DEL INCIDENTE\n\n"
            "Selecciona la causa o escribe manualmente:",
            _ikb_lista(CAUSAS, "ca"))
        return TESTIGOS

    if campo == "remedio":
        await _prompt(
            "🔧 REMEDIO DEFINITIVO\n\n"
            "Selecciona el tipo de remedio:",
            _ikb_lista(REMEDIOS, "rem"))
        return REMEDIO

    if campo == "coords":
        c.user_data["coords_lista"] = c.user_data.get("coords_lista", [])
        n = len(c.user_data["coords_lista"])
        await _prompt(
            f"🗺️ COORDENADAS GPS\n\n"
            f"Ya tienes {n} coordenada(s) guardada(s).\n\n"
            "📍 Comparte tu ubicación GPS desde Telegram\n"
            "✏️ O escribe: latitud, longitud\n"
            "   Ejemplo: -0.243225, -78.518738\n\n"
            "Puedes agregar varias. Toca Listo cuando termines.",
            _ikb_listo("coord"))
        return COORDS

    if campo == "personal":
        await _prompt(
            "👤 PERSONAL QUE INTERVINO\n\n"
            "✏️ Escribe los nombres separados por  /\n"
            "Ejemplo: Juan Pérez / María López")
        return PERSONAL

    if campo == "lugar":
        await _prompt(
            "📍 LUGAR DE LA INCIDENCIA\n\n"
            "✏️ Escribe el lugar donde ocurrió:")
        return LUGAR

    if campo == "obs":
        c.user_data["_causas_campo"] = "obs"
        await _prompt(
            "📝 OBSERVACIÓN GENERAL\n\n"
            "Selecciona la observación:",
            _ikb_lista(CAUSAS, "ca"))
        return OBS

    # Fallback seguro
    c.user_data.pop("_editing_campo", None)
    return await mostrar_menu(fu, c)


async def cb_enlaces_zona(upd, c):
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    if data == "fin":
        lista = c.user_data.get("enlaces_lista", [])
        c.user_data["enlaces"] = "\n".join(lista) if lista else "Ninguno"
        await q.edit_message_text(f"✅ {len(lista)} enlace(s) registrado(s).\n\nPersonal que intervino\nSeparados por / Ejemplo: Juan / Pedro")
        return PERSONAL
    if data == "m":
        await q.edit_message_text("✏️ Escribe el nombre del enlace:")
        return ENLACES_MANUAL
    zona = ZONAS_ORDEN[int(data)]
    c.user_data["_enlaces_zona_actual"] = zona
    tramos = TRAMOS_POR_ZONA.get(zona, [])
    kb = _ikb_lista(tramos, "et", extra=[("✏️ Escribir manualmente","et:m")])
    await q.edit_message_text(f"Zona: {zona}\n\nSelecciona el enlace caído:", reply_markup=kb)
    return ENLACES_TRAMO

async def p_enlaces_tramo(u, c):
    """Fallback para tramo manual."""
    tramo = u.message.text.strip()
    lista = c.user_data.get("enlaces_lista", [])
    lista.append(tramo); c.user_data["enlaces_lista"] = lista
    await u.message.reply_text(
        f"✅ Enlace {len(lista)}: {tramo}\n\n¿Agregar otro?",
        reply_markup=IKB_ENLACES_MAS)
    return ENLACES_MAS

async def cb_enlaces_tramo(upd, c):
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    if data == "m":
        await q.edit_message_text("✏️ Escribe el nombre exacto del enlace caído:")
        return ENLACES_MANUAL
    zona = c.user_data.get("_enlaces_zona_actual","")
    tramos = TRAMOS_POR_ZONA.get(zona, [])
    tramo = tramos[int(data)]
    lista = c.user_data.get("enlaces_lista", [])
    lista.append(tramo); c.user_data["enlaces_lista"] = lista
    await q.edit_message_text(
        f"✅ Enlace {len(lista)}: {tramo}\n\n¿Agregar otro?",
        reply_markup=IKB_ENLACES_MAS)
    return ENLACES_MAS

async def p_enlaces_mas(u, c):
    """Fallback texto para más enlaces."""
    await u.message.reply_text("¿Agregar otro enlace?", reply_markup=IKB_ENLACES_MAS)
    return ENLACES_MAS

async def cb_enlaces_mas(upd, c):
    q = upd.callback_query; await q.answer()
    data = q.data.split(":")[1]
    if data == "mas":
        await q.edit_message_text("Selecciona la zona del siguiente enlace:", reply_markup=_ikb_zona_enlaces())
        return ENLACES_ZONA
    if data == "manual":
        await q.edit_message_text("✏️ Escribe el nombre del enlace:")
        return ENLACES_MANUAL
    # fin
    lista = c.user_data.get("enlaces_lista", [])
    c.user_data["enlaces"] = "\n".join(lista) if lista else "Ninguno"

    class _FM:
        def __init__(self, chat_id, bot):
            self.chat_id = chat_id; self.bot = bot
        async def reply_text(self, txt, **kw):
            await self.bot.send_message(self.chat_id, txt, **kw)

    if c.user_data.get("_editing_campo") == "enlaces":
        await q.edit_message_text(f"✅ {len(lista)} enlace(s) registrado(s).")
        return await _volver_historial(_FM(q.message.chat_id, q.bot).reply_text,
                                       c.user_data, c)
    await q.edit_message_text(
        f"✅ {len(lista)} enlace(s) registrado(s).\n\nPersonal que intervino\nSeparados por / Ejemplo: Juan / Pedro")
    return PERSONAL

async def p_enlaces_manual(u, c):
    texto = u.message.text.strip()
    lista = c.user_data.get("enlaces_lista", [])
    lista.append(texto)
    c.user_data["enlaces_lista"] = lista
    n = len(lista)
    await u.message.reply_text(
        f"✅ Enlace {n} agregado: {texto}\n\n¿Deseas agregar otro?",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Agregar otro enlace",         callback_data="em:mas")],
            [InlineKeyboardButton("✏️ Escribir enlace manualmente", callback_data="em:manual")],
            [InlineKeyboardButton("✅ Finalizar enlaces",           callback_data="em:fin")],
        ]))
    return ENLACES_MAS

async def p_enlaces(u,c):
    """Fallback por compatibilidad."""
    c.user_data["enlaces"] = u.message.text.strip()
    await u.message.reply_text("Paso 14/16 — Personal que intervino. Separados por / Ejemplo: Juan / Pedro", reply_markup=KB_MENU)
    return PERSONAL

async def p_personal(u,c):
    c.user_data["personal"] = u.message.text.strip()
    d = c.user_data
    if c.user_data.get("_editing_campo") == "personal":
        await u.message.reply_text(f"✅ Personal registrado: {d['personal']}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, d, c)
    await u.message.reply_text("Paso 15/16 — Lugar de la incidencia", reply_markup=KB_MENU)
    return LUGAR

async def p_lugar(u, c):
    c.user_data["lugar"] = u.message.text.strip()
    d = c.user_data
    if c.user_data.get("_editing_campo") == "lugar":
        await u.message.reply_text(f"✅ Lugar registrado: {d['lugar'][:60]}", reply_markup=KB_MENU)
        return await _volver_historial(u.message.reply_text, d, c)
    c.user_data["coords_lista"] = []
    await u.message.reply_text(
        "Coordenadas GPS\n\n"
        "📍 Comparte tu ubicacion GPS desde Telegram\n"
        "✏️ O escribe la coordenada manualmente\n"
        "   Ejemplo: -0.243225, -78.518738\n\n"
        "Puedes agregar varias.",
        reply_markup=_ikb_listo("coord"))
    return COORDS

async def p_coords(u, c):
    if "coords_lista" not in c.user_data:
        c.user_data["coords_lista"] = []

    if u.message.location:
        lat = u.message.location.latitude
        lon = u.message.location.longitude
        cs  = f"{lat:.6f}, {lon:.6f}"
        c.user_data["coords_lista"].append(cs)
        n = len(c.user_data["coords_lista"])
        await u.message.reply_text(
            f"📍 Coordenada {n} guardada:\n{cs}\n\nEnvia otra o toca Listo.",
            reply_markup=_ikb_listo("coord"))
        return COORDS

    if u.message.text:
        t = u.message.text.strip()
        if t.lower() in ["listo", "omitir", "ok", "si"]:
            return await _fin_coords(u.message.chat_id, c, u)
        else:
            c.user_data["coords_lista"].append(t)
            n = len(c.user_data["coords_lista"])
            await u.message.reply_text(
                f"✅ Coordenada {n} guardada:\n{t}\n\nEnvia otra o toca Listo.",
                reply_markup=_ikb_listo("coord"))
            return COORDS

async def _fin_coords(chat_id, c, u_or_q):
    lista = c.user_data.get("coords_lista", [])
    if lista:
        c.user_data["coords"] = "\n".join(lista)
    else:
        c.user_data["coords"] = "No disponible"
    if c.user_data.get("_editing_campo") == "coords":
        await u_or_q.reply_text(f"✅ {len(lista)} coordenada(s) guardadas.", reply_markup=KB_MENU)
        return await _volver_historial(u_or_q.reply_text, c.user_data, c)
    c.user_data["_causas_campo"] = "obs"
    await u_or_q.reply_text(
        "✅ Coordenadas guardadas.\n\nObservaciones generales\n\nSelecciona la observacion:",
        reply_markup=_ikb_lista(CAUSAS, "ca"))
    return OBS


# ══════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN DE FOTOS — Brillo y overlay de coordenadas GPS
# ══════════════════════════════════════════════════════════════════════════════
async def _validar_foto_calidad(ruta: str, reply_fn, markup) -> bool:
    """
    Valida dos requisitos:
      1. Brillo suficiente (fotos de noche como la de referencia son aceptadas).
      2. Overlay de coordenadas GPS: texto blanco con fecha, GPS, ubicación.
    Retorna True si válida, False si rechazada (ya envió el mensaje de error).
    El archivo NO se elimina aquí — el llamador lo hace si devuelve False.
    """
    if not PIL_OK:
        logger.warning("Pillow no disponible — validación de fotos desactivada")
        return True
    try:
        img  = PILImage.open(ruta).convert("RGB")
        w, h = img.size
        px   = img.load()

        # Paso de muestreo: ~1 píxel cada 6-8 px para rapidez
        paso = max(1, min(w, h) // 160)

        # ── 1. Brillo: promedio global + detector de "foto negra con foco" ──
        # Fotos de noche válidas (calle, poste, pozo) tienen cielo oscuro PERO
        # muchos tonos intermedios de gris (edificios, calles, equipos).
        # El bokeh/flash-ciego tiene prácticamente TODO en negro puro (lum < 12)
        # más un único blob brillante → ratio de negro-absoluto > 0.82.
        suma_b = 0; n_b = 0; px_negro_abs = 0
        for y in range(0, h, paso):
            for x in range(0, w, paso):
                r, g, b = px[x, y]
                lum = 0.299*r + 0.587*g + 0.114*b
                suma_b += lum
                n_b += 1
                if lum < 12:          # negro casi puro
                    px_negro_abs += 1
        avg_b         = suma_b / max(n_b, 1)
        ratio_negro   = px_negro_abs / max(n_b, 1)

        # Rechazar si: muy oscura en promedio  O  casi todo es negro puro con avg bajo
        if avg_b < 28 or (ratio_negro > 0.82 and avg_b < 60):
            await reply_fn(
                "🌑 *Foto muy oscura — rechazada.*\n\n"
                "No se puede registrar esta foto.\n"
                "Vuelve a subir la foto con mejor iluminación 📷",
                parse_mode="Markdown",
                reply_markup=markup)
            return False

        # ── 1.5. Detección de logo TELCONET (esquina superior derecha) ───────
        #
        # La app Timestamp Cam con el logo Telconet imprime el logo azul en la
        # esquina superior-derecha. Buscamos píxeles con canal azul dominante
        # en el 40 % derecho y 30 % superior (zona ampliada para cubrir
        # logos de distintos tamaños y resoluciones).
        # Umbral relajado: 1.2 % de píxeles azules es suficiente — el triángulo
        # azul del logo es pequeño relativo a la zona de búsqueda.
        x0_logo = int(w * 0.60)          # 40 % derecho
        y1_logo = int(h * 0.30)          # 30 % superior

        pixeles_azules = 0
        pixeles_zona   = 0
        for y in range(0, y1_logo, paso):
            for x in range(x0_logo, w, paso):
                r, g, b = px[x, y]
                pixeles_zona += 1
                # Azul dominante con umbral relajado para logos claros y oscuros
                if b > 60 and b > r * 1.3 and b > g * 1.1:
                    pixeles_azules += 1

        ratio_azul = pixeles_azules / max(pixeles_zona, 1)

        if ratio_azul < 0.012:   # menos del 1.2 % → sin logo azul Telconet
            await reply_fn(
                "🏢 *Foto rechazada — faltan requisitos obligatorios.*\n\n"
                "La foto debe tener *los dos elementos* visibles:\n\n"
                "  🔷 Logo *TELCONET* en la esquina superior derecha\n"
                "  📅 Fecha y hora · 🌐 Coordenadas GPS · 📍 Lugar\n\n"
                "Usa la app *Timestamp Cam* con el logo Telconet activado "
                "y vuelve a enviar la foto 📷",
                parse_mode="Markdown",
                reply_markup=markup)
            return False

        # ── 2. Detección de texto-overlay en la franja inferior ───────────
        #
        # El overlay de coordenadas (fecha · GPS · ubicación · código)
        # es TEXTO BLANCO sobre fondo oscuro. Eso crea en cada fila de texto
        # muchas transiciones oscuro↔claro (los bordes de cada letra).
        #
        # Un fondo liso/metálico o un reflejo suave tiene POCAS transiciones.
        #
        # Algoritmo:
        #   Para cada fila muestreada en el 30 % inferior de la imagen:
        #     - Calcular luminancias de izquierda a derecha
        #     - Contar transiciones: píxel oscuro→claro o claro→oscuro
        #       (umbral de texto blanco = lum > 160)
        #     - Una "fila con texto" tiene: 4–100 transiciones
        #       Y entre 1 % y 55 % de píxeles blancos
        #   Contar cuántas filas-con-texto hay.
        #   Con 3-4 líneas de overlay (≈ 30 px c/u) esperamos ≥ 8 filas.

        y0 = int(h * 0.70)          # analizar el 30 % inferior
        UMBRAL_BLANCO = 158         # luminancia mínima para "píxel de texto"

        filas_con_texto = 0

        for y in range(y0, h, paso):
            lums = []
            for x in range(0, w, paso):
                r, g, b = px[x, y]
                lums.append(0.299*r + 0.587*g + 0.114*b)

            if len(lums) < 4:
                continue

            # Contar transiciones oscuro ↔ claro
            trans = 0
            prev  = lums[0] >= UMBRAL_BLANCO
            for lum in lums[1:]:
                curr = lum >= UMBRAL_BLANCO
                if curr != prev:
                    trans += 1
                prev = curr

            # Proporción de píxeles blancos en la fila
            blancos = sum(1 for l in lums if l >= UMBRAL_BLANCO)
            ratio   = blancos / len(lums)

            # Fila con texto: suficientes transiciones (letras) y
            # no toda la fila es blanca (no es un fondo sólido blanco)
            if trans >= 4 and 0.01 <= ratio <= 0.55:
                filas_con_texto += 1

        # Con 3 líneas de texto y paso ≈ 7px → esperamos ~12-18 filas
        # Umbral conservador: 7 filas de texto = overlay presente
        if filas_con_texto < 7:
            await reply_fn(
                "📍 *Foto sin coordenadas — rechazada.*\n\n"
                "Todas las fotos DEBEN tener el *overlay de coordenadas GPS* "
                "visible en la parte inferior:\n"
                "  📅 Fecha y hora\n"
                "  🌐 Latitud / Longitud\n"
                "  📍 Nombre del lugar\n"
                "  🔗 Código del enlace\n\n"
                "✅ Activa la app *Timestamp Cam* antes de fotografiar "
                "y vuelve a enviar la foto 📷",
                parse_mode="Markdown",
                reply_markup=markup)
            return False

        return True   # ← foto válida ✓

    except Exception as e:
        logger.warning(f"Error validando foto {ruta}: {e}")
        return True   # si el análisis falla técnicamente, no bloquear


def _guardar_foto_validada(uid, key, lista_actual, foto_tg_file, ruta):
    """
    Retorna la lista actualizada con la nueva ruta si la foto ya fue validada.
    Construye una COPIA nueva para no mutar la lista original en user_data
    hasta que estemos seguros de que la foto es válida.
    """
    nueva = list(lista_actual)   # copia explícita — nunca mutamos el original
    nueva.append(ruta)
    return nueva


# ── Fábrica de handlers de fotos ───────────────────────────────────────────────
def make_foto_handler(key, estado_actual, msg_siguiente, estado_siguiente, sec_listo, max_fotos=8):
    async def handler(u, c):
        uid = u.effective_user.id
        if u.message.photo:
            # Copia defensiva: nunca mutamos user_data hasta validar
            lista_actual = list(c.user_data.get(key, []))
            if len(lista_actual) >= max_fotos:
                await u.message.reply_text(
                    f"Limite de {max_fotos} fotos alcanzado.",
                    reply_markup=_ikb_listo(sec_listo))
                return estado_actual

            foto_tg  = u.message.photo[-1]
            file_obj = await c.bot.get_file(foto_tg.file_id)
            # Ruta temporal con prefijo "_tmp_" — sólo pasa a producción si válida
            ruta_tmp = os.path.join(FOTOS_DIR, f"_tmp_{uid}_{key}_{len(lista_actual)}.jpg")
            ruta_ok  = os.path.join(FOTOS_DIR, f"{uid}_{key}_{len(lista_actual)}.jpg")

            await file_obj.download_to_drive(ruta_tmp)
            await u.message.reply_text("🔍 Analizando foto...")

            # ── Validación ───────────────────────────────────────────────
            valida = await _validar_foto_calidad(
                ruta_tmp,
                lambda txt, **kw: u.message.reply_text(txt, **kw),
                _ikb_listo(sec_listo))

            if not valida:
                # Eliminar el archivo temporal — la foto NO se guarda
                try:
                    os.remove(ruta_tmp)
                    logger.info(f"Foto rechazada eliminada: {ruta_tmp}")
                except Exception as e:
                    logger.warning(f"No se pudo eliminar {ruta_tmp}: {e}")
                # user_data[key] queda INTACTO (lista_actual era una copia)
                return estado_actual

            # ── Foto válida: mover tmp → nombre definitivo ────────────────
            try:
                os.rename(ruta_tmp, ruta_ok)
            except Exception:
                ruta_ok = ruta_tmp   # si rename falla, usar el tmp

            lista_nueva = list(lista_actual)
            lista_nueva.append(ruta_ok)
            c.user_data[key] = lista_nueva   # sólo aquí actualizamos user_data

            n = len(lista_nueva); quedan = max_fotos - n
            msg = f"✅ Foto {n} guardada con coordenadas."
            msg += f" Puedes enviar {quedan} más." if quedan > 0 else " Límite alcanzado."
            await u.message.reply_text(msg, reply_markup=_ikb_listo(sec_listo))
            return estado_actual

        elif u.message.text:
            t = u.message.text.strip().lower()
            if t in ["listo", "omitir", "si", "ok"]:
                n = len(c.user_data.get(key, []))
                await u.message.reply_text(
                    f"✅ {n} foto(s) guardadas.\n\n{msg_siguiente}",
                    reply_markup=_ikb_listo(sec_listo.replace("fa","fd").replace("fd","ff") if "fa" in sec_listo or "fd" in sec_listo else sec_listo))
                return estado_siguiente
            else:
                await u.message.reply_text("Envia una foto o toca Listo.", reply_markup=_ikb_listo(sec_listo))
                return estado_actual
    return handler

p_fotos_antes   = make_foto_handler("fotos_antes",   FOTOS_ANTES,
    "FOTOS — DURANTE la reparacion\nEnvia fotos.", FOTOS_DURANTE, "fa")
p_fotos_durante = make_foto_handler("fotos_durante", FOTOS_DURANTE,
    "FOTOS — FIN de la reparacion\nEnvia fotos.", FOTOS_FIN, "fd")

async def p_fotos_fin(u, c):
    uid = u.effective_user.id; key = "fotos_fin"
    if u.message.photo:
        lista_actual = list(c.user_data.get(key, []))   # copia defensiva
        if len(lista_actual) >= 8:
            await u.message.reply_text("Limite alcanzado.", reply_markup=_ikb_listo("ff"))
            return FOTOS_FIN
        foto_tg  = u.message.photo[-1]
        file_obj = await c.bot.get_file(foto_tg.file_id)
        ruta_tmp = os.path.join(FOTOS_DIR, f"_tmp_{uid}_{key}_{len(lista_actual)}.jpg")
        ruta_ok  = os.path.join(FOTOS_DIR, f"{uid}_{key}_{len(lista_actual)}.jpg")
        await file_obj.download_to_drive(ruta_tmp)
        await u.message.reply_text("🔍 Analizando foto...")
        valida = await _validar_foto_calidad(
            ruta_tmp,
            lambda txt, **kw: u.message.reply_text(txt, **kw),
            _ikb_listo("ff"))
        if not valida:
            try: os.remove(ruta_tmp); logger.info(f"Foto FIN rechazada: {ruta_tmp}")
            except Exception as e: logger.warning(f"No se pudo eliminar {ruta_tmp}: {e}")
            return FOTOS_FIN   # user_data[key] intacto
        try: os.rename(ruta_tmp, ruta_ok)
        except: ruta_ok = ruta_tmp
        lista_nueva = list(lista_actual); lista_nueva.append(ruta_ok)
        c.user_data[key] = lista_nueva
        n = len(lista_nueva); quedan = 8 - n
        msg = f"✅ Foto FIN {n} guardada con coordenadas." + (f" Puedes enviar {quedan} más." if quedan > 0 else "")
        await u.message.reply_text(msg, reply_markup=_ikb_listo("ff"))
        return FOTOS_FIN
    elif u.message.text:
        t = u.message.text.strip().lower()
        if t in ["listo", "omitir", "si", "ok"]:
            n = len(c.user_data.get(key, []))
            await u.message.reply_text(f"✅ {n} foto(s) FIN guardadas.\n\n📸 Sección de fotos completada.", reply_markup=KB_MENU)
            return await mostrar_menu(u, c)
        else:
            await u.message.reply_text("Envia una foto o toca Listo.", reply_markup=_ikb_listo("ff"))
            return FOTOS_FIN

p_trazas_antes = make_foto_handler("trazas_antes", TRAZAS_ANTES,
    "TRAZAS OTDR — DESPUES de la reparacion\nEnvia capturas.", TRAZAS_DESPUES, "ta")

async def p_trazas_despues(u, c):
    uid = u.effective_user.id; key = "trazas_despues"
    if u.message.photo:
        lista_actual = list(c.user_data.get(key, []))   # copia defensiva
        if len(lista_actual) >= 8:
            await u.message.reply_text("Limite alcanzado.", reply_markup=_ikb_listo("td"))
            return TRAZAS_DESPUES
        foto_tg  = u.message.photo[-1]
        file_obj = await c.bot.get_file(foto_tg.file_id)
        ruta_tmp = os.path.join(FOTOS_DIR, f"_tmp_{uid}_{key}_{len(lista_actual)}.jpg")
        ruta_ok  = os.path.join(FOTOS_DIR, f"{uid}_{key}_{len(lista_actual)}.jpg")
        await file_obj.download_to_drive(ruta_tmp)
        await u.message.reply_text("🔍 Analizando foto...")
        valida = await _validar_foto_calidad(
            ruta_tmp,
            lambda txt, **kw: u.message.reply_text(txt, **kw),
            _ikb_listo("td"))
        if not valida:
            try: os.remove(ruta_tmp); logger.info(f"Traza rechazada: {ruta_tmp}")
            except Exception as e: logger.warning(f"No se pudo eliminar {ruta_tmp}: {e}")
            return TRAZAS_DESPUES   # user_data[key] intacto
        try: os.rename(ruta_tmp, ruta_ok)
        except: ruta_ok = ruta_tmp
        lista_nueva = list(lista_actual); lista_nueva.append(ruta_ok)
        c.user_data[key] = lista_nueva
        await u.message.reply_text(f"✅ Traza {len(lista_nueva)} guardada con coordenadas.", reply_markup=_ikb_listo("td"))
        return TRAZAS_DESPUES
    elif u.message.text:
        t = u.message.text.strip().lower()
        if t in ["listo", "omitir", "si", "ok"]:
            n = len(c.user_data.get(key, []))
            await u.message.reply_text(f"✅ {n} traza(s) guardadas.\n\n📡 Sección de trazas completada.", reply_markup=KB_MENU)
            return await mostrar_menu(u, c)
        else:
            await u.message.reply_text("Envia una captura o toca Listo.", reply_markup=_ikb_listo("td"))
            return TRAZAS_DESPUES

# ── Pasos texto restantes ──────────────────────────────────────────────────────
async def p_obs(u, c):
    texto = u.message.text.strip()
    if texto == "✏️ Escribir manualmente":
        await u.message.reply_text(
            "Escribe tu observacion:", reply_markup=KB_MENU)
        return OBS_MANUAL
    # Si seleccionó una causa predefinida, usar el texto largo
    obs_final = OBSERVACIONES_PREDEFINIDAS.get(texto, texto)
    c.user_data["obs"] = obs_final
    d = c.user_data
    campos_ok = sum(bool(d.get(k)) for k in
        ["codigo","cuadrilla","fecha_inc","fecha_res","personal","lugar","motivo","obs"])
    campos_vacios = 8 - campos_ok
    avance = "✅ ¡Todo completo!" if campos_vacios == 0 else f"⏳ Faltan {campos_vacios} campo(s)"
    await u.message.reply_text(
        f"✅ Observacion registrada.\n\n"
        f"📋 DATOS DEL REPORTE — HISTORIAL\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📁 Caso:        {d.get('codigo','—')}\n"
        f"👥 Cuadrilla:   {d.get('cuadrilla','—')}\n"
        f"📍 Lugar:       {str(d.get('lugar','—'))[:40]}\n"
        f"📅 Incidencia:  {d.get('fecha_inc','—')} {d.get('hora_inc','—')}\n"
        f"📅 Arribo:      {d.get('fecha_arr','—')} {d.get('hora_arr','—')}\n"
        f"📅 Reparación:  {d.get('fecha_rep','—')} {d.get('hora_rep','—')}\n"
        f"📅 Resolución:  {d.get('fecha_res','—')} {d.get('hora_res','—')}\n"
        f"⚡ Motivo:      {str(d.get('motivo','—'))[:40]}\n"
        f"🔗 Enlace:      {str(d.get('enlaces','—'))[:40]}\n"
        f"👤 Personal:    {str(d.get('personal','—'))[:40]}\n"
        f"📝 Observación: {str(d.get('obs','—'))[:50]}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Campos OK: {campos_ok}/8  —  {avance}\n\n"
        f"Toca el campo que quieres llenar o editar:",
        reply_markup=_ikb_datos_campos(d))
    return MENU

async def p_obs_manual(u, c):
    texto = u.message.text.strip()
    errores = _verificar_ortografia(texto)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "obs", "valor": texto, "errores": errores,
            "next_state": MENU, "current_state": OBS_MANUAL,
            "show_historial_obs": True,
        }
        await u.message.reply_text(_msg_spell(texto, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return OBS_MANUAL
    c.user_data["obs"] = texto
    d = c.user_data
    campos_ok = sum(bool(d.get(k)) for k in
        ["codigo","cuadrilla","fecha_inc","fecha_res","personal","lugar","motivo","obs"])
    campos_vacios = 8 - campos_ok
    avance = "✅ ¡Todo completo!" if campos_vacios == 0 else f"⏳ Faltan {campos_vacios} campo(s)"
    await u.message.reply_text(
        f"✅ Observacion registrada.\n\n"
        f"📋 DATOS DEL REPORTE — HISTORIAL\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📁 Caso:        {d.get('codigo','—')}\n"
        f"👥 Cuadrilla:   {d.get('cuadrilla','—')}\n"
        f"📍 Lugar:       {str(d.get('lugar','—'))[:40]}\n"
        f"📅 Incidencia:  {d.get('fecha_inc','—')} {d.get('hora_inc','—')}\n"
        f"📅 Arribo:      {d.get('fecha_arr','—')} {d.get('hora_arr','—')}\n"
        f"📅 Reparación:  {d.get('fecha_rep','—')} {d.get('hora_rep','—')}\n"
        f"📅 Resolución:  {d.get('fecha_res','—')} {d.get('hora_res','—')}\n"
        f"⚡ Motivo:      {str(d.get('motivo','—'))[:40]}\n"
        f"🔗 Enlace:      {str(d.get('enlaces','—'))[:40]}\n"
        f"👤 Personal:    {str(d.get('personal','—'))[:40]}\n"
        f"📝 Observación: {str(d.get('obs','—'))[:50]}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Campos OK: {campos_ok}/8  —  {avance}\n\n"
        f"Toca el campo que quieres llenar o editar:",
        reply_markup=_ikb_datos_campos(d))
    return MENU

async def p_resumen(u, c):
    t = u.message.text.strip()
    if t.lower() == "omitir":
        c.user_data["resumen_lista"] = []
        await u.message.reply_text("✅ Resumen omitido.\n\n📝 Sección completada.", reply_markup=KB_MENU)
        return await mostrar_menu(u, c)
    lista = c.user_data.get("resumen_lista", [])
    if len(lista) >= 20:
        await u.message.reply_text(
            f"⚠️ Límite de 20 líneas alcanzado.\n📊 Total: 20/20",
            reply_markup=_ikb_listo("res"))
        return RESUMEN
    errores = _verificar_ortografia(t)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "_resumen_line", "valor": t, "errores": errores,
            "next_state": RESUMEN, "current_state": RESUMEN,
        }
        await u.message.reply_text(_msg_spell(t, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return RESUMEN
    lista.append(t)
    c.user_data["resumen_lista"] = lista
    n = len(lista)
    await u.message.reply_text(
        f"✅ Línea {n} guardada.\n📊 Total: {n}/20\n\nEnvía otra línea o toca Listo:",
        reply_markup=_ikb_listo("res"))
    return RESUMEN

async def p_mant_confirm(u, c):
    """Fallback texto — muestra el panel principal."""
    nodos = c.user_data.setdefault("mant_nodos", {})
    await u.message.reply_text(
        _texto_mant_panel(nodos),
        reply_markup=_ikb_mant_panel(nodos))
    return MANT_CONFIRM

async def cb_mant_confirm(upd, c):
    """mc:si / mc:no (legado) → panel nuevo."""
    q = upd.callback_query; await q.answer()
    nodos = c.user_data.setdefault("mant_nodos", {})
    await q.edit_message_text(
        _texto_mant_panel(nodos),
        reply_markup=_ikb_mant_panel(nodos))
    return MANT_CONFIRM

# ── Panel principal callbacks (mp:) ──────────────────────────────────────────
async def cb_mant_panel(upd, c):
    q = upd.callback_query; await q.answer()
    partes = q.data.split(":", 2)
    acc    = partes[1]

    if acc == "nuevo":
        await q.edit_message_text(
            "📡 Nombre del nuevo nodo:\nEjemplo: NODO A  /  NODO B  /  EMPALME KM-12")
        return MANT_EST

    elif acc == "nodo":
        nombre = partes[2]
        c.user_data["_mant_cur_est"] = nombre
        nodos = c.user_data.setdefault("mant_nodos", {})
        hilos = nodos.setdefault(nombre, {})
        await q.edit_message_text(
            _texto_mant_nodo(nombre, hilos),
            reply_markup=_ikb_mant_nodo(nombre, hilos))
        return MANT_MORE_EST

    elif acc == "obs":
        await q.edit_message_text("📝 Observaciones del mantenimiento:\n(O escribe Omitir)")
        return MANT_OBS

    elif acc == "menu":
        nodos = c.user_data.get("mant_nodos", {})
        c.user_data["mant_filas"] = _mant_nodos_a_filas(nodos)
        c.user_data.setdefault("mant_obs", "")
        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)
        class _FU:
            def __init__(self, q):
                self.message = _FM(q.message.chat_id, q.bot)
                self.effective_user = q.from_user
        await q.edit_message_text("↩ Volviendo al menú...")
        return await mostrar_menu(_FU(q), c)

    return MANT_CONFIRM

# ── Panel de nodo callbacks (mn:) ─────────────────────────────────────────────
async def cb_mant_nodo(upd, c):
    q      = upd.callback_query; await q.answer()
    partes = q.data.split(":", 2)
    acc    = partes[1]
    nombre = c.user_data.get("_mant_cur_est", "")
    nodos  = c.user_data.setdefault("mant_nodos", {})
    hilos  = nodos.setdefault(nombre, {})

    if acc == "add":
        await q.edit_message_text(f"📡 {nombre}\n\nNúmero del hilo a agregar:\nEjemplo: 11")
        return MANT_HILO

    elif acc == "hilo":
        hilo_num = partes[2]
        c.user_data["_mant_cur_hilo"] = hilo_num
        hilos.setdefault(hilo_num, {"da":"","pa":"","dd":"","pd":""})
        datos = hilos[hilo_num]
        await q.edit_message_text(
            _texto_mant_hilo(nombre, hilo_num, datos),
            reply_markup=_ikb_mant_hilo(datos))
        return MANT_MORE_HILO

    elif acc == "back":
        nodos = c.user_data.get("mant_nodos", {})
        await q.edit_message_text(
            _texto_mant_panel(nodos),
            reply_markup=_ikb_mant_panel(nodos))
        return MANT_CONFIRM

    return MANT_MORE_EST

# ── Panel de hilo callbacks (mh:) ─────────────────────────────────────────────
async def cb_mant_hilo(upd, c):
    q   = upd.callback_query; await q.answer()
    acc = q.data.split(":")[1]
    nombre   = c.user_data.get("_mant_cur_est",  "")
    hilo_num = c.user_data.get("_mant_cur_hilo", "")

    if acc == "back":
        nodos = c.user_data.setdefault("mant_nodos", {})
        hilos = nodos.setdefault(nombre, {})
        await q.edit_message_text(
            _texto_mant_nodo(nombre, hilos),
            reply_markup=_ikb_mant_nodo(nombre, hilos))
        return MANT_MORE_EST

    CAMPO_INFO = {
        "da": ("📏 Distancia ANTES de la reparación (km):\nEjemplo: 151.41", MANT_DA),
        "pa": ("📉 Pérdida ANTES de la reparación (dB):\nEjemplo: 37.34",   MANT_PA),
        "dd": ("📏 Distancia DESPUÉS de la reparación (km):\nEjemplo: 154.33", MANT_DD),
        "pd": ("📉 Pérdida DESPUÉS de la reparación (dB):\nEjemplo: 33.34",  MANT_PD),
    }
    if acc in CAMPO_INFO:
        c.user_data["_mant_cur_campo"] = acc
        msg, estado = CAMPO_INFO[acc]
        nodos  = c.user_data.get("mant_nodos", {})
        actual = nodos.get(nombre, {}).get(hilo_num, {}).get(acc, "")
        extra  = f"\nValor actual: {actual}" if actual else ""
        await q.edit_message_text(f"📡 {nombre} — Hilo {hilo_num}\n\n{msg}{extra}")
        return estado

    return MANT_MORE_HILO

async def _guardar_campo_mant(u, c, campo):
    valor    = u.message.text.strip()
    nombre   = c.user_data.get("_mant_cur_est",  "")
    hilo_num = c.user_data.get("_mant_cur_hilo", "")
    nodos    = c.user_data.setdefault("mant_nodos", {})
    hilos    = nodos.setdefault(nombre, {})
    datos    = hilos.setdefault(hilo_num, {"da":"","pa":"","dd":"","pd":""})
    datos[campo] = valor
    await u.message.reply_text(
        f"✅ Guardado.\n\n" + _texto_mant_hilo(nombre, hilo_num, datos),
        reply_markup=_ikb_mant_hilo(datos))
    return MANT_MORE_HILO

async def p_mant_est(u, c):
    nombre = u.message.text.strip()
    c.user_data["_mant_cur_est"] = nombre
    nodos  = c.user_data.setdefault("mant_nodos", {})
    hilos  = nodos.setdefault(nombre, {})
    await u.message.reply_text(
        _texto_mant_nodo(nombre, hilos),
        reply_markup=_ikb_mant_nodo(nombre, hilos))
    return MANT_MORE_EST

async def p_mant_hilo(u, c):
    hilo_num = u.message.text.strip()
    c.user_data["_mant_cur_hilo"] = hilo_num
    nombre = c.user_data.get("_mant_cur_est", "")
    nodos  = c.user_data.setdefault("mant_nodos", {})
    hilos  = nodos.setdefault(nombre, {})
    datos  = hilos.setdefault(hilo_num, {"da":"","pa":"","dd":"","pd":""})
    await u.message.reply_text(
        _texto_mant_hilo(nombre, hilo_num, datos),
        reply_markup=_ikb_mant_hilo(datos))
    return MANT_MORE_HILO

async def p_mant_da(u, c):  return await _guardar_campo_mant(u, c, "da")
async def p_mant_pa(u, c):  return await _guardar_campo_mant(u, c, "pa")
async def p_mant_dd(u, c):  return await _guardar_campo_mant(u, c, "dd")
async def p_mant_pd(u, c):  return await _guardar_campo_mant(u, c, "pd")

async def p_mant_more_hilo(u, c):
    nombre   = c.user_data.get("_mant_cur_est",  "")
    hilo_num = c.user_data.get("_mant_cur_hilo", "")
    nodos    = c.user_data.get("mant_nodos", {})
    datos    = nodos.get(nombre, {}).get(hilo_num, {})
    await u.message.reply_text(
        _texto_mant_hilo(nombre, hilo_num, datos),
        reply_markup=_ikb_mant_hilo(datos))
    return MANT_MORE_HILO

async def cb_mant_more_hilo(upd, c):
    q = upd.callback_query; await q.answer()
    nombre   = c.user_data.get("_mant_cur_est",  "")
    hilo_num = c.user_data.get("_mant_cur_hilo", "")
    nodos    = c.user_data.get("mant_nodos", {})
    datos    = nodos.get(nombre, {}).get(hilo_num, {})
    await q.edit_message_text(
        _texto_mant_hilo(nombre, hilo_num, datos),
        reply_markup=_ikb_mant_hilo(datos))
    return MANT_MORE_HILO

async def p_mant_more_est(u, c):
    nombre = c.user_data.get("_mant_cur_est", "")
    nodos  = c.user_data.get("mant_nodos", {})
    hilos  = nodos.get(nombre, {})
    await u.message.reply_text(
        _texto_mant_nodo(nombre, hilos),
        reply_markup=_ikb_mant_nodo(nombre, hilos))
    return MANT_MORE_EST

async def cb_mant_more_est(upd, c):
    q = upd.callback_query; await q.answer()
    nodos = c.user_data.get("mant_nodos", {})
    await q.edit_message_text(
        _texto_mant_panel(nodos),
        reply_markup=_ikb_mant_panel(nodos))
    return MANT_CONFIRM

async def p_mant_obs(u, c):
    t = u.message.text.strip()
    if t.lower() == "omitir":
        c.user_data["mant_obs"] = ""
        await u.message.reply_text("✅ Tabla de mantenimiento completada.\n\n🔧 Sección completada.", reply_markup=KB_MENU)
        return await mostrar_menu(u, c)
    errores = _verificar_ortografia(t)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "mant_obs", "valor": t, "errores": errores,
            "next_state": MENU, "current_state": MANT_OBS,
            "show_mant_done": True,
        }
        await u.message.reply_text(_msg_spell(t, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return MANT_OBS
    c.user_data["mant_obs"] = t
    await u.message.reply_text("✅ Tabla de mantenimiento completada.\n\n🔧 Sección completada.", reply_markup=KB_MENU)
    return await mostrar_menu(u, c)

async def p_conclusiones(u, c):
    t = u.message.text.strip()
    if t.lower() == "omitir":
        c.user_data["conclusiones_lista"] = []
        await u.message.reply_text("✅ Conclusiones omitidas.\n\n📌 Sección completada.", reply_markup=KB_MENU)
        return await mostrar_menu(u, c)
    lista = c.user_data.get("conclusiones_lista", [])
    if len(lista) >= 20:
        await u.message.reply_text(
            f"⚠️ Límite de 20 líneas alcanzado.\n📊 Total: 20/20",
            reply_markup=_ikb_listo("con"))
        return CONCLUSIONES
    errores = _verificar_ortografia(t)
    if errores:
        c.user_data["_spell_pending"] = {
            "campo": "_conclusiones_line", "valor": t, "errores": errores,
            "next_state": CONCLUSIONES, "current_state": CONCLUSIONES,
        }
        await u.message.reply_text(_msg_spell(t, errores), parse_mode="Markdown", reply_markup=_ikb_spell())
        return CONCLUSIONES
    lista.append(t)
    c.user_data["conclusiones_lista"] = lista
    n = len(lista)
    await u.message.reply_text(
        f"✅ Línea {n} guardada.\n📊 Total: {n}/20\n\nEnvía otra línea o toca Listo:",
        reply_markup=_ikb_listo("con"))
    return CONCLUSIONES


# ── Generacion de Excel ────────────────────────────────────────────────────────
def _resize_foto(ruta_orig, w_px, h_px):
    if not PIL_OK or not os.path.exists(ruta_orig): return ruta_orig
    try:
        img = PILImage.open(ruta_orig).convert("RGB")
        img.thumbnail((w_px, h_px), PILImage.LANCZOS)
        ruta_res = ruta_orig.replace(".jpg", "_r.jpg")
        img.save(ruta_res, "JPEG", quality=85)
        return ruta_res
    except Exception as e:
        logger.warning(f"Resize fallido {ruta_orig}: {e}")
        return ruta_orig

def _insertar_fotos_seccion(ws, fotos, fila_ini, w_px=215, h_px=150):
    cols = ["A","B","C","D"]
    for idx, ruta_orig in enumerate(fotos[:8]):
        if not os.path.exists(ruta_orig): continue
        ruta = _resize_foto(ruta_orig, w_px, h_px)
        xl = XLImage(ruta); xl.width = w_px; xl.height = h_px
        xl.anchor = f"{cols[idx % 4]}{fila_ini + idx // 4}"
        ws.add_image(xl)

def generar_excel(d):
    wb = Workbook()
    BORDE      = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"),  bottom=Side(style="thin"))
    AZUL       = "0000FF"
    AMARILLO   = "FFFF99"
    AZUL_CLARO = "9BC2E6"
    GRIS_LBL   = "BFBFBF"
    BLANCO     = "FFFFFF"

    # Crear logo UNA sola vez para reutilizar en todas las pestañas
    _LOGO_PATH = _get_logo_path()

    def _celda(ws, row, col, valor="", bold=False, size=9, color="000000",
               bg=None, h="left", v="center", wrap=True):
        c = ws.cell(row=row, column=col, value=valor)
        c.font = Font(name="Arial", bold=bold, size=size, color=color)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        c.border = BORDE
        return c

    def _cabecera(ws, n_cols, logo_path=None):
        ws.row_dimensions[1].height = 65
        c1 = ws.cell(row=1, column=1)
        c1.border = BORDE
        c1.alignment = Alignment(horizontal="center", vertical="center")
        if logo_path:
            xl = XLImage(logo_path); xl.width = 185; xl.height = 58; xl.anchor = "A1"
            ws.add_image(xl)
        else:
            c1.value = "TELCONET"; c1.font = Font(name="Arial", bold=True, size=14, color="000000")
        for _c in range(2, n_cols): ws.cell(row=1, column=_c).border = BORDE
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n_cols-1)
        c = ws.cell(row=1, column=2, value="REPORTE DE INCIDENCIAS Y MANTENIMIENTOS PARA FIBRA OPTICA")
        c.font = Font(name="Arial", bold=True, size=13)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE
        _celda(ws, 1, n_cols, "Codigo: FOR FO 06\nVersion: 3 (04/08/2021)", bold=True, size=9, h="center")
        ws.row_dimensions[2].height = 5

    def _barra(ws, fila, texto, col_ini, col_fin, bg=AZUL):
        for _c in range(col_ini, col_fin + 1): ws.cell(row=fila, column=_c).border = BORDE
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
        c = ws.cell(row=fila, column=col_ini, value=texto)
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE; ws.row_dimensions[fila].height = 22

    def _barra_sec(ws, fila, texto, col_ini, col_fin, bg=GRIS_LBL, h="left"):
        for _c in range(col_ini, col_fin + 1): ws.cell(row=fila, column=_c).border = BORDE
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
        c = ws.cell(row=fila, column=col_ini, value=texto)
        c.font = Font(name="Arial", bold=True, size=9, color="000000")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
        c.border = BORDE; ws.row_dimensions[fila].height = 18

    def _fila_dato(ws, fila, label, valor, n_cols=4):
        _celda(ws, fila, 1, label, bold=True, size=9, bg=GRIS_LBL)
        for _c in range(2, n_cols + 1): ws.cell(row=fila, column=_c).border = BORDE
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=n_cols)
        _celda(ws, fila, 2, valor, size=9, h="center")
        ws.row_dimensions[fila].height = 35

    def _celdas_foto(ws, fila_ini, n_filas=2, n_cols=4, alto=120):
        for r in range(n_filas):
            ws.row_dimensions[fila_ini+r].height = alto
            for col in range(1, n_cols+1): ws.cell(row=fila_ini+r, column=col).border = BORDE
        return fila_ini + n_filas

    def _section_frame(ws, r1, c1, r2, c2):
        """Borde medio exterior + fino interior. Llamar ANTES de merge_cells."""
        MED = Side(style="medium"); THN = Side(style="thin")
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = Border(
                    left   = MED if c == c1 else THN,
                    right  = MED if c == c2 else THN,
                    top    = MED if r == r1 else THN,
                    bottom = MED if r == r2 else THN,
                )

    # ── HOJA 1 — Reporte_de_incidencias ───────────────────────────────────────
    ws1 = wb.active; ws1.title = "Reporte_de_incidencias"
    ws1.column_dimensions["A"].width = 52; ws1.column_dimensions["B"].width = 52
    ws1.column_dimensions["C"].width = 30; ws1.column_dimensions["D"].width = 22
    _cabecera(ws1, 4, _LOGO_PATH); _barra(ws1, 3, "REPORTE DE INCIDENCIAS EN F.O.", 1, 4)

    ws1.row_dimensions[4].height = 18; ws1.row_dimensions[5].height = 18
    _section_frame(ws1, 4, 1, 5, 4)   # borde medio exterior sección INICIAL (filas 4-5)
    ws1.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    c4a = ws1.cell(row=4, column=1, value="FECHA Y HORA INICIAL DE LA INCIDENCIA")
    c4a.font = Font(name="Arial", bold=True, size=8, color="000000")
    c4a.fill = PatternFill("solid", fgColor=GRIS_LBL)
    c4a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _celda(ws1, 4, 2, "OBSERVACION", bold=True, size=9, h="center")
    _celda(ws1, 4, 3, "FECHA",       bold=True, size=9, h="center")
    _celda(ws1, 4, 4, "HORA",        bold=True, size=9, h="center")
    _celda(ws1, 5, 2, "Incidencia presentada", size=9)
    _celda(ws1, 5, 3, d.get("fecha_inc",""), size=9, h="center")
    _celda(ws1, 5, 4, d.get("hora_inc",""),  size=9, h="center")

    # Fila 6 (Notificacion) forma parte del bloque ACCIONES — columna A queda dentro del merge 6-11
    ws1.row_dimensions[6].height = 18
    _celda(ws1, 6, 2, "Notificacion de Incidencia a F.O.:", size=9, bold=False)
    ws1.cell(row=6, column=3).border = BORDE; ws1.cell(row=6, column=4).border = BORDE
    ws1.merge_cells(start_row=6, start_column=3, end_row=6, end_column=4)
    c6r = ws1.cell(row=6, column=3, value=d.get("codigo",""))
    c6r.font = Font(name="Arial", bold=True, size=9, color="FF0000")
    c6r.alignment = Alignment(horizontal="center", vertical="center")
    c6r.border = BORDE

    for r in range(7, 12): ws1.row_dimensions[r].height = 18
    _section_frame(ws1, 6, 1, 11, 4)  # borde medio exterior sección ACCIONES (filas 6-11)
    ws1.merge_cells(start_row=6, start_column=1, end_row=11, end_column=1)
    c7a = ws1.cell(row=6, column=1, value="FECHA Y HORA DE LAS ACCIONES")
    c7a.font = Font(name="Arial", bold=True, size=8, color="000000")
    c7a.fill = PatternFill("solid", fgColor=GRIS_LBL)
    c7a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _celda(ws1, 7, 2, "OBSERVACION", bold=True, size=9, h="center")
    _celda(ws1, 7, 3, "FECHA",       bold=True, size=9, h="center")
    _celda(ws1, 7, 4, "HORA",        bold=True, size=9, h="center")
    for idx, (obs, fecha, hora) in enumerate([
        ("Arribo al sitio de la Incidencia:", d.get("fecha_arr",""), d.get("hora_arr","")),
        ("Inicio de la reparacion:",           d.get("fecha_rep",""), d.get("hora_rep","")),
        ("Hora de resolucion:",                d.get("fecha_res",""), d.get("hora_res","")),
    ], start=8):
        _celda(ws1, idx, 2, obs, size=9)
        _celda(ws1, idx, 3, fecha, size=9, h="center")
        _celda(ws1, idx, 4, hora,  size=9, h="center")

    _celda(ws1, 11, 2, "Tiempo de resolucion de la incidencia:", size=9, bold=False)
    try:
        from datetime import datetime as _dt
        fi = _dt.strptime(f"{d.get('fecha_inc','')} {d.get('hora_inc','')}","%d/%m/%Y %H:%M")
        fr = _dt.strptime(f"{d.get('fecha_res','')} {d.get('hora_res','')}","%d/%m/%Y %H:%M")
        delta = fr - fi
        h_tot = int(delta.total_seconds()//3600); m_tot = int((delta.total_seconds()%3600)//60)
        s_tot = int(delta.total_seconds()%60)
        meses = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        _fmt = lambda dt: f"{dt.day} {meses[dt.month]} {str(dt.year)[2:]}"
        rango = f"{_fmt(fi)}// {_fmt(fr)}"; duracion = f"{h_tot}:{m_tot:02d}:{s_tot:02d}"
    except:
        rango = f"{d.get('fecha_inc','')} // {d.get('fecha_res','')}"; duracion = ""
    for col, val in [(3, rango),(4, duracion)]:
        c = ws1.cell(row=11, column=col, value=val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDE

    n_fa = len(d.get("fotos_antes",[])); n_fd = len(d.get("fotos_durante",[]))
    n_ff = len(d.get("fotos_fin",[]));   n_ta = len(d.get("trazas_antes",[]))
    n_td = len(d.get("trazas_despues",[]))
    campos = [
        (12,"MOTIVO APARENTE DE LA INCIDENCIA",          d.get("motivo","")),
        (13,"INFORMACION DE TESTIGOS DEL INCIDENTE",     d.get("testigos","")),
        (14,"REMEDIO DEFINITIVO A LA INCIDENCIA",        d.get("remedio","")),
        (15,"ENLACES CAIDOS DEBIDO A LA INCIDENCIA",     d.get("enlaces","")),
        (16,"PERSONAL QUE INTERVINO EN LA SOLUCION",     d.get("personal","")),
        (17,"LUGAR DE LA INCIDENCIA",                    d.get("lugar","")),
        (18,"COORDENADAS",                               d.get("coords","No disponible")),
        (19,"FOTOS ANEXAS AL REPORTE",
             f"Antes:{n_fa} | Durante:{n_fd} | Fin:{n_ff}  (Total:{n_fa+n_fd+n_ff})"),
        (20,"TRAZAS ANEXAS DE OTDR",
             f"Antes:{n_ta} | Despues:{n_td}  (Total:{n_ta+n_td})"),
        (21,"RESUMEN DESCRIPTIVO DE LA INCIDENCIA",      "SI" if (d.get("resumen") or d.get("resumen_lista")) else "NO"),
        (22,"CONCLUSIONES TRABAJOS PROGRAMADOS",         "SI" if (d.get("conclusiones") or d.get("conclusiones_lista")) else "NO"),
        (23,"TABLA DE MANTENIMIENTOS",                   "SI" if (d.get("mant_nodos") or d.get("mant_filas")) else "NO"),
        (24,"OBSERVACIONES",                             d.get("obs","")),
    ]
    hipervinculos = {19:"Fotos_anexas_al_reporte",20:"Trazas_anexas_del_OTDR",
                     21:"Resumen_descriptivo_incidencia",22:"Conclusiones",23:"Tabla_para_mantenimiento"}
    for fila, label, valor in campos:
        _fila_dato(ws1, fila, label, valor, n_cols=4)
        if fila in hipervinculos:
            cl = ws1.cell(row=fila, column=1)
            cl.hyperlink = f"#{hipervinculos[fila]}!A1"
            cl.font = Font(name="Arial", bold=True, size=9, color="0000FF", underline="single")
    n_coords = len(d.get("coords_lista",[]))
    if n_coords > 1: ws1.row_dimensions[18].height = max(28, n_coords*16)
    ws1.row_dimensions[26].height = 8
    for r, txt in [(27,"CORDIALMENTE,"),(28,"Departamento de Fibra Optica"),(29,"TELCONET S.A.")]:
        c = ws1.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", bold=(r>27), size=10); ws1.row_dimensions[r].height = 16

    # ── HOJA 2 — Fotos_anexas_al_reporte ──────────────────────────────────────
    ws2 = wb.create_sheet("Fotos_anexas_al_reporte")
    for letra in ["A","B","C","D"]: ws2.column_dimensions[letra].width = 52
    _cabecera(ws2, 4, _LOGO_PATH); _barra(ws2, 3, "FOTOS DE LA INCIDENCIA", 1, 4)
    fila = 4
    for titulo_sec, key_sec, fila_fotos in [
        ("ANTES DE LA REPARACION","fotos_antes",5),
        ("DURANTE LA REPARACION","fotos_durante",9),
        ("FIN DE LA REPARACION","fotos_fin",13)]:
        _barra_sec(ws2, fila, titulo_sec, 1, 4, bg="D9D9D9", h="center"); fila += 1
        _celdas_foto(ws2, fila, n_filas=2, n_cols=4, alto=175)
        _insertar_fotos_seccion(ws2, d.get(key_sec,[]), fila_ini=fila_fotos, w_px=340, h_px=220)
        fila += 3

    # ── HOJA 3 — Trazas_anexas_del_OTDR ───────────────────────────────────────
    ws3 = wb.create_sheet("Trazas_anexas_del_OTDR")
    for letra in ["A","B","C","D"]: ws3.column_dimensions[letra].width = 52
    _cabecera(ws3, 4, _LOGO_PATH); _barra(ws3, 3, "MEDICIONES DE OTDR", 1, 4)
    fila = 4
    for titulo_sec, key_sec, fila_trazas in [
        ("ANTES DE LA REPARACION","trazas_antes",5),
        ("DESPUES DE LA REPARACION","trazas_despues",9)]:
        _barra_sec(ws3, fila, titulo_sec, 1, 4, bg="D9D9D9", h="center"); fila += 1
        _celdas_foto(ws3, fila, n_filas=2, n_cols=4, alto=185)
        _insertar_fotos_seccion(ws3, d.get(key_sec,[]), fila_ini=fila_trazas, w_px=340, h_px=240)
        fila += 3

    # ── HOJA 4 — Resumen_descriptivo_incidencia ────────────────────────────────
    ws4 = wb.create_sheet("Resumen_descriptivo_incidencia")
    ws4.column_dimensions["A"].width = 30; ws4.column_dimensions["B"].width = 85
    ws4.column_dimensions["C"].width = 22
    _cabecera(ws4, 3, _LOGO_PATH); _barra(ws4, 3, "RESUMEN DESCRIPTIVO DE LA INCIDENCIA", 1, 3)
    # Usar lista si existe, si no usar campo legacy
    lineas_res = d.get("resumen_lista", [])
    if not lineas_res:
        txt_res = d.get("resumen","")
        lineas_res = [l.strip() for l in txt_res.split("\n") if l.strip()] if txt_res else []
    lineas = lineas_res if lineas_res else ["(Sin resumen registrado)"]
    for i, linea in enumerate(lineas, start=4):
        ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        for col in range(1,4): ws4.cell(row=i, column=col).border = BORDE
        c = ws4.cell(row=i, column=1, value=linea)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws4.row_dimensions[i].height = 20

    # ── HOJA 5 — Tabla_para_mantenimiento ─────────────────────────────────────
    ws5 = wb.create_sheet("Tabla_para_mantenimiento")
    anchos = {"A":22,"B":22,"C":10,"D":22,"E":24,"F":22,"G":24,"H":35,"I":10,"J":20,"K":18}
    for cl, w in anchos.items(): ws5.column_dimensions[cl].width = w
    ws5.row_dimensions[1].height = 60
    ws5.cell(row=1, column=1).border = BORDE; ws5.cell(row=1, column=2).border = BORDE
    ws5.merge_cells("A1:B1")
    c1_ws5 = ws5.cell(row=1, column=1)
    c1_ws5.border = BORDE; c1_ws5.alignment = Alignment(horizontal="center", vertical="center")
    _logo_tmp5 = _LOGO_PATH
    if _logo_tmp5:
        xl5 = XLImage(_logo_tmp5); xl5.width = 160; xl5.height = 52; xl5.anchor = "A1"
        ws5.add_image(xl5)
    else:
        c1_ws5.value = "TELCONET"; c1_ws5.font = Font(name="Arial", bold=True, size=14)
    for _c in range(3, 11): ws5.cell(row=1, column=_c).border = BORDE
    ws5.merge_cells("C1:J1")
    ct = ws5.cell(row=1, column=3, value="REPORTE DE INCIDENCIAS Y MANTENIMIENTOS PARA FIBRA ÓPTICA")
    ct.font = Font(name="Arial", bold=True, size=12)
    ct.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ct.border = BORDE
    ck = ws5.cell(row=1, column=11, value="Código: FOR FO 06\nVersión: 3 (04/08/2021)")
    ck.font = Font(name="Arial", bold=True, size=9)
    ck.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ck.border = BORDE
    ws5.row_dimensions[2].height = 6

    CAM = "ED7D31"; CBas = "404040"; YELL = "FFFF99"
    ws5.row_dimensions[3].height = 30; ws5.row_dimensions[4].height = 36
    for col, txt in [(1,"Mantenimiento"),(2,"Estación"),(3,"Hilos")]:
        for _r in [3,4]: ws5.cell(row=_r, column=col).border = BORDE
        ws5.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        c = ws5.cell(row=3, column=col, value=txt)
        c.font = Font(name="Arial", bold=True, italic=True, size=9, color=CBas)
        c.fill = PatternFill("solid", fgColor=YELL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDE
    for (c1,c2,txt) in [(4,5,"Valores Medidos Antes"),(6,7,"Valores Medidos Después"),(8,11,"Observaciones y Novedades Luego del Trabajo")]:
        for _col in range(c1, c2+1): ws5.cell(row=3, column=_col).border = BORDE
        ws5.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
        c = ws5.cell(row=3, column=c1, value=txt)
        c.font = Font(name="Arial", bold=True, italic=True, size=9, color=CAM)
        c.fill = PatternFill("solid", fgColor=YELL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDE
    for cn, txt in [(4,"Distancia (Km)"),(5,"Pérdida en la\nlínea (dB)"),(6,"Distancia (Km)"),
                    (7,"Pérdida en la\nlínea (dB)"),(8,"Según lecturas\nde OTDR"),
                    (9,"Hilo"),(10,"Distancias\n(Km)"),(11,"Pérdidas\n(dB)")]:
        c = ws5.cell(row=4, column=cn, value=txt)
        c.font = Font(name="Arial", bold=True, italic=True, size=8, color=CAM)
        c.fill = PatternFill("solid", fgColor=YELL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDE

    filas_mant = d.get("mant_filas",[])
    obs_mant   = d.get("mant_obs","")
    if not filas_mant and d.get("mant_est"):
        filas_mant = [{"est":d.get("mant_est",""),"hilo":d.get("mant_hilo",""),
                       "da":d.get("mant_da",""),"pa":d.get("mant_pa",""),
                       "dd":d.get("mant_dd",""),"pd":d.get("mant_pd","")}]
    if not filas_mant:
        filas_mant = [{"est":"(No aplica)","hilo":"","da":"","pa":"","dd":"","pd":""}]
        obs_mant   = "Tabla no aplicable para esta incidencia"
    n_filas = len(filas_mant); FILA_INI = 5
    grupos_est = []
    for idx, fila in enumerate(filas_mant):
        est = fila.get("est","")
        if grupos_est and grupos_est[-1][0] == est: grupos_est[-1][1].append(idx)
        else: grupos_est.append([est,[idx]])
    for i, fila in enumerate(filas_mant):
        row = FILA_INI + i; ws5.row_dimensions[row].height = 22
        for col in range(1,12): ws5.cell(row=row, column=col).border = BORDE
        c = ws5.cell(row=row, column=3, value=fila.get("hilo",""))
        c.font = Font(name="Arial", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDE
        for col, key in [(4,"da"),(5,"pa"),(6,"dd"),(7,"pd")]:
            c = ws5.cell(row=row, column=col, value=fila.get(key,""))
            c.font = Font(name="Arial", size=9); c.fill = PatternFill("solid", fgColor=BLANCO)
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDE
        for col, key in [(9,"hilo"),(10,"dd"),(11,"pd")]:
            c = ws5.cell(row=row, column=col, value=fila.get(key,""))
            c.font = Font(name="Arial", bold=(key=="hilo"), size=9)
            c.fill = PatternFill("solid", fgColor=BLANCO)
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDE
    if n_filas > 1:
        ws5.merge_cells(start_row=FILA_INI, start_column=1, end_row=FILA_INI+n_filas-1, end_column=1)
    for r in range(FILA_INI, FILA_INI+n_filas): ws5.cell(row=r, column=1).border = BORDE
    c_t = ws5.cell(row=FILA_INI, column=1, value="TRAMO")
    c_t.font = Font(name="Arial", bold=True, size=10)
    c_t.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    c_t.alignment = Alignment(horizontal="center", vertical="center"); c_t.border = BORDE
    for nombre_est, indices in grupos_est:
        r_ini = FILA_INI+indices[0]; r_fin = FILA_INI+indices[-1]
        if len(indices) > 1:
            ws5.merge_cells(start_row=r_ini, start_column=2, end_row=r_fin, end_column=2)
        for r in range(r_ini, r_fin+1): ws5.cell(row=r, column=2).border = BORDE
        c = ws5.cell(row=r_ini, column=2, value=nombre_est)
        c.font = Font(name="Arial", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDE
    if n_filas > 1:
        ws5.merge_cells(start_row=FILA_INI, start_column=8, end_row=FILA_INI+n_filas-1, end_column=8)
    for r in range(FILA_INI, FILA_INI+n_filas): ws5.cell(row=r, column=8).border = BORDE
    c_o = ws5.cell(row=FILA_INI, column=8, value=obs_mant)
    c_o.font = Font(name="Arial", size=9); c_o.fill = PatternFill("solid", fgColor=BLANCO)
    c_o.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c_o.border = BORDE

    # ── HOJA 6 — Conclusiones ──────────────────────────────────────────────────
    ws6 = wb.create_sheet("Conclusiones")
    ws6.column_dimensions["A"].width = 30; ws6.column_dimensions["B"].width = 85
    ws6.column_dimensions["C"].width = 22
    _cabecera(ws6, 3, _LOGO_PATH); _barra(ws6, 3, "CONCLUSIONES", 1, 3)
    # Usar lista si existe, si no usar campo legacy
    lineas_con = d.get("conclusiones_lista", [])
    if not lineas_con:
        txt_con = d.get("conclusiones","")
        lineas_con = [l.strip() for l in txt_con.split("\n") if l.strip()] if txt_con else []
    lineas = lineas_con if lineas_con else ["(Sin conclusiones registradas)"]
    for i, linea in enumerate(lineas, start=4):
        ws6.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        for col in range(1,4): ws6.cell(row=i, column=col).border = BORDE
        c = ws6.cell(row=i, column=1, value=linea)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws6.row_dimensions[i].height = 20

    codigo = d.get("codigo","SIN_CODIGO").replace("/","-").replace(":","-")
    cuadrilla = d.get("cuadrilla","SIN_CUADRILLA").replace(" ","_")
    fecha_hoy = datetime.now().strftime('%d%m%Y')
    nombre_archivo = f"TT_INC_{codigo}-{fecha_hoy}-{cuadrilla}"
    ruta = os.path.join(SALIDA, f"{nombre_archivo}.xlsx")
    wb.save(ruta); return ruta

# ── CALLBACK UNIFICADO LISTO/OMITIR ───────────────────────────────────────────
async def cb_listo_seccion(upd, c):
    """Maneja todos los botones Listo/Omitir inline de secciones."""
    q = upd.callback_query; await q.answer()
    partes = q.data.split(":")
    sec    = partes[1]  # coord, fa, fd, ff, ta, td, gen
    accion = partes[2]  # l=listo, o=omitir

    async def _siguiente(texto, estado_sig, sec_sig=None):
        await q.edit_message_text(texto)
        if estado_sig == MENU:
            return await mostrar_menu(upd, c)
        if sec_sig:
            await c.bot.send_message(q.message.chat_id,
                f"Envia imagenes o toca Listo/Omitir.",
                reply_markup=_ikb_listo(sec_sig))
        return estado_sig

    if sec == "coord":
        lista = c.user_data.get("coords_lista", [])
        c.user_data["coords"] = "\n".join(lista) if lista else "No disponible"
        c.user_data["_causas_campo"] = "obs"
        await q.edit_message_text(f"✅ {len(lista)} coordenada(s) guardadas.\n\nObservaciones — selecciona:",
            reply_markup=_ikb_lista(CAUSAS, "ca"))
        return OBS
    elif sec == "fa":
        n = len(c.user_data.get("fotos_antes",[]))
        await q.edit_message_text(f"✅ {n} fotos ANTES guardadas.")
        await c.bot.send_message(q.message.chat_id,
            "FOTOS — DURANTE la reparacion\nEnvia fotos.",
            reply_markup=_ikb_listo("fd"))
        return FOTOS_DURANTE
    elif sec == "fd":
        n = len(c.user_data.get("fotos_durante",[]))
        await q.edit_message_text(f"✅ {n} fotos DURANTE guardadas.")
        await c.bot.send_message(q.message.chat_id,
            "FOTOS — FIN de la reparacion\nEnvia fotos.",
            reply_markup=_ikb_listo("ff"))
        return FOTOS_FIN
    elif sec == "ff":
        n = len(c.user_data.get("fotos_fin",[]))
        await q.edit_message_text(f"✅ {n} fotos FIN guardadas.\n\n📸 Sección completada.")
        return await mostrar_menu(upd, c)
    elif sec == "ta":
        n = len(c.user_data.get("trazas_antes",[]))
        await q.edit_message_text(f"✅ {n} trazas ANTES guardadas.")
        await c.bot.send_message(q.message.chat_id,
            "TRAZAS OTDR — DESPUES de la reparacion\nEnvia capturas.",
            reply_markup=_ikb_listo("td"))
        return TRAZAS_DESPUES
    elif sec == "td":
        n = len(c.user_data.get("trazas_despues",[]))
        await q.edit_message_text(f"✅ {n} trazas DESPUES guardadas.\n\n📡 Sección completada.")
        return await mostrar_menu(upd, c)
    elif sec == "res":
        lista = c.user_data.get("resumen_lista", [])
        c.user_data["resumen"] = "\n".join(lista)
        await q.edit_message_text(f"✅ {len(lista)} línea(s) de resumen guardadas.\n\n📝 Sección completada.")
        return await mostrar_menu(upd, c)
    elif sec == "con":
        lista = c.user_data.get("conclusiones_lista", [])
        c.user_data["conclusiones"] = "\n".join(lista)
        await q.edit_message_text(f"✅ {len(lista)} conclusión(es) guardadas.\n\n📌 Sección completada.")
        return await mostrar_menu(upd, c)
    else:
        return await mostrar_menu(upd, c)

async def cb_confirmar(upd, c):
    """Callback para confirmar/cancelar generación del Excel."""
    q = upd.callback_query; await q.answer()
    if q.data.split(":")[1] == "no":
        await q.edit_message_text("Cancelado. Usa el menú para continuar.")
        return await mostrar_menu(upd, c)
    await q.edit_message_text("Generando reporte Excel...")
    try:
        # Sincronizar mant_filas desde mant_nodos antes de generar
        nodos = c.user_data.get("mant_nodos", {})
        if nodos:
            c.user_data["mant_filas"] = _mant_nodos_a_filas(nodos)
        ruta = generar_excel(c.user_data)
        d = c.user_data
        n_fa = len(d.get("fotos_antes",[])); n_fd = len(d.get("fotos_durante",[]))
        n_ff = len(d.get("fotos_fin",[]));   n_ta = len(d.get("trazas_antes",[]))
        n_td = len(d.get("trazas_despues",[])); total = n_fa+n_fd+n_ff+n_ta+n_td
        with open(ruta,"rb") as f:
            await c.bot.send_document(
                chat_id=q.message.chat_id,
                document=f,
                filename=os.path.basename(ruta),
                caption=(f"✅ Reporte generado\nCaso: {d.get('codigo','')}\n"
                         f"Fotos: Antes={n_fa} | Durante={n_fd} | Fin={n_ff}\n"
                         f"OTDR:  Antes={n_ta} | Despues={n_td}\n"
                         f"Total: {total}\nEscribe /start para nuevo reporte."))

        # Enviar a los coordinadores
        caption_coord = (
            f"📋 *Nuevo reporte generado*\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"📁 Caso:       {d.get('codigo','—')}\n"
            f"👥 Cuadrilla:  {d.get('cuadrilla','—')}\n"
            f"📍 Zona:       {d.get('cuadrilla_zona','—')}\n"
            f"📅 Incidencia: {d.get('fecha_inc','')} {d.get('hora_inc','')}\n"
            f"📅 Resolución: {d.get('fecha_res','')} {d.get('hora_res','')}\n"
            f"👤 Técnico:    {d.get('auth_email','—')}")
        for coord_id in COORDINADORES_TG_IDS:
            try:
                with open(ruta,"rb") as f:
                    await c.bot.send_document(
                        chat_id=coord_id,
                        document=f,
                        filename=os.path.basename(ruta),
                        caption=caption_coord,
                        parse_mode="Markdown")
            except Exception as e_coord:
                logger.warning(f"No se pudo notificar al coordinador {coord_id}: {e_coord}")
    except Exception as e:
        logger.error(f"Error Excel: {e}", exc_info=True)
        await c.bot.send_message(q.message.chat_id, f"Error al generar: {e}")
    return ConversationHandler.END

# ── MENÚ PRINCIPAL POR PESTAÑAS ───────────────────────────────────────────────
KB_MENU = ReplyKeyboardMarkup([
    ["📋 Datos del reporte"],
    ["📸 Fotos del reporte"],
    ["📡 Trazas OTDR"],
    ["📝 Resumen descriptivo"],
    ["🔧 Tabla de mantenimiento"],
    ["📌 Conclusiones"],
    ["✅ GENERAR EXCEL"],
], resize_keyboard=True, one_time_keyboard=False, is_persistent=True)

# ── Inline keyboard de pestañas (siempre visible al lado de mensajes clave) ────
def _ikb_tabs(d=None):
    """Teclado inline con las 7 pestañas + indicadores de estado."""
    if d is None: d = {}
    s1 = "✅" if d.get("obs") is not None and d.get("codigo") else "⬜"
    n_f = len(d.get("fotos_antes",[])) + len(d.get("fotos_durante",[])) + len(d.get("fotos_fin",[]))
    s2 = f"✅{n_f}📸" if "fotos_antes" in d else "⬜"
    n_t = len(d.get("trazas_antes",[])) + len(d.get("trazas_despues",[]))
    s3 = f"✅{n_t}📡" if "trazas_antes" in d else "⬜"
    n_r = len(d.get("resumen_lista", []))
    s4 = f"✅{n_r}📝" if n_r > 0 else "⬜"
    s5 = f"✅{len(d.get('mant_filas',[]))}🔧" if d.get("mant_obs") is not None else "⬜"
    n_c = len(d.get("conclusiones_lista", []))
    s6 = f"✅{n_c}📌" if n_c > 0 else "⬜"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"📋 Datos {s1}",        callback_data="tab:datos")],
        [InlineKeyboardButton(f"📸 Fotos {s2}",        callback_data="tab:fotos")],
        [InlineKeyboardButton(f"📡 Trazas {s3}",       callback_data="tab:trazas")],
        [InlineKeyboardButton(f"📝 Resumen {s4}",      callback_data="tab:resumen")],
        [InlineKeyboardButton(f"🔧 Mantenimiento {s5}",callback_data="tab:mant")],
        [InlineKeyboardButton(f"📌 Conclusiones {s6}", callback_data="tab:conclusiones")],
        [InlineKeyboardButton("✅ GENERAR EXCEL",       callback_data="tab:excel")],
    ])

def _ikb_fotos_sel(d):
    """Teclado inline para elegir qué sub-sección de fotos llenar."""
    n_a = len(d.get("fotos_antes",  []))
    n_d = len(d.get("fotos_durante", []))
    n_f = len(d.get("fotos_fin",     []))
    def ic(n): return "✅" if n >= 1 else "⬜"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"{ic(n_a)} 📸 ANTES     {n_a}/8 fotos",   callback_data="fsec:fa")],
        [InlineKeyboardButton(f"{ic(n_d)} 📸 DURANTE   {n_d}/8 fotos",   callback_data="fsec:fd")],
        [InlineKeyboardButton(f"{ic(n_f)} 📸 AL FINAL  {n_f}/8 fotos",   callback_data="fsec:ff")],
        [InlineKeyboardButton("✅ Listo con fotos", callback_data="fsec:listo")],
    ])

def _ikb_trazas_sel(d):
    """Teclado inline para elegir qué sub-sección de trazas llenar."""
    n_a = len(d.get("trazas_antes",   []))
    n_d = len(d.get("trazas_despues", []))
    def ic(n): return "✅" if n >= 1 else "⬜"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"{ic(n_a)} 📡 ANTES    {n_a}/8 trazas",  callback_data="tsec:ta")],
        [InlineKeyboardButton(f"{ic(n_d)} 📡 DESPUÉS  {n_d}/8 trazas",  callback_data="tsec:td")],
        [InlineKeyboardButton("✅ Listo con trazas", callback_data="tsec:listo")],
    ])

def _ikb_resumen_acc(n):
    """Acciones para resumen/conclusiones."""
    btns = [[InlineKeyboardButton("➕ Agregar línea",   callback_data="racc:add")]]
    if n > 0:
        btns.append([InlineKeyboardButton("🗑 Borrar todo", callback_data="racc:clear")])
    btns.append([InlineKeyboardButton("✅ Listo",        callback_data="racc:listo")])
    return InlineKeyboardMarkup(btns)

def _ikb_concl_acc(n):
    """Acciones para conclusiones."""
    btns = [[InlineKeyboardButton("➕ Agregar línea",   callback_data="cacc:add")]]
    if n > 0:
        btns.append([InlineKeyboardButton("🗑 Borrar todo", callback_data="cacc:clear")])
    btns.append([InlineKeyboardButton("✅ Listo",        callback_data="cacc:listo")])
    return InlineKeyboardMarkup(btns)

# Teclados de sección — incluyen botón de regreso al menú
# Estos son INLINE — no reemplazan el menú persistente de abajo
SI_NO  = IKB_MANT_CONF   # alias para compatibilidad
OMITIR = InlineKeyboardMarkup([[InlineKeyboardButton("⏭ Omitir sección", callback_data="lo:gen:o")]])
LISTO  = InlineKeyboardMarkup([[
    InlineKeyboardButton("✅ Listo",  callback_data="lo:gen:l"),
    InlineKeyboardButton("⏭ Omitir", callback_data="lo:gen:o"),
]])

async def cb_fotos_sec(upd, c):
    """Callback: usuario elige sub-sección de fotos (Antes/Durante/Fin)."""
    q = upd.callback_query; await q.answer()
    sec = q.data.split(":")[1]
    d = c.user_data

    if sec == "listo":
        await q.edit_message_text(
            f"📸 Fotos guardadas: "
            f"Antes={len(d.get('fotos_antes',[]))} | "
            f"Durante={len(d.get('fotos_durante',[]))} | "
            f"Fin={len(d.get('fotos_fin',[]))}\n\n"
            "✅ Sección fotos completada.")

        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
                self.text = ""; self.location = None; self.photo = None
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)

        class _FU:
            def __init__(self, q):
                self.message = _FM(q.message.chat_id, q.bot)
                self.effective_user = q.from_user

        return await mostrar_menu(_FU(q), c)

    MAP = {"fa": (FOTOS_ANTES,   "fa", "ANTES"),
           "fd": (FOTOS_DURANTE, "fd", "DURANTE"),
           "ff": (FOTOS_FIN,     "ff", "AL FINAL")}
    estado, sec_btn, nombre = MAP[sec]
    n = len(d.get({"fa":"fotos_antes","fd":"fotos_durante","ff":"fotos_fin"}[sec], []))
    await q.edit_message_text(
        f"📸 FOTOS — {nombre} de la reparación\n"
        f"📊 Ya tienes {n}/8 fotos en esta sección.\n\n"
        "Envía fotos o toca Listo:",
        reply_markup=_ikb_listo(sec_btn))
    return estado

async def cb_trazas_sec(upd, c):
    """Callback: usuario elige sub-sección de trazas (Antes/Después)."""
    q = upd.callback_query; await q.answer()
    sec = q.data.split(":")[1]
    d = c.user_data

    async def _edit(text, **kw):
        if q.message.photo:
            await q.edit_message_caption(caption=text, **kw)
        else:
            await q.edit_message_text(text, **kw)

    if sec == "listo":
        await _edit(
            f"📡 Trazas guardadas: "
            f"Antes={len(d.get('trazas_antes',[]))} | "
            f"Después={len(d.get('trazas_despues',[]))}\n\n"
            "✅ Sección trazas completada.")

        class _FM:
            def __init__(self, chat_id, bot):
                self.chat_id = chat_id; self.bot = bot
                self.text = ""; self.location = None; self.photo = None
            async def reply_text(self, txt, **kw):
                await self.bot.send_message(self.chat_id, txt, **kw)

        class _FU:
            def __init__(self, q):
                self.message = _FM(q.message.chat_id, q.bot)
                self.effective_user = q.from_user

        return await mostrar_menu(_FU(q), c)

    MAP = {"ta": (TRAZAS_ANTES,   "ta", "ANTES"),
           "td": (TRAZAS_DESPUES, "td", "DESPUÉS")}
    estado, sec_btn, nombre = MAP[sec]
    n = len(d.get({"ta":"trazas_antes","td":"trazas_despues"}[sec], []))
    await _edit(
        f"📡 TRAZAS — {nombre} de la reparación\n"
        f"📊 Ya tienes {n}/8 trazas en esta sección.\n\n"
        "Envía capturas o toca Listo:",
        reply_markup=_ikb_listo(sec_btn))
    return estado

async def cb_racc(upd, c):
    """Callback acciones de resumen: agregar / borrar / listo."""
    q = upd.callback_query; await q.answer()
    acc = q.data.split(":")[1]
    lista = c.user_data.get("resumen_lista", [])

    class _FM:
        def __init__(self, chat_id, bot):
            self.chat_id = chat_id; self.bot = bot
            self.text = ""; self.location = None; self.photo = None
        async def reply_text(self, txt, **kw):
            await self.bot.send_message(self.chat_id, txt, **kw)

    class _FU:
        def __init__(self, q):
            self.message = _FM(q.message.chat_id, q.bot)
            self.effective_user = q.from_user

    if acc == "clear":
        c.user_data["resumen_lista"] = []
        await q.edit_message_text("🗑 Resumen borrado.")
        return await mostrar_menu(_FU(q), c)
    if acc == "listo":
        await q.edit_message_text(f"✅ Resumen guardado. {len(lista)} línea(s).")
        return await mostrar_menu(_FU(q), c)
    # "add" → entrar a estado RESUMEN para escribir
    n = len(lista)
    preview = ""
    if lista:
        preview = "\n📋 Ya guardadas:\n" + "\n".join(
            f"  {i+1}. {l[:55]}{'…' if len(l)>55 else ''}" for i,l in enumerate(lista))
    await q.edit_message_text(
        f"📝 RESUMEN — Escribe la línea {n+1}{preview}\n\nEnvía el texto o toca Listo:",
        reply_markup=_ikb_listo("res"))
    return RESUMEN

async def cb_cacc(upd, c):
    """Callback acciones de conclusiones: agregar / borrar / listo."""
    q = upd.callback_query; await q.answer()
    acc = q.data.split(":")[1]
    lista = c.user_data.get("conclusiones_lista", [])

    class _FM:
        def __init__(self, chat_id, bot):
            self.chat_id = chat_id; self.bot = bot
            self.text = ""; self.location = None; self.photo = None
        async def reply_text(self, txt, **kw):
            await self.bot.send_message(self.chat_id, txt, **kw)

    class _FU:
        def __init__(self, q):
            self.message = _FM(q.message.chat_id, q.bot)
            self.effective_user = q.from_user

    if acc == "clear":
        c.user_data["conclusiones_lista"] = []
        await q.edit_message_text("🗑 Conclusiones borradas.")
        return await mostrar_menu(_FU(q), c)
    if acc == "listo":
        await q.edit_message_text(f"✅ Conclusiones guardadas. {len(lista)} línea(s).")
        return await mostrar_menu(_FU(q), c)
    # "add" → entrar a estado CONCLUSIONES para escribir
    n = len(lista)
    preview = ""
    if lista:
        preview = "\n📋 Ya guardadas:\n" + "\n".join(
            f"  {i+1}. {l[:55]}{'…' if len(l)>55 else ''}" for i,l in enumerate(lista))
    await q.edit_message_text(
        f"📌 CONCLUSIONES — Escribe la línea {n+1}{preview}\n\nEnvía el texto o toca Listo:",
        reply_markup=_ikb_listo("con"))
    return CONCLUSIONES

async def cb_tabs(upd, c):
    """Callback de los botones inline de pestañas — navega a cualquier sección."""
    q = upd.callback_query; await q.answer()
    tab = q.data.split(":")[1]
    d   = c.user_data

    # Actualizar el teclado inline con el estado actual
    try:
        await q.edit_message_reply_markup(reply_markup=_ikb_tabs(d))
    except Exception:
        pass

    # Simular la selección del menú según la pestaña tocada
    class _FakeMsg:
        def __init__(self, chat_id, bot, text):
            self.chat_id = chat_id; self.bot = bot; self.text = text
            self.location = None; self.photo = None
        async def reply_text(self, txt, **kw):
            await self.bot.send_message(self.chat_id, txt, **kw)

    class _FakeUpd:
        def __init__(self, q):
            self.message = _FakeMsg(q.message.chat_id, q.bot, "")
            self.effective_user = q.from_user

    fu = _FakeUpd(q)

    TAB_MAP = {
        "datos":       "📋 Datos del reporte",
        "fotos":       "📸 Fotos del reporte",
        "trazas":      "📡 Trazas OTDR",
        "resumen":     "📝 Resumen descriptivo",
        "mant":        "🔧 Tabla de mantenimiento",
        "conclusiones":"📌 Conclusiones",
        "excel":       "✅ GENERAR EXCEL",
    }
    fu.message.text = TAB_MAP.get(tab, "")
    return await p_menu(fu, c)

async def mostrar_menu(u, c):
    d = c.user_data
    # Indicadores de estado por sección
    s1 = "✅" if d.get("obs") is not None and d.get("codigo") else "⏳"
    n_fotos = len(d.get("fotos_antes",[])) + len(d.get("fotos_durante",[])) + len(d.get("fotos_fin",[]))
    s2 = f"✅ {n_fotos}📸" if "fotos_antes" in d else "⏳"
    n_trazas = len(d.get("trazas_antes",[])) + len(d.get("trazas_despues",[]))
    s3 = f"✅ {n_trazas}📡" if "trazas_antes" in d else "⏳"
    s4_n = len(d.get("resumen_lista", []))
    s4 = f"✅ {s4_n}📝" if s4_n > 0 else "⏳"
    s5 = f"✅ {len(d.get('mant_filas',[]))}🔧" if d.get("mant_obs") is not None else "⏳"
    s6_n = len(d.get("conclusiones_lista", []))
    s6 = f"✅ {s6_n}📌" if s6_n > 0 else "⏳"

    texto = (
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "📋  MENÚ — PESTAÑAS DEL INFORME\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"1️⃣  Datos del reporte       {s1}\n"
        f"2️⃣  Fotos del reporte       {s2}\n"
        f"3️⃣  Trazas OTDR             {s3}\n"
        f"4️⃣  Resumen descriptivo     {s4}\n"
        f"5️⃣  Tabla de mantenimiento  {s5}\n"
        f"6️⃣  Conclusiones            {s6}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "Usa los botones de abajo ↓ o los botones\n"
        "inline ↕ para ir a cualquier pestaña:"
    )
    # Enviar con teclado persistente (abajo) + teclado inline (pestañas visibles)
    await u.message.reply_text(texto, reply_markup=KB_MENU)
    await u.message.reply_text(
        "🗂 Acceso rápido a pestañas:",
        reply_markup=_ikb_tabs(d))
    return MENU

async def p_menu(u, c):
    opcion = u.message.text.strip()

    if opcion == "📋 Datos del reporte":
        d = c.user_data
        tiene_datos = bool(d.get("codigo") or d.get("cuadrilla"))
        if tiene_datos:
            campos_ok = sum(bool(d.get(k)) for k in
                ["codigo","cuadrilla","fecha_inc","fecha_res","personal","lugar","motivo","obs"])
            # Calcular campos vacíos para mostrar avance
            campos_vacios = 8 - campos_ok
            avance = "✅ ¡Todo completo!" if campos_vacios == 0 else f"⏳ Faltan {campos_vacios} campo(s)"
            historial = (
                f"📋 DATOS DEL REPORTE — HISTORIAL\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"📁 Caso:        {d.get('codigo','—')}\n"
                f"👥 Cuadrilla:   {d.get('cuadrilla','—')}\n"
                f"📍 Lugar:       {str(d.get('lugar','—'))[:40]}\n"
                f"📅 Incidencia:  {d.get('fecha_inc','—')} {d.get('hora_inc','—')}\n"
                f"📅 Arribo:      {d.get('fecha_arr','—')} {d.get('hora_arr','—')}\n"
                f"📅 Reparación:  {d.get('fecha_rep','—')} {d.get('hora_rep','—')}\n"
                f"📅 Resolución:  {d.get('fecha_res','—')} {d.get('hora_res','—')}\n"
                f"⚡ Motivo:      {str(d.get('motivo','—'))[:40]}\n"
                f"🔗 Enlace:      {str(d.get('enlaces','—'))[:40]}\n"
                f"👤 Personal:    {str(d.get('personal','—'))[:40]}\n"
                f"📝 Observación: {str(d.get('obs','—'))[:50]}\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"✅ Campos OK: {campos_ok}/8  —  {avance}\n\n"
                f"Toca el campo que quieres llenar o editar:"
            )
            await u.message.reply_text(historial, reply_markup=_ikb_datos_campos(d))
            return MENU
        else:
            await u.message.reply_text(
                "📋 DATOS DEL REPORTE\n\nSelecciona la ZONA de la cuadrilla:",
                reply_markup=_ikb_zonas("cz"))
            return CUADRILLA_ZONA

    elif opcion == "📸 Fotos del reporte":
        d = c.user_data
        n_a = len(d.get("fotos_antes",  [])); n_d = len(d.get("fotos_durante",[])); n_f = len(d.get("fotos_fin",[]))
        total = n_a + n_d + n_f
        await u.message.reply_text(
            f"📸 FOTOS DEL REPORTE\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 Estado actual:\n"
            f"  • Antes:   {n_a}/8 📸  {'✅' if n_a>0 else '⬜'}\n"
            f"  • Durante: {n_d}/8 📸  {'✅' if n_d>0 else '⬜'}\n"
            f"  • Al final:{n_f}/8 📸  {'✅' if n_f>0 else '⬜'}\n"
            f"  Total: {total} fotos\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"Selecciona qué sección llenar:",
            reply_markup=_ikb_fotos_sel(d))
        return MENU

    elif opcion == "📡 Trazas OTDR":
        d = c.user_data
        n_a = len(d.get("trazas_antes",[])); n_d = len(d.get("trazas_despues",[]))
        total = n_a + n_d
        _otdr_path = _get_otdr_path()
        _caption = (
            f"📡 TRAZAS OTDR\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 Estado actual:\n"
            f"  • Antes:   {n_a}/8 📡  {'✅' if n_a>0 else '⬜'}\n"
            f"  • Después: {n_d}/8 📡  {'✅' if n_d>0 else '⬜'}\n"
            f"  Total: {total} trazas\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"Selecciona qué sección llenar:"
        )
        if _otdr_path:
            with open(_otdr_path, "rb") as _img:
                await u.message.reply_photo(photo=_img, caption=_caption, reply_markup=_ikb_trazas_sel(d))
        else:
            await u.message.reply_text(_caption, reply_markup=_ikb_trazas_sel(d))
        return MENU

    elif opcion == "📝 Resumen descriptivo":
        lista = c.user_data.get("resumen_lista", [])
        n = len(lista)
        resumen_preview = ""
        if lista:
            resumen_preview = "\n📋 Líneas guardadas:\n" + "\n".join(
                f"  {i+1}. {l[:60]}{'…' if len(l)>60 else ''}" for i,l in enumerate(lista))
        await u.message.reply_text(
            f"📝 RESUMEN DESCRIPTIVO\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 Total: {n}/20 líneas{resumen_preview}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"¿Qué deseas hacer?",
            reply_markup=_ikb_resumen_acc(n))
        return MENU

    elif opcion == "🔧 Tabla de mantenimiento":
        nodos = c.user_data.setdefault("mant_nodos", {})
        await u.message.reply_text(
            _texto_mant_panel(nodos),
            reply_markup=_ikb_mant_panel(nodos))
        return MANT_CONFIRM

    elif opcion == "📌 Conclusiones":
        lista = c.user_data.get("conclusiones_lista", [])
        n = len(lista)
        con_preview = ""
        if lista:
            con_preview = "\n📋 Líneas guardadas:\n" + "\n".join(
                f"  {i+1}. {l[:60]}{'…' if len(l)>60 else ''}" for i,l in enumerate(lista))
        await u.message.reply_text(
            f"📌 CONCLUSIONES\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 Total: {n}/20 líneas{con_preview}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"¿Qué deseas hacer?",
            reply_markup=_ikb_concl_acc(n))
        return MENU

    elif opcion == "✅ GENERAR EXCEL":
        d = c.user_data
        n_fa = len(d.get("fotos_antes",[])); n_fd = len(d.get("fotos_durante",[]))
        n_ff = len(d.get("fotos_fin",[]));   n_ta = len(d.get("trazas_antes",[]))
        n_td = len(d.get("trazas_despues",[]))
        await u.message.reply_text(
            "📊 RESUMEN FINAL\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"Caso:       {d.get('codigo','⏳ Sin llenar')}\n"
            f"Cuadrilla:  {d.get('cuadrilla','⏳ Sin llenar')}\n"
            f"Incidencia: {d.get('fecha_inc','')} {d.get('hora_inc','')}\n"
            f"Resolución: {d.get('fecha_res','')} {d.get('hora_res','')}\n"
            f"Personal:   {d.get('personal','')[:50]}\n"
            f"Lugar:      {d.get('lugar','')[:50]}\n"
            f"Fotos:  Antes={n_fa} | Durante={n_fd} | Fin={n_ff}\n"
            f"OTDR:   Antes={n_ta} | Despues={n_td}\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "¿Confirmas y generas el Excel?",
            reply_markup=IKB_CONFIRMAR)
        return CONFIRMAR

    else:
        # Incluye el caso de que presionen 🔙 Menú desde el menú mismo
        return await mostrar_menu(u, c)

# ── Confirmar y generar ────────────────────────────────────────────────────────
async def confirmar(u, c):
    if u.message.text.strip().upper() not in ["SI","S"]:
        await u.message.reply_text("Cancelado. Escribe /start para comenzar.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END
    await u.message.reply_text("Generando reporte Excel...", reply_markup=ReplyKeyboardRemove())
    try:
        # Sincronizar mant_filas desde mant_nodos antes de generar
        nodos = c.user_data.get("mant_nodos", {})
        if nodos:
            c.user_data["mant_filas"] = _mant_nodos_a_filas(nodos)
        ruta = generar_excel(c.user_data)
        d = c.user_data
        n_fa = len(d.get("fotos_antes",[]));   n_fd = len(d.get("fotos_durante",[]))
        n_ff = len(d.get("fotos_fin",[]));     n_ta = len(d.get("trazas_antes",[]))
        n_td = len(d.get("trazas_despues",[])); total = n_fa+n_fd+n_ff+n_ta+n_td

        # Enviar al técnico
        with open(ruta,"rb") as f:
            await u.message.reply_document(
                document=f, filename=os.path.basename(ruta),
                caption=(f"Reporte generado\nCaso: {d.get('codigo','')}\n"
                         f"Fotos: Antes={n_fa} | Durante={n_fd} | Fin={n_ff}\n"
                         f"OTDR:  Antes={n_ta} | Despues={n_td}\n"
                         f"Total imagenes: {total}\n6 pestanas completas.\n"
                         f"Escribe /start para nuevo reporte."))

        # Enviar a los coordinadores
        caption_coord = (
            f"📋 *Nuevo reporte generado*\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"📁 Caso:       {d.get('codigo','—')}\n"
            f"👥 Cuadrilla:  {d.get('cuadrilla','—')}\n"
            f"📍 Zona:       {d.get('cuadrilla_zona','—')}\n"
            f"📅 Incidencia: {d.get('fecha_inc','')} {d.get('hora_inc','')}\n"
            f"📅 Resolución: {d.get('fecha_res','')} {d.get('hora_res','')}\n"
            f"👤 Técnico:    {d.get('auth_email','—')}")
        for coord_id in COORDINADORES_TG_IDS:
            try:
                with open(ruta,"rb") as f:
                    await u.message.bot.send_document(
                        chat_id=coord_id,
                        document=f,
                        filename=os.path.basename(ruta),
                        caption=caption_coord,
                        parse_mode="Markdown")
            except Exception as e_coord:
                logger.warning(f"No se pudo notificar al coordinador {coord_id}: {e_coord}")

    except Exception as e:
        logger.error(f"Error al generar Excel: {e}", exc_info=True)
        await u.message.reply_text(f"Error al generar: {e}")
    return ConversationHandler.END

async def cmd_menu(u, c):
    """Comando /menu — vuelve al menú desde cualquier punto."""
    await u.message.reply_text("↩️ Volviendo al menú...")
    return await mostrar_menu(u, c)

async def p_volver_menu(u, c):
    """Handler del botón 🔙 Menú — vuelve al menú desde cualquier estado."""
    return await mostrar_menu(u, c)

async def cancelar(u, c):
    await u.message.reply_text("Cancelado. Escribe /start.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ══════════════════════════════════════════════════════════════════════════════
# CALLBACK: Confirmar o rechazar texto con errores ortográficos
# ══════════════════════════════════════════════════════════════════════════════
async def cb_spell_confirm(upd, c):
    """El usuario elige Aceptar o Reescribir tras el aviso de ortografía."""
    q = upd.callback_query; await q.answer()
    action = q.data.split(":")[1]
    pending = c.user_data.get("_spell_pending")
    if not pending:
        await q.edit_message_text("⚠️ Sesión expirada. Escribe el texto de nuevo.")
        return MENU

    campo         = pending["campo"]
    valor         = pending["valor"]
    next_state    = pending["next_state"]
    current_state = pending["current_state"]
    c.user_data.pop("_spell_pending", None)

    # ── Reescribir: el usuario vuelve al mismo estado sin guardar ─────────────
    if action == "retry":
        await q.edit_message_text("✏️ Por favor, escribe el texto de nuevo correctamente:")
        return current_state

    # ── Confirmar corrección (sc:fix): aplicar y guardar ─────────────────────
    if action == "fix":
        errores_guardados = pending.get("errores", [])
        valor = _aplicar_correcciones(valor, errores_guardados)
        await q.edit_message_text(f"✅ Texto corregido guardado:\n`{valor}`", parse_mode="Markdown")

    # ── Aceptar: guardar y continuar ──────────────────────────────────────────

    # Casos especiales: líneas de resumen / conclusiones
    if campo == "_resumen_line":
        lista = c.user_data.get("resumen_lista", [])
        lista.append(valor)
        c.user_data["resumen_lista"] = lista
        n = len(lista)
        await q.edit_message_text(f"✅ Línea {n} guardada.\n📊 Total: {n}/20\n\nEnvía otra línea o toca Listo:")
        await c.bot.send_message(q.message.chat_id, "Continúa:", reply_markup=_ikb_listo("res"))
        return RESUMEN

    if campo == "_conclusiones_line":
        lista = c.user_data.get("conclusiones_lista", [])
        lista.append(valor)
        c.user_data["conclusiones_lista"] = lista
        n = len(lista)
        await q.edit_message_text(f"✅ Línea {n} guardada.\n📊 Total: {n}/20\n\nEnvía otra línea o toca Listo:")
        await c.bot.send_message(q.message.chat_id, "Continúa:", reply_markup=_ikb_listo("con"))
        return CONCLUSIONES

    # Caso general: guardar campo
    c.user_data[campo] = valor

    # Modo edición: volver al historial
    if c.user_data.get("_editing_campo") == campo:
        await q.edit_message_text("✅ Texto guardado.")
        class _FMsc:
            def __init__(self, chat_id, bot): self.chat_id=chat_id; self.bot=bot
            async def reply_text(self, txt, **kw): await self.bot.send_message(self.chat_id, txt, **kw)
        return await _volver_historial(_FMsc(q.message.chat_id, q.bot).reply_text, c.user_data, c)

    # ── Post-guardado según campo ─────────────────────────────────────────────
    if pending.get("show_historial_obs"):
        # obs_manual → mostrar historial de datos
        await q.edit_message_text("✅ Observación aceptada.")
        d = c.user_data
        campos_ok = sum(bool(d.get(k)) for k in
            ["codigo","cuadrilla","fecha_inc","fecha_res","personal","lugar","motivo","obs"])
        campos_vacios = 8 - campos_ok
        avance = "✅ ¡Todo completo!" if campos_vacios == 0 else f"⏳ Faltan {campos_vacios} campo(s)"
        await c.bot.send_message(
            q.message.chat_id,
            f"✅ Observacion registrada.\n\n"
            f"📋 DATOS DEL REPORTE — HISTORIAL\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📁 Caso:        {d.get('codigo','—')}\n"
            f"👥 Cuadrilla:   {d.get('cuadrilla','—')}\n"
            f"📍 Lugar:       {str(d.get('lugar','—'))[:40]}\n"
            f"📅 Incidencia:  {d.get('fecha_inc','—')} {d.get('hora_inc','—')}\n"
            f"📅 Arribo:      {d.get('fecha_arr','—')} {d.get('hora_arr','—')}\n"
            f"📅 Reparación:  {d.get('fecha_rep','—')} {d.get('hora_rep','—')}\n"
            f"📅 Resolución:  {d.get('fecha_res','—')} {d.get('hora_res','—')}\n"
            f"⚡ Motivo:      {str(d.get('motivo','—'))[:40]}\n"
            f"🔗 Enlace:      {str(d.get('enlaces','—'))[:40]}\n"
            f"👤 Personal:    {str(d.get('personal','—'))[:40]}\n"
            f"📝 Observación: {str(d.get('obs','—'))[:50]}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"✅ Campos OK: {campos_ok}/8  —  {avance}\n\n"
            f"Toca el campo que quieres llenar o editar:",
            reply_markup=_ikb_datos_campos(d))
        return MENU

    if pending.get("show_mant_done"):
        # mant_obs → volver al menú principal
        await q.edit_message_text("✅ Tabla de mantenimiento completada.")
        return MENU

    # Flujo estándar con next_msg y next_markup opcionales
    next_msg    = pending.get("next_msg", "✅ Guardado.")
    next_markup = pending.get("next_markup")
    await q.edit_message_text("✅ Texto aceptado.")
    if next_markup:
        await c.bot.send_message(q.message.chat_id, next_msg, reply_markup=next_markup)
    else:
        await c.bot.send_message(q.message.chat_id, next_msg, reply_markup=KB_MENU)
    return next_state


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    threading.Thread(target=iniciar_servidor, daemon=True).start()
    persistence = PicklePersistence(filepath="/tmp/bot_data.pkl")
    app = Application.builder().token(BOT_TOKEN).persistence(persistence).build()
    TXT_O_FOTO = (filters.TEXT & ~filters.COMMAND) | filters.PHOTO
    CQ = CallbackQueryHandler  # alias corto
    CT  = CQ(cb_tabs,       pattern="^tab:")   # Tab shortcut
    CDA = CQ(cb_datos_campo, pattern="^dc:")   # Datos campo selector
    CFS = CQ(cb_fotos_sec,  pattern="^fsec:")  # Fotos sub-sección
    CTS = CQ(cb_trazas_sec, pattern="^tsec:")  # Trazas sub-sección
    CMP = CQ(cb_mant_panel,    pattern="^mp:")  # Mant Panel principal
    CMN = CQ(cb_mant_nodo,     pattern="^mn:")  # Mant Panel nodo
    CMH = CQ(cb_mant_hilo,     pattern="^mh:")  # Mant Panel hilo campos
    CRA = CQ(cb_racc,       pattern="^racc:")  # Resumen acciones
    CCA = CQ(cb_cacc,       pattern="^cacc:")  # Conclusiones acciones
    CSS = CQ(cb_spell_confirm, pattern="^sc:") # Spell-check Confirmar/Reescribir

    # ── ATAJO GLOBAL: cualquier botón del menú funciona desde cualquier estado ──
    _MENU_RE = (
        "^(📋 Datos del reporte|📸 Fotos del reporte|📡 Trazas OTDR"
        "|📝 Resumen descriptivo|🔧 Tabla de mantenimiento"
        "|📌 Conclusiones|✅ GENERAR EXCEL)$"
    )
    MH = lambda h: MessageHandler(filters.TEXT & ~filters.COMMAND, h)
    MS = MessageHandler(filters.Regex(_MENU_RE), p_menu)  # Menu Shortcut

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            AUTH:          [MH(p_auth)],
            MENU:          [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MH(p_menu)],
            CODIGO:        [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_codigo)],
            CUADRILLA_ZONA:[CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_cuadrilla_zona, pattern="^cz:"), MS,
                            MH(p_cuadrilla_zona)],
            CUADRILLA_SEL: [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_cuadrilla_sel,  pattern="^cs:"), MS,
                            MH(p_cuadrilla_sel)],
            FECHA_INC:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_fecha_inc)],
            HORA_INC:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_hora_inc)],
            FECHA_ARR:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_fecha_arr)],
            HORA_ARR:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_hora_arr)],
            FECHA_REP:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_fecha_rep)],
            HORA_REP:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_hora_rep)],
            FECHA_RES:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_fecha_res)],
            HORA_RES:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_hora_res)],
            MOTIVO_ZONA:   [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_motivo_zona,    pattern="^mz:"), MS,
                            MH(p_motivo_zona)],
            MOTIVO_TIPO:   [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_motivo_tipo,    pattern="^mt:"), MS,
                            MH(p_motivo_tipo)],
            MOTIVO_TRAMO:  [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_motivo_tramo,   pattern="^tr:"), MS,
                            MH(p_motivo_tramo)],
            MOTIVO:        [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, MS, MH(p_motivo)],
            TESTIGOS:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, CQ(cb_causas,         pattern="^ca:"), MS,
                            MH(p_testigos)],
            REMEDIO:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_remedio,        pattern="^rem:"), MS,
                            MH(p_remedio)],
            REMEDIO_METROS:[CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_remedio_metros)],
            REMEDIO_MANUAL:[CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, MS, MH(p_remedio_manual)],
            ENLACES:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_enlaces)],
            ENLACES_ZONA:  [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_enlaces_zona,   pattern="^ez:"), MS,
                            MH(p_enlaces_zona)],
            ENLACES_TRAMO: [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_enlaces_tramo,  pattern="^et:"), MS,
                            MH(p_enlaces_tramo)],
            ENLACES_MAS:   [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_enlaces_mas,    pattern="^em:"), MS,
                            MH(p_enlaces_mas)],
            ENLACES_MANUAL:[CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_enlaces_manual)],
            PERSONAL:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_personal)],
            LUGAR:         [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_lugar)],
            COORDS:        [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_listo_seccion,  pattern="^lo:coord:"), MS,
                            MessageHandler((filters.TEXT & ~filters.COMMAND) | filters.LOCATION, p_coords)],
            FOTOS_ANTES:   [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_listo_seccion,  pattern="^lo:fa:"), MS,
                            MessageHandler(TXT_O_FOTO, p_fotos_antes)],
            FOTOS_DURANTE: [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_listo_seccion,  pattern="^lo:fd:"), MS,
                            MessageHandler(TXT_O_FOTO, p_fotos_durante)],
            FOTOS_FIN:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_listo_seccion,  pattern="^lo:ff:"), MS,
                            MessageHandler(TXT_O_FOTO, p_fotos_fin)],
            TRAZAS_ANTES:  [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_listo_seccion,  pattern="^lo:ta:"), MS,
                            MessageHandler(TXT_O_FOTO, p_trazas_antes)],
            TRAZAS_DESPUES:[CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_listo_seccion,  pattern="^lo:td:"), MS,
                            MessageHandler(TXT_O_FOTO, p_trazas_despues)],
            OBS:           [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_causas,         pattern="^ca:"), MS,
                            MH(p_obs)],
            OBS_MANUAL:    [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, MS, MH(p_obs_manual)],
            RESUMEN:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, CQ(cb_listo_seccion,  pattern="^lo:res:"), MS,
                            MH(p_resumen)],
            MANT_CONFIRM:  [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_mant_confirm,   pattern="^mc:"), MS,
                            MH(p_mant_confirm)],
            MANT_EST:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_mant_est)],
            MANT_HILO:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_mant_hilo)],
            MANT_DA:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_mant_da)],
            MANT_PA:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_mant_pa)],
            MANT_DD:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_mant_dd)],
            MANT_PD:       [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, MS, MH(p_mant_pd)],
            MANT_MORE_HILO:[CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_mant_more_hilo, pattern="^mh:"), MS,
                            MH(p_mant_more_hilo)],
            MANT_MORE_EST: [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_mant_more_est,  pattern="^me:"), MS,
                            MH(p_mant_more_est)],
            MANT_OBS:      [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, MS, MH(p_mant_obs)],
            CONCLUSIONES:  [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CSS, CQ(cb_listo_seccion,  pattern="^lo:con:"), MS,
                            MH(p_conclusiones)],
            CONFIRMAR:     [CT, CDA, CFS, CTS, CMP, CMN, CMH, CRA, CCA, CQ(cb_confirmar,      pattern="^cf:"), MS,
                            MH(confirmar)],
        },
        fallbacks=[
            CommandHandler("cancelar", cancelar),
            CommandHandler("menu", cmd_menu),
            MessageHandler(filters.Regex("^🔙 Menú$"), p_volver_menu),
            CQ(cb_datos_campo, pattern="^dc:"),
            CQ(cb_tabs,        pattern="^tab:"),
            CQ(cb_mant_panel,  pattern="^mp:"),
            CQ(cb_mant_nodo,   pattern="^mn:"),
            CQ(cb_mant_hilo,   pattern="^mh:"),
        ],
        allow_reentry=True
    )
    app.add_handler(conv)
    # ── Red de seguridad: dc: funciona aunque el ConversationHandler no tenga estado ──
    app.add_handler(CQ(cb_datos_campo, pattern="^dc:"), group=1)
    logger.info("Bot iniciado — Autenticacion TOTP + Telconet activa")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)

if __name__ == "__main__":
    main()
